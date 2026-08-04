"""
base/cbv/roster.py

Class-based views for the Shift Roster Planner.
Provides roster grid, cell editing, publish, my roster, and import/export.
"""

import json
from datetime import date, timedelta

import openpyxl
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import TemplateView
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from base.filters import RosterFilter
from base.forms import RosterCellUpdateForm
from base.models import CompanyLeaves, EmployeeShift, Holidays, Roster, RosterPublishLog
from employee.models import Employee
from horilla_views.cbv_methods import login_required, paginator_qry
from horilla_views.generic.cbv.views import HorillaCardView, HorillaNavView

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_date(value, fallback):
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError):
        return fallback


def _week_start():
    today = date.today()
    return today - timedelta(days=today.weekday())


# ---------------------------------------------------------------------------
# Roster Home
# ---------------------------------------------------------------------------


@method_decorator(login_required, name="dispatch")
class RosterHomeView(TemplateView):
    template_name = "base/roster/roster_home.html"


# ---------------------------------------------------------------------------
# Roster Nav
# ---------------------------------------------------------------------------


@method_decorator(login_required, name="dispatch")
class RosterNavView(HorillaNavView):
    nav_title = _("Roster Planner")
    template_name = "generic/inline_nav.html"
    search_url = reverse_lazy("roster-grid")
    search_swap_target = "#rosterGridContainer"
    filter_body_template = "base/roster/roster_filter.html"
    apply_first_filter = True

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.actions = [
            {
                "action": _("Import Roster"),
                "attrs": f"""
                    data-toggle="oh-modal-toggle"
                    data-target="#genericModal"
                    hx-target="#genericModalBody"
                    hx-get="{reverse('roster-import-form')}"
                """,
            },
            {
                "action": _("Publish Roster"),
                "attrs": f"""
                    data-toggle="oh-modal-toggle"
                    data-target="#genericModal"
                    hx-target="#genericModalBody"
                    hx-get="{reverse('roster-publish-form')}"
                """,
            },
        ]

    def get_context_data(self, **kwargs):
        from base.models import Department

        context = super().get_context_data(**kwargs)
        context["departments"] = Department.objects.all()
        today = _week_start()
        context["default_from_date"] = self.request.GET.get(
            "from_date", today.isoformat()
        )
        context["default_to_date"] = self.request.GET.get(
            "to_date", (today + timedelta(days=6)).isoformat()
        )
        return context


# ---------------------------------------------------------------------------
# Roster Grid
# ---------------------------------------------------------------------------


@method_decorator(login_required, name="dispatch")
class RosterGridView(HorillaCardView):
    model = Roster
    filter_class = RosterFilter
    template_name = "base/roster/roster_grid.html"
    search_url = reverse_lazy("roster-grid")

    def get_queryset(self):
        queryset = super().get_queryset()
        today = _week_start()
        if not self.request.GET.get("from_date"):
            queryset = queryset.filter(
                date__gte=today, date__lte=today + timedelta(days=6)
            )
        return queryset

    def get_context_data(self, **kwargs):
        from employee.models import Employee

        context = super().get_context_data(**kwargs)

        today = _week_start()
        from_date = _parse_date(self.request.GET.get("from_date"), today)
        to_date = _parse_date(
            self.request.GET.get("to_date"), from_date + timedelta(days=6)
        )
        date_range = [
            from_date + timedelta(days=i) for i in range((to_date - from_date).days + 1)
        ]

        emp_ids = list(self.queryset.values_list("employee_id", flat=True).distinct())
        employees_qs = (
            Employee.objects.filter(pk__in=emp_ids)
            .select_related("employee_work_info__department_id")
            .order_by("employee_first_name", "employee_last_name")
        )
        paginated = paginator_qry(
            employees_qs, self.request.GET.get("page"), self.records_per_page
        )

        page_emp_ids = [emp.pk for emp in paginated]
        roster_entries = self.queryset.filter(
            employee_id__in=page_emp_ids
        ).select_related(
            "shift", "employee", "employee__employee_work_info__department_id"
        )

        roster_map = {}
        for entry in roster_entries:
            roster_map[(entry.employee_id, entry.date)] = entry

        rows = [
            {
                "employee": emp,
                "cells": list(
                    zip(date_range, [roster_map.get((emp.pk, d)) for d in date_range])
                ),
            }
            for emp in paginated
        ]

        get_params = self.request.GET.copy()
        get_params.pop("page", None)

        context.update(
            {
                "date_range": date_range,
                "rows": rows,
                "from_date": from_date,
                "to_date": to_date,
                "paginated_employees": paginated,
                "roster_filter_params": get_params.urlencode(),
                "view_id": "rosterGrid",
                "selected_instances_key_id": "rosterEmployeeInstances",
                "select_all_ids": json.dumps(emp_ids),
                "queryset": paginated,
                "bulk_select_option": True,
            }
        )
        return context


# ---------------------------------------------------------------------------
# Roster Cell Update (inline HTMX)
# ---------------------------------------------------------------------------


@method_decorator(login_required, name="dispatch")
class RosterCellUpdateView(View):
    template_name = "base/roster/roster_cell_form.html"

    def _get_entry(self, employee_id, cell_date):
        return Roster.objects.filter(employee_id=employee_id, date=cell_date).first()

    def get(self, request, *args, **kwargs):
        from employee.models import Employee

        employee_id = request.GET.get("employee")
        cell_date = _parse_date(request.GET.get("date"), None)
        dept_id = request.GET.get("department")

        if not employee_id or not cell_date:
            return HttpResponse(status=400)

        employee = get_object_or_404(Employee, pk=employee_id)
        entry = self._get_entry(employee_id, cell_date)
        form = RosterCellUpdateForm(instance=entry)

        return render(
            request,
            self.template_name,
            {
                "form": form,
                "employee": employee,
                "cell_date": cell_date,
                "dept_id": dept_id,
                "entry": entry,
            },
        )

    def post(self, request, *args, **kwargs):
        from employee.models import Employee

        employee_id = request.POST.get("employee")
        cell_date = _parse_date(request.POST.get("date"), None)
        dept_id = request.POST.get("department")

        if not employee_id or not cell_date:
            return HttpResponse(status=400)

        employee = get_object_or_404(Employee, pk=employee_id)
        department = employee.get_department()
        entry = self._get_entry(employee_id, cell_date)
        form = RosterCellUpdateForm(request.POST, instance=entry)

        if form.is_valid():
            roster = form.save(commit=False)
            roster.employee = employee
            roster.date = cell_date
            if department:
                roster.department = department
            if not roster.created_by_id:
                roster.created_by = getattr(request.user, "employee_get", None)
            roster.save()
            return render(
                request,
                "base/roster/roster_cell.html",
                {
                    "entry": roster,
                    "cell_date": cell_date,
                    "employee": employee,
                    "dept_id": dept_id,
                },
            )

        return render(
            request,
            self.template_name,
            {
                "form": form,
                "employee": employee,
                "cell_date": cell_date,
                "dept_id": dept_id,
                "entry": entry,
            },
        )


# ---------------------------------------------------------------------------
# Roster Publish Form (GET) + Publish Action (POST)
# ---------------------------------------------------------------------------


@method_decorator(login_required, name="dispatch")
class RosterPublishFormView(View):
    template_name = "base/roster/publish_form.html"

    def get(self, request, *args, **kwargs):
        from base.models import Department

        return render(
            request,
            self.template_name,
            {
                "departments": Department.objects.all(),
            },
        )


@method_decorator(login_required, name="dispatch")
class RosterPublishView(View):

    def post(self, request, *args, **kwargs):
        from base.models import Department
        from notifications.signals import notify

        dept_id = request.POST.get("department") or None
        from_date = _parse_date(request.POST.get("from_date"), None)
        to_date = _parse_date(request.POST.get("to_date"), None)

        if not from_date:
            return JsonResponse({"error": "From date is required."}, status=400)

        if dept_id:
            departments = Department.objects.filter(pk=dept_id)
        else:
            departments = Department.objects.all()

        publisher = getattr(request.user, "employee_get", None)
        all_employee_ids = set()
        for department in departments:
            qs = Roster.objects.filter(
                department=department,
                date__gte=from_date,
                is_published=False,
            )
            if to_date:
                qs = qs.filter(date__lte=to_date)

            emp_ids = set(qs.values_list("employee_id", flat=True))
            all_employee_ids.update(emp_ids)
            qs.update(is_published=True)

            RosterPublishLog.objects.create(
                department=department,
                from_date=from_date,
                to_date=to_date or from_date,
                published_by=publisher,
                total_employees=len(emp_ids),
            )

        my_roster_url = reverse("my-roster")
        date_label = f"{from_date}" if not to_date else f"{from_date} - {to_date}"
        seen = set()
        notify_qs = Roster.objects.filter(
            employee_id__in=all_employee_ids,
            date__gte=from_date,
            is_published=True,
        ).select_related("employee__employee_user_id")
        if to_date:
            notify_qs = notify_qs.filter(date__lte=to_date)
        for entry in notify_qs:
            if entry.employee_id in seen:
                continue
            seen.add(entry.employee_id)
            try:
                notify.send(
                    publisher,
                    recipient=entry.employee.employee_user_id,
                    verb=f"Your roster from {date_label} has been published.",
                    redirect=my_roster_url,
                    icon="calendar-outline",
                )
            except Exception:
                pass

        dept_label = departments.first() if dept_id else _("All Departments")
        messages.success(
            request,
            _("Roster published for %(dept)s (%(range)s).")
            % {"dept": dept_label, "range": date_label},
        )
        grid_params = f"from_date={from_date}"
        if to_date:
            grid_params += f"&to_date={to_date}"
        if dept_id:
            grid_params += f"&department={dept_id}"
        return HttpResponse(
            "<script>$('#genericModal').removeClass('oh-modal--show');"
            f"htmx.ajax('GET', '{reverse('roster-grid')}?{grid_params}', "
            "{'target':'#rosterGridContainer','swap':'innerHTML'});"
            "$('#reloadMessagesButton').click();</script>"
        )


# ---------------------------------------------------------------------------
# Roster Bulk Publish by Employee
# ---------------------------------------------------------------------------


@method_decorator(login_required, name="dispatch")
class RosterEmployeeBulkPublishView(View):

    def post(self, request, *args, **kwargs):
        from notifications.signals import notify

        employee_ids = request.POST.getlist("employee_ids")
        from_date = _parse_date(request.POST.get("from_date"), None)
        to_date = _parse_date(request.POST.get("to_date"), None)

        if not employee_ids or not from_date:
            return HttpResponse(status=400)

        publisher = getattr(request.user, "employee_get", None)

        qs = Roster.objects.filter(
            employee_id__in=employee_ids,
            date__gte=from_date,
            is_published=False,
        )
        if to_date:
            qs = qs.filter(date__lte=to_date)

        count = qs.values("employee_id").distinct().count()
        qs.update(is_published=True)

        my_roster_url = reverse("my-roster")
        seen = set()
        notify_qs = Roster.objects.filter(
            employee_id__in=employee_ids,
            date__gte=from_date,
            is_published=True,
        ).select_related("employee__employee_user_id")
        if to_date:
            notify_qs = notify_qs.filter(date__lte=to_date)
        for entry in notify_qs:
            if entry.employee_id in seen:
                continue
            seen.add(entry.employee_id)
            try:
                notify.send(
                    publisher,
                    recipient=entry.employee.employee_user_id,
                    verb="Your roster has been published.",
                    redirect=my_roster_url,
                    icon="calendar-outline",
                )
            except Exception:
                pass

        messages.success(
            request,
            _("Roster published for %(count)s employee(s).") % {"count": count},
        )
        grid_params = f"from_date={from_date}"
        if to_date:
            grid_params += f"&to_date={to_date}"
        dept_id = request.POST.get("department")
        if dept_id:
            grid_params += f"&department={dept_id}"
        return HttpResponse(
            "<script>"
            f"htmx.ajax('GET', '{reverse('roster-grid')}?{grid_params}', "
            "{'target':'#rosterGridContainer','swap':'innerHTML'});"
            "$('#reloadMessagesButton').click();</script>"
        )


# ---------------------------------------------------------------------------
# My Roster (employee self-service)
# ---------------------------------------------------------------------------


@method_decorator(login_required, name="dispatch")
class MyRosterView(View):
    template_name = "base/roster/my_roster.html"

    def get(self, request, *args, **kwargs):
        employee = getattr(request.user, "employee_get", None)
        today = date.today()
        end = today + timedelta(days=13)

        entries = []
        if employee:
            entries = (
                Roster.objects.filter(
                    employee=employee,
                    date__range=(today, end),
                    is_published=True,
                )
                .select_related("shift", "department")
                .order_by("date")
            )

        return render(
            request,
            self.template_name,
            {
                "entries": entries,
                "today": today,
                "end_date": end,
            },
        )


# ---------------------------------------------------------------------------
# Roster Import
# ---------------------------------------------------------------------------


@method_decorator(login_required, name="dispatch")
class RosterImportFormView(View):
    template_name = "base/roster/roster_import_form.html"

    def get(self, request, *args, **kwargs):
        today = _week_start()
        return render(
            request,
            self.template_name,
            {
                "default_start": today.isoformat(),
                "default_end": (today + timedelta(days=6)).isoformat(),
            },
        )


@method_decorator(login_required, name="dispatch")
class RosterTemplateDownloadView(View):

    def get(self, request, *args, **kwargs):

        today = _week_start()
        start_date = _parse_date(request.GET.get("start_date"), today)
        end_date = _parse_date(request.GET.get("end_date"), today + timedelta(days=6))
        if (end_date - start_date).days > 27:
            end_date = start_date + timedelta(days=27)
        date_range = [
            start_date + timedelta(days=i)
            for i in range((end_date - start_date).days + 1)
        ]

        off_dates = {}

        holidays = Holidays.objects.filter(
            start_date__lte=end_date, end_date__gte=start_date
        )
        for h in holidays.filter(is_specific=False):
            h_end = h.end_date or h.start_date
            for d in date_range:
                if h.start_date <= d <= h_end:
                    off_dates.setdefault(d, h.name)

        # Pre-build {employee_pk: {date: holiday_name}} for specific holidays — avoids per-row DB queries
        specific_employee_off = {}
        for h in holidays.filter(is_specific=True).prefetch_related("employees"):
            h_end = h.end_date or h.start_date
            applicable_dates = [
                d
                for d in date_range
                if (
                    h.recurring
                    and h.start_date.month == d.month
                    and h.start_date.day == d.day
                )
                or (not h.recurring and h.start_date <= d <= h_end)
            ]
            for emp in h.employees.all():
                emp_off = specific_employee_off.setdefault(emp.pk, {})
                for d in applicable_dates:
                    emp_off.setdefault(d, h.name)

        company_leaves = list(CompanyLeaves.objects.all())
        for d in date_range:
            if d in off_dates:
                continue
            first_day = d.replace(day=1)
            week_no = str((d.day + first_day.weekday() - 1) // 7)
            week_day = str(d.weekday())
            for cl in company_leaves:
                if cl.based_on_week_day == week_day and (
                    not cl.based_on_week or cl.based_on_week == week_no
                ):
                    off_dates[d] = "Weekly Off Day"
                    break

        selected_company = request.session.get("selected_company")
        emp_qs = (
            Employee.objects.filter(is_active=True, employee_work_info__isnull=False)
            .select_related(
                "employee_work_info__department_id", "employee_work_info__company_id"
            )
            .order_by("employee_first_name", "employee_last_name")
        )
        if selected_company and selected_company != "all":
            emp_qs = emp_qs.filter(employee_work_info__company_id=selected_company)

        shifts = list(EmployeeShift.objects.values_list("employee_shift", flat=True))
        shift_options = ["OFF"] + list(shifts)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roster"

        header_fill = PatternFill("solid", fgColor="5C6BC0")
        weekend_fill = PatternFill("solid", fgColor="ECEEF6")
        off_fill = PatternFill("solid", fgColor="E0E0E0")
        meta_fill = PatternFill("solid", fgColor="F0F2F8")
        even_fill = PatternFill("solid", fgColor="FAFBFF")
        white_font = Font(bold=True, color="FFFFFF", size=10)
        meta_font = Font(bold=True, color="4A5173", size=10)
        off_font = Font(bold=False, color="000000", size=10)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        for col, label in enumerate(["Employee Name", "Employee ID"], start=1):
            c = ws.cell(1, col, label)
            c.fill = meta_fill
            c.font = meta_font
            c.alignment = left if col == 1 else center

        for col_i, d in enumerate(date_range, start=3):
            c = ws.cell(1, col_i)
            c.value = d
            c.number_format = "ddd dd/mm"
            c.alignment = center
            off_reason = off_dates.get(d)
            if off_reason:
                c.fill = off_fill
                c.font = Font(bold=True, color="757575", size=10)
            elif d.weekday() >= 5:
                c.fill = weekend_fill
                c.font = Font(bold=True, color="8B91B0", size=10)
            else:
                c.fill = header_fill
                c.font = white_font
            ws.column_dimensions[get_column_letter(col_i)].width = 13

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 13
        ws.row_dimensions[1].height = 26

        dv_formula = '"' + ",".join(opt[:31] for opt in shift_options)[:255] + '"'
        dv = DataValidation(
            type="list", formula1=dv_formula, allow_blank=True, showErrorMessage=False
        )
        last_row = max(emp_qs.count() + 2, 100)
        dv.sqref = f"C2:{get_column_letter(len(date_range) + 2)}{last_row}"
        ws.add_data_validation(dv)

        for row_i, emp in enumerate(emp_qs, start=2):
            ws.cell(row_i, 1, emp.get_full_name()).alignment = left
            id_cell = ws.cell(row_i, 2, emp.pk)
            id_cell.alignment = center
            id_cell.font = Font(size=10, color="777777")
            if row_i % 2 == 0:
                ws.cell(row_i, 1).fill = even_fill
                ws.cell(row_i, 2).fill = even_fill
            ws.row_dimensions[row_i].height = 20
            for col_i, d in enumerate(date_range, start=3):
                c = ws.cell(row_i, col_i)
                c.alignment = center
                off_reason = off_dates.get(d) or specific_employee_off.get(
                    emp.pk, {}
                ).get(d)
                if off_reason:
                    c.value = "OFF"
                    c.fill = off_fill
                    c.font = off_font
                    if off_reason != "Weekly Off Day":
                        c.comment = Comment(f"Holiday: {off_reason}", "Horilla")
                elif d.weekday() >= 5:
                    c.fill = weekend_fill

        ws2 = wb.create_sheet("Available Shifts")
        ws2.cell(1, 1, "Shift Name").font = Font(bold=True, color="4A5173", size=10)
        ws2.cell(2, 1, "OFF").font = Font(bold=True, color="E65100", size=10)
        for i, s in enumerate(shifts, start=3):
            ws2.cell(i, 1, s).font = Font(size=10)
        ws2.column_dimensions["A"].width = 32

        ws.freeze_panes = "C2"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = (
            f"roster_template_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


@method_decorator(login_required, name="dispatch")
class RosterImportView(View):

    def post(self, request, *args, **kwargs):
        from datetime import datetime as dt

        import openpyxl

        from base.models import EmployeeShift
        from employee.models import Employee

        uploaded = request.FILES.get("file")
        if not uploaded:
            return self._form_error(request, _("No file uploaded."))
        if not uploaded.name.endswith(".xlsx"):
            return self._form_error(request, _("Please upload a valid .xlsx file."))

        try:
            wb = openpyxl.load_workbook(uploaded, data_only=True)
        except Exception:
            return self._form_error(
                request,
                _("Could not read the file. Please upload the unmodified template."),
            )

        ws = wb.active
        dates = []
        for col in range(3, ws.max_column + 1):
            val = ws.cell(1, col).value
            if val is None:
                break
            if isinstance(val, dt):
                dates.append(val.date())
            elif isinstance(val, date):
                dates.append(val)
            else:
                dates.append(None)

        if not any(dates):
            return self._form_error(
                request,
                _("Could not read date headers. Please use the downloaded template."),
            )

        shift_map = {s.employee_shift.lower(): s for s in EmployeeShift.objects.all()}
        publisher = getattr(request.user, "employee_get", None)
        created = updated = skipped = 0
        errors = []

        for row in range(2, ws.max_row + 1):
            emp_id_val = ws.cell(row, 2).value
            if emp_id_val is None:
                continue
            try:
                emp = Employee.objects.select_related(
                    "employee_work_info__department_id",
                ).get(pk=int(emp_id_val))
            except (Employee.DoesNotExist, ValueError, TypeError):
                errors.append(f"Row {row}: Employee ID {emp_id_val!r} not found.")
                continue

            work_info = getattr(emp, "employee_work_info", None)
            dept = getattr(work_info, "department_id", None)

            if not dept:
                errors.append(
                    f"Row {row}: {emp.get_full_name()} has no department — skipped."
                )
                continue

            for col_offset, roster_date in enumerate(dates):
                if not roster_date:
                    skipped += 1
                    continue
                raw = ws.cell(row, col_offset + 3).value
                if raw is None or str(raw).strip() == "":
                    skipped += 1
                    continue

                value = str(raw).strip()
                is_off = value.lower() == "off"
                shift = None

                if not is_off:
                    shift = shift_map.get(value.lower())
                    if not shift:
                        errors.append(
                            f"Row {row}, {roster_date}: Unknown shift '{value}'."
                        )
                        skipped += 1
                        continue

                entry, was_created = Roster.objects.get_or_create(
                    employee=emp,
                    date=roster_date,
                    defaults={
                        "shift": shift,
                        "department": dept,
                        "is_off": is_off,
                        "is_published": False,
                        "created_by": publisher,
                    },
                )
                if was_created:
                    created += 1
                else:
                    entry.shift = shift
                    entry.is_off = is_off
                    entry.save(update_fields=["shift", "is_off"])
                    updated += 1

        return render(
            request,
            "base/roster/roster_import_result.html",
            {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            },
        )

    def _form_error(self, request, message):
        return render(
            request,
            "base/roster/roster_import_form.html",
            {
                "default_week": _week_start().isoformat(),
                "error": message,
            },
        )
