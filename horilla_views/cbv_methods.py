"""
horilla/cbv_methods.py
"""

import json
import types
import uuid
from io import BytesIO
from typing import Any
from urllib.parse import urlencode
from venv import logger

from django import forms, template
from django.contrib import messages
from django.core.cache import cache as CACHE
from django.core.paginator import Paginator
from django.db import models
from django.db.models.fields.related import ForeignKey
from django.db.models.fields.related_descriptors import (
    ForwardManyToOneDescriptor,
    ReverseOneToOneDescriptor,
)
from django.http import HttpResponse
from django.middleware.csrf import get_token
from django.shortcuts import redirect, render
from django.template import loader
from django.template.defaultfilters import register
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.functional import lazy
from django.utils.html import format_html
from django.utils.safestring import SafeString
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from horilla import settings
from horilla.horilla_middlewares import _thread_locals
from horilla_views.templatetags.generic_template_filters import getattribute

FIELD_WIDGET_MAP = {
    models.CharField: forms.TextInput(attrs={"class": "oh-input w-100"}),
    models.ImageField: forms.FileInput(
        attrs={"type": "file", "class": "oh-input w-100"}
    ),
    models.FileField: forms.FileInput(
        attrs={"type": "file", "class": "oh-input w-100"}
    ),
    models.TextField: forms.Textarea(
        {
            "class": "oh-input w-100",
            "rows": 2,
            "cols": 40,
        }
    ),
    models.IntegerField: forms.NumberInput(attrs={"class": "oh-input w-100"}),
    models.FloatField: forms.NumberInput(attrs={"class": "oh-input w-100"}),
    models.DecimalField: forms.NumberInput(attrs={"class": "oh-input w-100"}),
    models.EmailField: forms.EmailInput(attrs={"class": "oh-input w-100"}),
    models.DateField: forms.DateInput(
        attrs={"type": "date", "class": "oh-input w-100"}
    ),
    models.DateTimeField: forms.DateTimeInput(
        attrs={"type": "date", "class": "oh-input w-100"}
    ),
    models.TimeField: forms.TimeInput(
        attrs={"type": "time", "class": "oh-input w-100"}
    ),
    models.BooleanField: forms.Select({"class": "oh-select oh-select-2 w-100"}),
    models.ForeignKey: forms.Select({"class": "oh-select oh-select-2 w-100"}),
    models.ManyToManyField: forms.SelectMultiple(
        attrs={"class": "oh-select oh-select-2 select2-hidden-accessible"}
    ),
    models.OneToOneField: forms.Select({"class": "oh-select oh-select-2 w-100"}),
}

MODEL_FORM_FIELD_MAP = {
    models.CharField: forms.CharField,
    models.TextField: forms.CharField,  # Textarea can be specified as a widget
    models.IntegerField: forms.IntegerField,
    models.FloatField: forms.FloatField,
    models.DecimalField: forms.DecimalField,
    models.ImageField: forms.FileField,
    models.FileField: forms.FileField,
    models.EmailField: forms.EmailField,
    models.DateField: forms.DateField,
    models.DateTimeField: forms.DateTimeField,
    models.TimeField: forms.TimeField,
    models.BooleanField: forms.BooleanField,
    models.ForeignKey: forms.ModelChoiceField,
    models.ManyToManyField: forms.ModelMultipleChoiceField,
    models.OneToOneField: forms.ModelChoiceField,
}


BOOLEAN_CHOICES = (
    ("", "----------"),
    (True, "Yes"),
    (False, "No"),
)


def decorator_with_arguments(decorator):
    """
    Decorator that allows decorators to accept arguments and keyword arguments.

    Args:
        decorator (function): The decorator function to be wrapped.

    Returns:
        function: The wrapper function.

    """

    def wrapper(*args, **kwargs):
        """
        Wrapper function that captures the arguments and keyword arguments.

        Args:
            *args: Variable length argument list.
            **kwargs: Arbitrary keyword arguments.

        Returns:
            function: The inner wrapper function.

        """

        def inner_wrapper(func):
            """
            Inner wrapper function that applies the decorator to the function.

            Args:
                func (function): The function to be decorated.

            Returns:
                function: The decorated function.

            """
            return decorator(func, *args, **kwargs)

        return inner_wrapper

    return wrapper


def login_required(view_func):
    """
    Decorator to check authenticity of users
    """

    def wrapped_view(self, *args, **kwargs):
        request = getattr(_thread_locals, "request")
        if not getattr(self, "request", None):
            self.request = request
        path = request.path
        res = path.split("/", 2)[1].capitalize().replace("-", " ").upper()
        if res == "PMS":
            res = "Performance"
        request.session["title"] = res
        if path == "" or path == "/":
            request.session["title"] = "Dashboard".upper()
        if not request.user.is_authenticated:
            login_url = reverse("login")
            params = urlencode(request.GET)
            url = f"{login_url}?next={request.path}"
            if params:
                url += f"&{params}"
            return redirect(url)
        try:
            func = view_func(self, request, *args, **kwargs)
        except Exception as e:
            logger.exception(e)
            if not settings.DEBUG:
                return render(request, "went_wrong.html")
            return view_func(self, *args, **kwargs)
        return func

    return wrapped_view


@decorator_with_arguments
def permission_required(function, perm):
    """
    Decorator to validate user permissions
    """

    def _function(self, *args, **kwargs):
        request = getattr(_thread_locals, "request")
        if not getattr(self, "request", None):
            self.request = request

        if request.user.has_perm(perm):
            return function(self, *args, **kwargs)
        else:
            messages.info(request, "You dont have permission.")
            previous_url = request.META.get("HTTP_REFERER", "/")
            key = "HTTP_HX_REQUEST"
            if key in request.META.keys():
                return render(request, "decorator_404.html")
            script = f'<script>window.location.href = "{previous_url}"</script>'
            return HttpResponse(script)

    return _function


@decorator_with_arguments
def check_feature_enabled(function, feature_name, model_class: models.Model):
    """
    Decorator for check feature enabled in singlton model
    """

    def _function(self, request, *args, **kwargs):
        general_setting = model_class.objects.first()
        enabled = getattr(general_setting, feature_name, False)
        if enabled:
            return function(self, request, *args, **kwargs)
        messages.info(request, _("Feature is not enabled on the settings"))
        previous_url = request.META.get("HTTP_REFERER", "/")
        key = "HTTP_HX_REQUEST"
        if key in request.META.keys():
            return render(request, "decorator_404.html")
        script = f'<script>window.location.href = "{previous_url}"</script>'
        return HttpResponse(script)

    return _function


def hx_request_required(function):
    """
    Decorator method that only allow HTMX metod to enter
    """

    def _function(request, *args, **kwargs):
        key = "HTTP_HX_REQUEST"
        if key not in request.META.keys():
            return render(request, "405.html")
        return function(request, *args, **kwargs)

    return _function


def csrf_input(request):
    return format_html(
        '<input type="hidden" name="csrfmiddlewaretoken" value="{}">',
        get_token(request),
    )


@register.simple_tag(takes_context=True)
def csrf_token(context):
    """
    to access csrf token inside the render_template method
    """
    try:
        request = context["request"]
    except:
        request = getattr(_thread_locals, "request")
    csrf_input_lazy = lazy(csrf_input, SafeString, str)
    return csrf_input_lazy(request)


def get_all_context_variables(request) -> dict:
    """
    This method will return dictionary format of context processors
    """
    if getattr(request, "all_context_variables", None) is None:
        all_context_variables = {}
        for processor_path in settings.TEMPLATES[0]["OPTIONS"]["context_processors"]:
            module_path, func_name = processor_path.rsplit(".", 1)
            module = __import__(module_path, fromlist=[func_name])
            func = getattr(module, func_name)
            context = func(request)
            all_context_variables.update(context)
        all_context_variables["csrf_token"] = csrf_token(all_context_variables)
        request.all_context_variables = all_context_variables
    return request.all_context_variables


def render_template(
    path: str,
    context: dict,
    decoding: str = "utf-8",
    status: int = None,
    _using=None,
) -> str:
    """
    This method is used to render HTML text with context.
    """

    request = getattr(_thread_locals, "request", None)
    context.update(get_all_context_variables(request))
    template_loader = loader.get_template(path)
    template_body = template_loader.template.source
    template_bdy = template.Template(template_body)
    context_instance = template.Context(context)
    rendered_content = template_bdy.render(context_instance)
    return HttpResponse(rendered_content, status=status).content.decode(decoding)


def paginator_qry(qryset, page_number, records_per_page=50):
    """
    This method is used to paginate queryset
    """
    if hasattr(qryset, "ordered") and not qryset.ordered:
        qryset = (
            qryset.order_by("-created_at")
            if hasattr(qryset.model, "created_at")
            else qryset.order_by("-id")
        )  # 803

    paginator = Paginator(qryset, records_per_page)
    qryset = paginator.get_page(page_number)
    return qryset


def get_short_uuid(length: int, prefix: str = "hlv"):
    """
    Short uuid generating method
    """
    uuid_str = str(uuid.uuid4().hex)
    return prefix + str(uuid_str[:length]).replace("-", "")


def update_initial_cache(request: object, cache: dict, view: object):

    if cache.get(request.session.session_key + "cbv"):
        cache.get(request.session.session_key + "cbv").update({view: {}})
        return
    cache.set(request.session.session_key + "cbv", {view: {}})
    return


class Reverse:
    reverse: bool = True
    page: str = ""

    def __str__(self) -> str:
        return str(self.reverse)


def getmodelattribute(value: models.Model, attr: str):
    """
    Gets an attribute of a model dynamically, handling related fields.
    """
    result = value
    attrs = attr.split("__")
    for attr in attrs:
        if hasattr(result, attr):
            result = getattr(result, attr)
            if isinstance(result, ForwardManyToOneDescriptor):
                result = result.field.related_model
        elif hasattr(result, "field") and isinstance(result.field, ForeignKey):
            result = getattr(result.field.remote_field.model, attr, None)
        elif hasattr(result, "related") and isinstance(
            result, ReverseOneToOneDescriptor
        ):
            result = getattr(result.related.related_model, attr, None)
    return result


def sortby(
    query_dict, queryset, key: str, page: str = "page", is_first_sort: bool = False
):
    """
    New simplified method to sort the queryset/lists
    """
    request = getattr(_thread_locals, "request", None)
    sort_key = query_dict[key]
    if not CACHE.get(request.session.session_key + "cbvsortby"):
        CACHE.set(request.session.session_key + "cbvsortby", Reverse())
        CACHE.get(request.session.session_key + "cbvsortby").page = (
            "1" if not query_dict.get(page) else query_dict.get(page)
        )
    reverse_object = CACHE.get(request.session.session_key + "cbvsortby")
    reverse = reverse_object.reverse
    none_ids = []
    none_queryset = []
    model = queryset.model
    model_attr = getmodelattribute(model, sort_key)
    is_method = (
        isinstance(model_attr, types.FunctionType)
        or model_attr not in model._meta.get_fields()
    )
    if not is_method:
        none_queryset = queryset.filter(**{f"{sort_key}__isnull": True})
        none_ids = list(none_queryset.values_list("id", flat=True))
        queryset = queryset.exclude(id__in=none_ids)

    def _sortby(object):
        result = getattribute(object, attr=sort_key)
        if result is None:
            none_ids.append(object.pk)
        return result

    order = not reverse
    current_page = query_dict.get(page)
    if current_page or is_first_sort:
        order = not order
        if reverse_object.page == current_page and not is_first_sort:
            order = not order
        reverse_object.page = current_page
    try:
        queryset = sorted(queryset, key=_sortby, reverse=order)
    except TypeError:
        none_queryset = list(queryset.filter(id__in=none_ids))
        queryset = sorted(queryset.exclude(id__in=none_ids), key=_sortby, reverse=order)

    reverse_object.reverse = order
    if order:
        order = "asc"
        queryset = list(queryset) + list(none_queryset)
    else:
        queryset = list(none_queryset) + list(queryset)
        order = "desc"
    setattr(request, "sort_order", order)
    setattr(request, "sort_key", sort_key)
    CACHE.set(request.session.session_key + "cbvsortby", reverse_object)
    return queryset


def update_saved_filter_cache(request, cache):
    """
    Method to save filter on cache
    """
    if cache.get(request.session.session_key + request.path + "cbv"):
        cache.get(request.session.session_key + request.path + "cbv").update(
            {
                "path": request.path,
                "query_dict": request.GET,
                # "request": request,
            }
        )
        return cache
    cache.set(
        request.session.session_key + request.path + "cbv",
        {
            "path": request.path,
            "query_dict": request.GET,
            # "request": request,
        },
    )
    return cache


def get_nested_field(model_class: models.Model, field_name: str) -> object:
    """
    Recursion function to execute nested field logic
    """
    if "__" in field_name:
        splits = field_name.split("__", 1)
        related_model_class = getmodelattribute(
            model_class,
            splits[0],
        ).related.related_model
        return get_nested_field(related_model_class, splits[1])
    field = getattribute(model_class, field_name)
    return field


def get_field_class_map(model_class: models.Model, bulk_update_fields: list) -> dict:
    """
    Returns a dictionary mapping field names to their corresponding field classes
    for a given model class, including related fields(one-to-one).
    """
    field_class_map = {}
    for field_name in bulk_update_fields:
        field = get_nested_field(model_class, field_name)
        field_class_map[field_name] = field.field
    return field_class_map


def structured(self):
    """
    Render the form fields as HTML table rows with Bootstrap styling.
    """
    request = getattr(_thread_locals, "request", None)
    context = {
        "form": self,
        "request": request,
    }
    table_html = render_to_string("generic/form.html", context)
    return table_html


def get_original_model_field(historical_model):
    """
    Given a historical model and a field name,
    return the actual model field from the original model.
    """
    model_name = historical_model.__name__.replace("Historical", "")
    app_label = historical_model._meta.app_label
    try:
        original_model = apps.get_model(app_label, model_name)
        return original_model
    except Exception as e:
        return historical_model


def value_to_field(field: object, value: list) -> Any:
    """
    return value according to the format of the field
    """
    from base.methods import eval_validate

    if isinstance(field, models.ManyToManyField):
        return [int(val) for val in value]
    elif isinstance(
        field,
        (
            models.DateField,
            models.DateTimeField,
            models.CharField,
            models.EmailField,
            models.TextField,
            models.TimeField,
        ),
    ):
        value = value[0]
        return value
    value = eval_validate(str(value[0]))
    return value


def merge_dicts(dict1, dict2):
    """
    Method to merge two dicts
    """
    merged_dict = dict1.copy()

    for key, value in dict2.items():
        if key in merged_dict:
            for model_class, instances in value.items():
                if model_class in merged_dict[key]:
                    merged_dict[key][model_class].extend(instances)
                else:
                    merged_dict[key][model_class] = instances
        else:
            merged_dict[key] = value

    return merged_dict


def flatten_dict(d, parent_key=""):
    """Recursively flattens a nested dictionary"""
    items = []
    for k, v in d.items():
        new_key = k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key).items())
        else:
            items.append((new_key, v))
    return dict(items)


def export_xlsx(json_data, columns, file_name="quick_export"):
    """
    Quick export method
    """
    top_fields = [col[0] for col in columns if len(col) == 2]

    nested_fields = [
        col for col in columns if len(col) == 3 and isinstance(col[2], dict)
    ]

    # Discover dynamic keys for each nested column
    dynamic_columns = {}
    for title, key, mappings in nested_fields:
        dyn_keys = set()
        for entry in json_data:
            try:
                nested_data = json.loads(entry.get(key, "[]").replace("'", '"'))
                for item in nested_data:
                    flat = flatten_dict(item)
                    dyn_keys.update(flat.keys())
            except Exception:
                continue
        dynamic_columns[key] = {
            "title": title,
            "keys": [k for k in mappings if k in dyn_keys],
            "display_names": mappings,
        }

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Quick Export"

    # Header row
    header = top_fields[:]
    for nested_info in dynamic_columns.values():
        for dyn_key in nested_info["keys"]:
            display_name = nested_info["display_names"].get(dyn_key, dyn_key)
            header.append(display_name)
    ws.append(list(str(title) for title in header))

    # Style definitions
    header_fill = PatternFill(
        start_color="FFD700", end_color="FFD700", fill_type="solid"
    )
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Apply styles to header
    for col_idx, title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_index = 2

    for entry in json_data:
        all_nested_records = []
        max_nested_rows = 1

        for key, nested_info in dynamic_columns.items():
            try:
                nested_data = json.loads(entry.get(key, "[]").replace("'", '"'))
                if not isinstance(nested_data, list):
                    nested_data = []
            except Exception:
                nested_data = []
            all_nested_records.append(nested_data)
            max_nested_rows = max(max_nested_rows, len(nested_data))

        for i in range(max_nested_rows):
            row = []

            # Top fields
            for tf in top_fields:
                row.append(entry.get(tf, "") if i == 0 else "")

            # Nested fields
            for idx, (key, nested_info) in enumerate(dynamic_columns.items()):
                nested_data = all_nested_records[idx]
                flat_ans = flatten_dict(nested_data[i]) if i < len(nested_data) else {}
                for dyn_key in nested_info["keys"]:
                    row.append(flat_ans.get(dyn_key, ""))

            ws.append(row)

            # Apply border to row
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=row_index, column=col_idx)
                cell.border = thin_border

            row_index += 1

        # Merge top fields if needed
        if max_nested_rows > 1:
            for col_idx in range(1, len(top_fields) + 1):
                ws.merge_cells(
                    start_row=row_index - max_nested_rows,
                    start_column=col_idx,
                    end_row=row_index - 1,
                    end_column=col_idx,
                )
                top_cell = ws.cell(row=row_index - max_nested_rows, column=col_idx)
                top_cell.alignment = Alignment(vertical="center")
                top_cell.border = thin_border  # Re-apply border

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Output file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}.xlsx"'
    return response


from django.apps import apps
from django.core.exceptions import FieldDoesNotExist
from django.db.models import Model
from django.db.models.fields.related import (
    ForeignKey,
    ManyToManyRel,
    ManyToOneRel,
    OneToOneField,
    OneToOneRel,
)
from openpyxl import Workbook


def get_verbose_name_from_field_path(model, field_path, import_mapping):
    """
    Get verbose name
    """
    parts = field_path.split("__")
    current_model = model
    verbose_name = None

    for i, part in enumerate(parts):
        try:
            field = current_model._meta.get_field(part)

            # Skip reverse relations (e.g., OneToOneRel)
            if isinstance(field, (OneToOneRel, ManyToOneRel, ManyToManyRel)):
                related_model = field.related_model
                field = getattr(related_model, parts[-1]).field
                return field.verbose_name.title()

            verbose_name = field.verbose_name

            if isinstance(field, (ForeignKey, OneToOneField)):
                current_model = field.related_model

        except FieldDoesNotExist:
            return f"[Invalid: {field_path}]"

    return verbose_name.title() if verbose_name else field_path


def generate_import_excel(
    base_model, import_fields, reference_field="id", import_mapping={}, queryset=[]
):
    """
    Generate import excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Sheet"

    # Style definitions
    header_fill = PatternFill(
        start_color="FFD700", end_color="FFD700", fill_type="solid"
    )
    bold_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Generate headers
    headers = [
        get_verbose_name_from_field_path(base_model, field, import_mapping)
        for field in import_fields
    ]
    headers = [
        f"{get_verbose_name_from_field_path(base_model, reference_field,import_mapping)} | Reference"
    ] + headers
    ws.append(headers)

    # Apply styles to header row
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment
        cell.border = thin_border

        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 30

    for obj in queryset:
        row = [str(getattribute(obj, reference_field))] + [
            str(getattribute(obj, import_mapping.get(field, field)))
            for field in import_fields
        ]
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.freeze_panes = "B2"
    return wb


def split_by_import_reference(employee_data):
    with_import_reference = []
    without_import_reference = []

    for record in employee_data:
        if record.get("id_import_reference") is not None:
            with_import_reference.append(record)
        else:
            without_import_reference.append(record)

    return with_import_reference, without_import_reference


def resolve_foreign_keys(
    base_model,
    record,
    import_column_mapping,
    model_lookup,
    primary_key_mapping,
    pk_values_mapping,
    prefix="",
):
    resolved = {}

    for key, value in record.items():
        full_key = f"{prefix}__{key}" if prefix else key

        if isinstance(value, dict):
            try:
                field = base_model._meta.get_field(key)
                related_model = field.related_model
            except Exception:
                resolved[key] = value
                continue

            # Recursively resolve nested foreign keys
            nested_data = resolve_foreign_keys(
                related_model,
                value,
                import_column_mapping,
                model_lookup,
                primary_key_mapping,
                pk_values_mapping,
                prefix=full_key,
            )
            instance = related_model.objects.create(**nested_data)
            resolved[key] = instance

        else:
            model_class = model_lookup.get(full_key)
            lookup_field = primary_key_mapping.get(full_key)

            if model_class and lookup_field:
                if value in [None, ""]:
                    resolved[key] = None
                    continue

                try:
                    instance, _ = model_class.objects.get_or_create(
                        **{lookup_field: value}
                    )
                    resolved[key] = instance
                except Exception as e:
                    raise ValueError(
                        f"Failed to get_or_create '{model_class.__name__}' using {lookup_field}={value}: {e}"
                    )
            else:
                resolved[key] = value

    return resolved


def update_related(
    obj,
    record,
    primary_key_mapping,
    reverse_model_relation_to_base_model,
):
    related_objects = {
        key: getattribute(obj, key) or None
        for key in reverse_model_relation_to_base_model
    }
    for relation in reverse_model_relation_to_base_model:
        related_record_info = record.get(relation)
        for key, value in related_record_info.items():
            related_object = related_objects[relation]
            obj_related_field = relation + "__" + key
            pk_mapping = primary_key_mapping.get(obj_related_field)
            if obj_related_field in primary_key_mapping and pk_mapping:
                previous_obj = getattr(related_object, key, None)
                if previous_obj and value is not None:
                    new_obj = previous_obj._meta.model.objects.get(
                        **{pk_mapping: value}
                    )
                    setattr(related_object, key, new_obj)
            else:
                if value is not None:
                    setattr(related_object, key, value)
            if related_object:
                related_object.save()


def assign_related(
    record,
    reverse_field,
    pk_values_mapping,
    pk_field_mapping,
):
    """
    Method to assign related records
    """
    reverse_obj_dict = {}
    if reverse_field in record:
        if isinstance(record[reverse_field], dict):
            for field, value in record[reverse_field].items():
                full_field = reverse_field + "__" + field
                if full_field in pk_values_mapping:
                    reverse_obj_dict.update(
                        {
                            field: data
                            for data in pk_values_mapping[full_field]
                            if getattr(data, pk_field_mapping[full_field], None)
                            == value
                        }
                    )
                else:
                    reverse_obj_dict[field] = value
        else:
            instances = [
                data
                for data in pk_values_mapping[reverse_field]
                if getattr(
                    data,
                    pk_field_mapping[reverse_field],
                    record[reverse_field],
                )
                == record[reverse_field]
            ]
            if instances:
                instance = instances[0]
                reverse_obj_dict.update({reverse_field: instance})
    return reverse_obj_dict
