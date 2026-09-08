import csv
import importlib
import json
import os
import re
from collections import defaultdict
from io import BytesIO

from arabic_reshaper import ArabicReshaper
from bidi.algorithm import get_display
from django import forms
from django.apps import apps
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.utils import NestedObjects
from django.contrib.contenttypes.models import ContentType
from django.contrib.staticfiles import finders
from django.core.cache import cache as CACHE
from django.core.exceptions import FieldDoesNotExist
from django.db import connection, router
from django.db.models.fields.related import ForeignKey, OneToOneField
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
from django.utils.translation import gettext as _
from django.views import View
from django.views.decorators.csrf import csrf_protect
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from xhtml2pdf import pisa

from base.methods import eval_validate, has_export_access
from horilla.decorators import login_required as func_login_required
from horilla.export_safety import safe_cell
from horilla.http.response import HorillaRedirect
from horilla.signals import post_generic_delete, pre_generic_delete
from horilla_views import models
from horilla_views.cbv_methods import (
    get_nested_field,
    get_short_uuid,
    hx_request_required,
    login_required,
    merge_dicts,
)
from horilla_views.forms import SavedFilterForm
from horilla_views.generic.cbv.views import HorillaFormView, HorillaListView
from horilla_views.templatetags.generic_template_filters import getattribute

# Create your views here.


reshaper = ArabicReshaper(
    {
        "support_ligatures": False,
    }
)


@method_decorator(login_required, name="dispatch")
class ToggleColumn(View):
    """
    ToggleColumn
    """

    def get(self, *args, **kwargs):
        """
        method to toggle columns
        """

        query_dict = self.request.GET
        path = query_dict.get("path")

        if not path:
            return HorillaRedirect(
                self.request,
                message=_("No matching query found."),
            )
        query_dict = dict(query_dict)
        del query_dict["path"]

        field_order = [key for key, v in query_dict.items()]
        # First column in order is the primary/fixed column — always keep it visible.
        primary_column = field_order[0] if field_order else None
        hidden_fields = [
            key
            for key, value in query_dict.items()
            if value[0] and key != primary_column
        ]

        existing_instance = models.ToggleColumn.objects.filter(
            user_id=self.request.user, path=path
        ).first()

        existing_field_order = models.ColumnOrder.objects.filter(
            employee=self.request.user.employee_get, path=path
        ).first()

        instance = models.ToggleColumn() if not existing_instance else existing_instance
        instance.path = path
        instance.excluded_columns = hidden_fields
        instance.save()

        column_order = (
            models.ColumnOrder() if not existing_field_order else existing_field_order
        )
        column_order.path = path
        column_order.column_order = field_order
        column_order.save()

        return HttpResponse("success")


@method_decorator(login_required, name="dispatch")
class ReloadField(View):
    """
    ReloadField
    """

    def get(self, request, *args, **kwargs):
        """
        Http method to reload dynamic create fields
        """
        class_path = request.GET.get("form_class_path")
        reload_field = request.GET.get("dynamic_field")

        if not class_path:
            return HorillaRedirect(
                request,
                message=_("No matching query found."),
            )

        module_name, class_name = class_path.rsplit(".", 1)
        module = importlib.import_module(module_name)
        form_class = getattr(module, class_name, None)
        if not (
            isinstance(form_class, type) and issubclass(form_class, forms.BaseForm)
        ):
            return HorillaRedirect(
                request,
                message=_("No matching query found."),
            )
        parent_form = form_class()

        dynamic_cache = CACHE.get(request.session.session_key + "cbv" + reload_field)
        onchange = CACHE.get(
            request.session.session_key + "cbv" + reload_field + "onchange"
        )
        if not onchange:
            onchange = ""

        model: models.HorillaModel = dynamic_cache["model"]
        value = dynamic_cache.get("value", "")

        cache_field = dynamic_cache["dynamic_field"]
        if cache_field != reload_field:
            cache_field = reload_field
        field = parent_form.fields[cache_field]

        queryset = model.objects.all()
        queryset = field.queryset
        choices = [(instance.id, instance) for instance in queryset]
        choices.insert(0, ("", "Select option"))
        choices.append(("dynamic_create", "Dynamic create"))

        form_field = forms.ChoiceField
        if isinstance(field, forms.ModelMultipleChoiceField):
            form_field = forms.MultipleChoiceField
            dynamic_initial = request.GET.get("dynamic_initial", [])
            value = eval_validate(f"""[{dynamic_cache["value"]},{dynamic_initial}]""")
        else:
            if not value and self.request.GET.get("dynamic_initial"):
                value = eval_validate(self.request.GET.get("dynamic_initial"))

        parent_form.fields[cache_field] = form_field(
            choices=choices,
            label=field.label,
            required=field.required,
        )
        parent_form.fields[cache_field].widget.option_template_name = (
            "horilla_widgets/select_option.html",
        )
        parent_form.fields[cache_field].widget.attrs = field.widget.attrs
        parent_form.fields[cache_field].initial = value
        parent_form.fields[cache_field].widget.attrs["onchange"] = onchange

        field = parent_form[cache_field]
        dynamic_id: str = get_short_uuid(4)
        return render(
            request,
            "generic/reload_select_field.html",
            {"field": field, "dynamic_id": dynamic_id},
        )


@method_decorator(login_required, name="dispatch")
class ActiveTab(View):
    def get(self, *args, **kwargs):
        """
        CBV method to handle active tab
        """
        path = self.request.GET.get("path")
        target = self.request.GET.get("target")
        if path and target and self.request.user:
            existing_instance = models.ActiveTab.objects.filter(
                created_by=self.request.user, path=path
            ).first()

            instance = (
                models.ActiveTab() if not existing_instance else existing_instance
            )
            instance.path = path
            instance.tab_target = target
            instance.save()
        return JsonResponse({"message": "Success"})


@method_decorator(login_required, name="dispatch")
class ActiveGroup(View):
    def get(self, *args, **kwargs):
        """
        ActiveGroup
        """
        path = self.request.GET.get("path")
        target = self.request.GET.get("target")
        group_field = self.request.GET.get("field")
        if path and target and group_field and self.request.user:
            existing_instance = models.ActiveGroup.objects.filter(
                created_by=self.request.user,
                path=path,
                group_by_field=group_field,
            ).first()

            instance = (
                models.ActiveGroup() if not existing_instance else existing_instance
            )
            instance.path = path
            instance.group_by_field = group_field
            instance.group_target = target
            instance.save()
        return JsonResponse({"message": "Success"})


@method_decorator(login_required, name="dispatch")
class SavedFilter(HorillaFormView):
    """
    SavedFilter
    """

    model = models.SavedFilter
    form_class = SavedFilterForm
    new_display_title = "Save Applied Filter"
    template_name = "generic/saved_filter_form.html"
    form_disaply_attr = "Blah"

    def form_valid(self, form: SavedFilterForm) -> HttpResponse:
        referrer = self.request.POST.get("referrer", "")
        path = self.request.POST.get("path", "/")
        result_dict = {key: value[0] for key, value in self.request.GET.lists()}
        if form.is_valid():
            instance: models.SavedFilter = form.save(commit=False)
            if not instance.pk:
                instance.path = path
                instance.referrer = referrer
                instance.filter = result_dict
                instance.urlencode = self.request.GET.urlencode()
            instance.save()
            messages.success(self.request, _("Filter Saved"))
            return self.HttpResponse()
        return super().form_valid(form)

    def get_context_data(self, **kwargs) -> dict:
        context = super().get_context_data(**kwargs)
        referrer = self.request.GET.get("referrer", "")
        if referrer:
            # Remove the protocol and domain part
            referrer = "/" + "/".join(referrer.split("/")[3:])
        context["path"] = self.request.GET.get("path", "")
        context["referrer"] = referrer
        return context


@method_decorator(login_required, name="dispatch")
@method_decorator(hx_request_required, name="dispatch")
class DeleteSavedFilter(View):
    """
    Delete saved filter
    """

    def get(self, *args, **kwargs):
        pk = kwargs["pk"]
        models.SavedFilter.objects.filter(created_by=self.request.user, pk=pk).delete()
        return HttpResponse("")


@method_decorator(login_required, name="dispatch")
class ActiveView(View):
    """
    ActiveView CBV
    """

    def get(self, *args, **kwargs):
        path = self.request.GET.get("path")
        view_type = self.request.GET.get("view")

        if not path:
            return HorillaRedirect(
                self.request,
                message=_("No matching query found."),
            )
        active_view = models.ActiveView.objects.filter(
            path=path, created_by=self.request.user
        ).first()

        active_view = active_view if active_view else models.ActiveView()
        active_view.path = path
        active_view.type = view_type
        active_view.save()
        return HttpResponse("")


@method_decorator(login_required, name="dispatch")
@method_decorator(hx_request_required, name="dispatch")
@method_decorator(csrf_protect, name="dispatch")
class SearchInIds(View):
    """
    Search in ids view
    """

    def get(self, *args, **kwargs):
        """
        Search in instance ids method
        """
        cache_key = f"{self.request.session.session_key}search_in_instance_ids"
        context: dict = CACHE.get(cache_key)
        if context:
            context["instances"] = context["filter_class"](self.request.GET).qs
        return render(self.request, "generic/filter_result.html", context)


@method_decorator(login_required, name="dispatch")
@method_decorator(hx_request_required, name="dispatch")
class LastAppliedFilter(View):
    """
    Class view to handle last applied filter
    """

    def get(self, *args, **kwargs):
        """
        Get method
        """

        nav_path = self.request.GET.get(
            "nav_url",
        )
        if nav_path:
            CACHE.set(
                self.request.session.session_key + "last-applied-filter" + nav_path,
                self.request.GET,
                timeout=600,
            )
        return HttpResponse("success")


@method_decorator(login_required, name="dispatch")
class DynamiListView(HorillaListView):
    """
    DynamicListView for Generic Delete
    """

    instances = []

    def get_queryset(self):
        search = self.request.GET.get("search", "")

        def _search_filter(instance):
            return search in str(instance).lower()

        return filter(_search_filter, self.instances)


@method_decorator(login_required, name="dispatch")
class HorillaDeleteConfirmationView(View):
    """
    Generic Delete Confirmation View
    """

    confirmation_target = "deleteConfirmationBody"
    # URL name used by delete_confirmation.html for hx-get / hx-post on this flow
    generic_delete_url_name = "generic-delete"

    def get(self, *args, **kwargs):
        """
        GET method
        """
        from horilla.urls import path, urlpatterns

        pk = self.request.GET.get("pk")
        try:
            app, MODEL_NAME = self.request.GET.get("model").split(".")
        except:
            messages.error(self.request, _("Invalid model parameter format."))
            return HorillaFormView.HttpResponse()

        if not self.request.user.has_perm(app + ".delete_" + MODEL_NAME.lower()):
            return render(self.request, "no_perm.html")
        model = apps.get_model(app, MODEL_NAME)

        delete_object = model.objects.filter(pk=pk).first()
        if not delete_object:
            messages.error(self.request, _("Record not found."))
            return HorillaFormView.HttpResponse()
        objs = [delete_object]
        using = router.db_for_write(delete_object._meta.model)
        collector = NestedObjects(using=using, origin=objs)
        collector.collect(objs)
        MODEL_MAP = {}
        PROTECTED_MODEL_MAP = {}
        DYNAMIC_PATH_MAP = {}
        MODEL_RELATED_FIELD_MAP = {}
        MODEL_RELATED_PROTECTED_FIELD_MAP = {}

        def format_callback(instance, protected=False):
            if not MODEL_RELATED_FIELD_MAP.get(instance._meta.model):
                MODEL_RELATED_FIELD_MAP[instance._meta.model] = []
                MODEL_RELATED_PROTECTED_FIELD_MAP[instance._meta.model] = []

            def find_related_field(obj, related_instance):
                for field in obj._meta.get_fields():
                    # Check if the field is a foreign key (or related model)
                    if isinstance(
                        field, (models.models.ForeignKey, models.models.OneToOneField)
                    ):
                        # Get the field value
                        field_value = getattr(obj, field.name)
                        # If the field value matches the related instance, return the field name
                        if field_value == related_instance:
                            if "PROTECT" in field.remote_field.on_delete.__name__:
                                MODEL_RELATED_PROTECTED_FIELD_MAP[
                                    instance._meta.model
                                ].append((field.name, field.verbose_name))
                            MODEL_RELATED_FIELD_MAP[instance._meta.model].append(
                                field.name
                            )

            find_related_field(instance, delete_object)
            app_label = instance._meta.app_label
            app_label = apps.get_app_config(app_label).verbose_name
            model = instance._meta.model

            model.verbose_name = model.__name__.split("_")[0]

            model_map = PROTECTED_MODEL_MAP if protected else MODEL_MAP

            if app_label not in model_map:
                model_map[app_label] = {}

            if model not in model_map[app_label]:
                model_map[app_label][model] = []
                DYNAMIC_PATH_MAP[model.verbose_name] = (
                    f"{get_short_uuid(prefix='generic-delete',length=10)}"
                )

                class DynamiListView(HorillaListView):
                    """
                    DynamicListView for Generic Delete
                    """

                    instances = []
                    columns = [
                        (
                            "Record",
                            "dynamic_display_name_generic_delete",
                        ),
                    ]
                    records_per_page = 5
                    filter_selected = False
                    quick_export = False

                    selected_instances_key_id = "storedIds" + app_label

                    def dynamic_display_name_generic_delete(self):

                        is_protected = False
                        classname = self.__class__.__name__
                        app_label = self._meta.app_label

                        app_verbose_name = apps.get_app_config(app_label).verbose_name
                        protected = PROTECTED_MODEL_MAP.get(app_verbose_name, {}).get(
                            self._meta.model, []
                        )
                        ids = [instance.pk for instance in protected]
                        if self.pk in ids:
                            is_protected = True

                        if "_" in classname:
                            field_name = classname.split("_", 1)[1]
                            classname = classname.split("_")[0]

                            object_field_name = classname.lower()
                            model = apps.get_model(app_label, classname)

                            field = model._meta.get_field(field_name)

                            return f"""
                            {getattr(self, object_field_name)}
                            <i style="color:#989898;">(In {field.verbose_name})</i>
                            """
                        indication = f"""
                        {self}
                        """
                        if is_protected:
                            verbose_names = [
                                str(i[1])
                                for i in list(
                                    set(
                                        MODEL_RELATED_PROTECTED_FIELD_MAP.get(
                                            self._meta.model, ""
                                        )
                                    )
                                )
                            ]
                            indication = (
                                indication
                                + f"""
                            <i style="color:red;">(Record in {",".join(verbose_names)})</i>
                            """
                            )
                        return indication

                    def __init__(self, **kwargs):
                        super().__init__(**kwargs)
                        self._saved_filters = self.request.GET

                    def get_context_data(self, **kwargs):
                        context = super().get_context_data(**kwargs)
                        context["search_url"] = "/" + self.search_url
                        return context

                    def get_queryset(self):
                        search = self.request.GET.get("search", "")

                        def _search_filter(instance):
                            return search in str(instance).lower()

                        self.instances = list(
                            set(
                                (
                                    self.instances
                                    + MODEL_MAP.get(
                                        apps.get_app_config(
                                            self.model._meta.app_label
                                        ).verbose_name,
                                        {},
                                    ).get(self.model, [])
                                    + PROTECTED_MODEL_MAP.get(
                                        apps.get_app_config(
                                            self.model._meta.app_label
                                        ).verbose_name,
                                        {},
                                    ).get(self.model, [])
                                )
                            )
                        )

                        queryset = self.model.objects.filter(
                            pk__in=[
                                instance.pk
                                for instance in filter(_search_filter, self.instances)
                            ]
                        )
                        return queryset

                model.dynamic_display_name_generic_delete = (
                    DynamiListView.dynamic_display_name_generic_delete
                )

                DynamiListView.model = model
                if "_" in model.__name__:
                    DynamiListView.bulk_update_fields = [MODEL_NAME.lower()]
                else:
                    DynamiListView.bulk_update_fields = MODEL_RELATED_FIELD_MAP.get(
                        model, []
                    )
                DynamiListView.instances = model_map[app_label][model]
                DynamiListView.search_url = DYNAMIC_PATH_MAP[model.verbose_name]
                DynamiListView.selected_instances_key_id = (
                    DynamiListView.selected_instances_key_id + model.verbose_name
                )

                urlpatterns.append(
                    path(
                        DynamiListView.search_url,
                        DynamiListView.as_view(),
                        name=DynamiListView.search_url,
                    )
                )
            model_map[app_label][model].append(instance)

            return instance

        _to_delete = collector.nested(format_callback)
        protected = [
            format_callback(obj, protected=True) for obj in collector.protected
        ]

        model_count = {
            model._meta.verbose_name_plural: len(objs)
            for model, objs in collector.model_objs.items()
        }

        protected_model_count = defaultdict(int)

        for obj in collector.protected:
            model = type(obj)
            protected_model_count[model._meta.verbose_name_plural] += 1
        protected_model_count = dict(protected_model_count)
        context = {
            "model_map": merge_dicts(MODEL_MAP, PROTECTED_MODEL_MAP),
            "dynamic_list_path": DYNAMIC_PATH_MAP,
            "delete_object": delete_object,
            "protected": protected,
            "model_count_sum": sum(model_count.values()),
            "related_objects_count": model_count,
            "protected_objects_count": protected_model_count,
        }
        for key, value in self.get_context_data().items():
            context[key] = value

        context["generic_delete_url_name"] = getattr(
            self, "generic_delete_url_name", "generic-delete"
        )

        return render(self.request, "generic/delete_confirmation.html", context)

    def post(self, *args, **kwargs):
        """
        Post method to handle the delete
        """
        confirmations = ["action", "revert", "confirm"]
        if not all(self.request.POST.get(key) == "on" for key in confirmations):
            messages.error(
                self.request,
                "All confirmation checkboxes must be acknowledged before deleting.",
            )
            return self.get(*args, **kwargs)

        pk = self.request.GET["pk"]
        app, MODEL_NAME = self.request.GET["model"].split(".")
        if not self.request.user.has_perm(app + ".delete_" + MODEL_NAME.lower()):
            return render(self.request, "no_perm.html")
        model = apps.get_model(app, MODEL_NAME)
        delete_object = model.objects.get(pk=pk)
        objs = [delete_object]
        using = router.db_for_write(delete_object._meta.model)
        collector = NestedObjects(using=using, origin=objs)
        collector.collect(objs)

        def delete_callback(instance, protected=False):
            try:
                if self.request.user.has_perm(
                    f"{instance._meta.app_label}.delete_{instance._meta.model.__name__.lower()}"
                ):
                    pre_generic_delete.send(
                        sender=instance._meta.model,
                        instance=instance,
                        args=args,
                        view_instance=self,
                        kwargs=kwargs,
                    )
                    instance.delete()
                    post_generic_delete.send(
                        sender=instance._meta.model,
                        instance=instance,
                        args=args,
                        view_instance=self,
                        kwargs=kwargs,
                    )
                    messages.success(
                        self.request, _("Deleted %(instance)s") % {"instance": instance}
                    )
                else:
                    messages.info(
                        self.request,
                        _("You don't have permission to delete %(instance)s")
                        % {"instance": instance},
                    )
            except:
                messages.error(
                    self.request,
                    _("Cannot delete : %(instance)s") % {"instance": instance},
                )

        # deleting protected objects
        for obj in collector.protected:
            delete_callback(obj, protected=True)
        # deleting related objects
        collector.nested(delete_callback)
        reload_target = self.request.GET.get("reload_target")
        script = ""
        if reload_target:
            script = f"$('{reload_target}').first().click();"

        return HorillaFormView.HttpResponse(script=script)

    def get_context_data(self, **kwargs) -> dict:
        context = {}
        context["confirmation_target"] = self.confirmation_target
        return context


_getattibute = getattribute


def sanitize_filename(filename):
    return re.sub(r'[<>:"/\\|?*\[\]]+', "_", filename)[:200]  # limit to 200 chars


def link_callback(uri, rel):
    """
    Convert html URIs to absolute system paths so xhtml2pdf can access them.
    Called by pisa.CreatePDF(..., link_callback=link_callback)
    """
    # If absolute URL (http/https/file) return as-is
    if (
        uri.startswith("http://")
        or uri.startswith("https://")
        or uri.startswith("file://")
    ):
        return uri

    # Try static files first
    static_path = None
    if uri.startswith(settings.STATIC_URL):
        # remove STATIC_URL prefix
        rel_path = uri.replace(settings.STATIC_URL, "")
        # find with staticfiles finders
        found = finders.find(rel_path)
        if found:
            static_path = found

    # Try media files next
    media_path = None
    if uri.startswith(settings.MEDIA_URL):
        rel_path = uri.replace(settings.MEDIA_URL, "")
        media_path = os.path.join(settings.MEDIA_ROOT, rel_path)

    # If a path found, return it
    for path in (static_path, media_path, uri):
        if path and os.path.exists(path):
            return path

    # Last resort: maybe it's relative to your project root
    project_path = os.path.join(settings.BASE_DIR, uri)
    if os.path.exists(project_path):
        return project_path

    raise Exception("File not found for URI: %s" % uri)


def reshape_text(text):
    """
    Make text safe for xhtml2pdf:
    - Reshape Arabic
    - Apply bidi ordering
    - Leave all other languages untouched
    """

    if not isinstance(text, str):
        return text
    try:
        reshaped = reshaper.reshape(text)
        return get_display(reshaped)
    except Exception:
        return text


# Attribute names a client-supplied export column path may never traverse.
_EXPORT_FORBIDDEN_PARTS = frozenset(
    {
        "password",
        "employee_user_id",
        "user",
        "session",
        "token",
        "secret",
        "api_key",
        "api_secret",
        "otp",
    }
)


@func_login_required
def export_data(request, *args, **kwargs):

    # =====================================================
    # INPUT
    # =====================================================
    ids = eval_validate(request.POST.get("ids", "[]"))
    columns = eval_validate(request.POST.get("columns", "[]"))

    export_format = request.POST.get("format", "csv").lower()
    file_name = request.POST.get("export_file_name", "quick_export")

    date_range = request.session.get("report_date_range", "")

    # =====================================================
    # COMPANY
    # =====================================================
    company_name = "All Company"
    logo_path = None

    company = getattr(request, "selected_company_instance", None)
    if company:
        company_name = company.company or company_name
        if company.icon:
            logo_path = company.icon.path

    # =====================================================
    # MODEL
    # =====================================================
    model_path = request.GET.get("model")
    if not model_path:
        return HorillaRedirect(request, message=_("No matching query found."))
    app_label = model_path.split(".")[0]
    model_name = model_path.split(".")[-1]
    try:
        model = apps.get_model(app_label, model_name)
    except LookupError:
        return HorillaRedirect(request, message=_("No matching query found."))
    base_table = model._meta.db_table

    # =====================================================
    # EXPORT ACCESS CONTROL
    # =====================================================
    if not has_export_access(request, model):
        messages.info(request, _("You dont have access to export this data"))
        return HorillaRedirect(request)

    # =====================================================
    # SQL BUILD
    # =====================================================
    headers = []
    select_sql = [f"{base_table}.id"]
    method_columns = {}
    select_index = 1

    for label, field in columns:
        headers.append(label)

        if not field:
            select_sql.append("''")
            select_index += 1
            continue

        parts = field.split("__")

        if len(parts) == 1:
            try:
                f = model._meta.get_field(parts[0])
                if isinstance(f, (ForeignKey, OneToOneField)):
                    select_sql.append(f"{base_table}.id")
                    method_columns[select_index] = parts[0]
                else:
                    select_sql.append(f"{base_table}.{f.column}")
            except FieldDoesNotExist:
                select_sql.append(f"{base_table}.id")
                method_columns[select_index] = parts[0]

            select_index += 1
            continue

        select_sql.append(f"{base_table}.id")
        method_columns[select_index] = field
        select_index += 1

    # =====================================================
    # EXECUTE SQL
    # =====================================================
    if not ids:
        return HttpResponse("No IDs provided")

    # Re-derive the id list through the model's own manager instead of
    # trusting the client-supplied `ids` as-is: `model.objects` applies
    # this app's company/permission scoping (HorillaCompanyManager), the
    # same scoping the originating list view relies on, while the raw SQL
    # below has no access control of its own and would otherwise return
    # rows for any id an authenticated user cares to submit, including
    # records outside their company.
    authorized_ids = list(model.objects.filter(id__in=ids).values_list("id", flat=True))
    if not authorized_ids:
        return HttpResponse("No IDs provided")

    placeholders = ", ".join(["%s"] * len(authorized_ids))

    query = f"""
        SELECT {", ".join(select_sql)}
        FROM {base_table}
        WHERE {base_table}.id IN ({placeholders})
    """

    with connection.cursor() as cursor:
        cursor.execute(query, authorized_ids)
        rows = cursor.fetchall()

    # =====================================================
    # ORM CACHE
    # =====================================================
    objs = {o.id: o for o in model.objects.filter(id__in=authorized_ids)}

    method_maps = {}
    for idx, attr in method_columns.items():
        method_maps[idx] = {}
        for obj_id, obj in objs.items():
            value = obj
            for part in attr.split("__"):
                if value is None:
                    break
                # The path is client-supplied. Refuse private attributes and
                # the handful of names that would turn a quick export into a
                # credential dump (employee_user_id__password, tokens, OTPs).
                if part.startswith("_") or part in _EXPORT_FORBIDDEN_PARTS:
                    value = None
                    break
                value = getattr(value, part, None)
                if callable(value):
                    # Client-supplied column paths can name any attribute,
                    # so only ever invoke this codebase's own read-only
                    # accessor convention (get_full_name, get_avatar,
                    # get_<field>_display, ...). Calling anything else
                    # would let an export column trigger an arbitrary
                    # zero-arg instance method, including mutating ones
                    # like delete()/save().
                    value = value() if part.startswith("get_") else None
            method_maps[idx][obj_id] = str(value) if value is not None else ""

    # =====================================================
    # FINAL ROWS
    # =====================================================
    final_rows = []

    for row in rows:
        row = list(row)
        obj_id = row[0]
        for idx, mmap in method_maps.items():
            row[idx] = mmap.get(obj_id, "")
        final_rows.append(row[1:])

    # =====================================================
    # XLSX EXPORT (FIXED WIDTH – PERFORMANCE SAFE)
    # =====================================================
    if export_format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Quick Export"

        total_columns = len(headers)
        center = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="FFD700", fill_type="solid")
        header_font = Font(bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Logo
        if logo_path and os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 120
            img.height = 60
            ws.add_image(img, "A1")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        ws.cell(row=1, column=1).value = company_name
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        ws.cell(row=1, column=1).alignment = center

        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=total_columns)
        ws.cell(row=2, column=1).value = file_name
        ws.cell(row=2, column=1).font = Font(size=14, bold=True, color="FF0000")
        ws.cell(row=2, column=1).alignment = center

        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_columns)
        ws.cell(row=4, column=1).value = date_range
        ws.cell(row=4, column=1).alignment = center

        HEADER_ROW = 5

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center

            # ✅ FIXED WIDTH (NO PERFORMANCE HIT)
            ws.column_dimensions[cell.column_letter].width = 25

        for r_idx, row in enumerate(final_rows, start=HEADER_ROW + 1):
            for c_idx, val in enumerate(row, start=1):
                # Text carried through from user-entered data can execute when
                # the workbook is opened, so guard it on the way in.
                ws.cell(row=r_idx, column=c_idx).value = safe_cell(val)

        buf = BytesIO()
        wb.save(buf)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{file_name}.xlsx"'
        return response

    # =====================================================
    # PDF EXPORT
    # =====================================================
    if export_format == "pdf":

        html = render_to_string(
            "generic/export_pdf.html",
            {
                "headers": headers,
                "rows": [dict(zip(headers, r)) for r in final_rows],
                "company_name": company_name,
                "date_range": date_range,
                "report_title": file_name,
                "logo_path": (
                    logo_path if logo_path and os.path.exists(logo_path) else None
                ),
                "landscape": len(headers) > 5,
            },
        )

        buf = BytesIO()
        pisa.CreatePDF(html, dest=buf)

        return HttpResponse(
            buf.getvalue(),
            content_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{file_name}.pdf"'},
        )

    # =====================================================
    # CSV EXPORT
    # =====================================================
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{file_name}.csv"'

    writer = csv.writer(response)
    writer.writerow([company_name])
    writer.writerow([file_name])
    writer.writerow([date_range])
    writer.writerow([])
    writer.writerow(headers)
    writer.writerows(final_rows)

    return response


@method_decorator(login_required, name="dispatch")
class DynamicView(View):
    """
    DynamicView
    """

    def get(self, request, *args, **kwargs):

        field = kwargs.get("field")
        session_key = kwargs.get("session_key")
        if session_key != request.session.session_key:
            return HttpResponseForbidden("Invalid session key.")

        return render(request, "dynamic.html", {"field": field})
