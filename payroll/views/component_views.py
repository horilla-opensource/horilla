"""
component_views.py

This module is used to write methods to the component_urls patterns respectively
"""

import json
import operator
from collections import defaultdict
from datetime import date, datetime, timedelta
from itertools import groupby
from urllib.parse import parse_qs

import pandas as pd
from django.apps import apps
from django.contrib import messages
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, QueryDict
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.views.decorators.cache import never_cache
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from base.backends import ConfiguredEmailBackend
from base.methods import (
    closest_numbers,
    eval_validate,
    filter_own_records,
    get_key_instances,
    get_next_month_same_date,
    sortby,
)
from base.models import Company
from employee.models import Employee, EmployeeWorkInformation
from horilla.decorators import (
    hx_request_required,
    login_required,
    owner_can_enter,
    permission_required,
)
from horilla.group_by import group_by_queryset
from horilla.horilla_settings import HORILLA_DATE_FORMATS
from horilla.methods import dynamic_attr, get_horilla_model_class, get_urlencode

# from leave.models import AvailableLeave
from notifications.signals import notify
from payroll.filters import (
    AllowanceFilter,
    DeductionFilter,
    LoanAccountFilter,
    PayslipFilter,
    PayslipReGroup,
    ReimbursementFilter,
)
from payroll.forms import component_forms as forms
from payroll.methods.deductions import create_deductions, update_compensation_deduction
from payroll.methods.methods import (
    calculate_employer_contribution,
    compute_net_pay,
    compute_salary_on_period,
    paginator_qry,
    save_payslip,
)
from payroll.methods.payslip_calc import (
    calculate_allowance,
    calculate_gross_pay,
    calculate_net_pay_deduction,
    calculate_post_tax_deduction,
    calculate_pre_tax_deduction,
    calculate_tax_deduction,
    calculate_taxable_gross_pay,
)
from payroll.methods.tax_calc import calculate_taxable_amount
from payroll.models.models import (
    Allowance,
    Contract,
    Deduction,
    LoanAccount,
    Payslip,
    Reimbursement,
    ReimbursementMultipleAttachment,
)
from payroll.threadings.mail import MailSendThread


def return_none(a, b):
    return None


operator_mapping = {
    "equal": operator.eq,
    "notequal": operator.ne,
    "lt": operator.lt,
    "gt": operator.gt,
    "le": operator.le,
    "ge": operator.ge,
    "icontains": operator.contains,
    "range": return_none,
}


def payroll_calculation(employee, start_date, end_date):
    """
    Calculate payroll components for the specified employee within the given date range.


    Args:
        employee (Employee): The employee for whom the payroll is calculated.
        start_date (date): The start date of the payroll period.
        end_date (date): The end date of the payroll period.


    Returns:
        dict: A dictionary containing the calculated payroll components:
    """

    basic_pay_details = compute_salary_on_period(employee, start_date, end_date)
    contract = basic_pay_details["contract"]
    contract_wage = basic_pay_details["contract_wage"]
    basic_pay = basic_pay_details["basic_pay"]
    loss_of_pay = basic_pay_details["loss_of_pay"]
    paid_days = basic_pay_details["paid_days"]
    unpaid_days = basic_pay_details["unpaid_days"]

    working_days_details = basic_pay_details["month_data"]

    updated_basic_pay_data = update_compensation_deduction(
        employee, basic_pay, "basic_pay", start_date, end_date
    )
    basic_pay = updated_basic_pay_data["compensation_amount"]
    basic_pay_deductions = updated_basic_pay_data["deductions"]

    loss_of_pay_amount = 0
    if not contract.deduct_leave_from_basic_pay:
        loss_of_pay_amount = loss_of_pay
    else:
        basic_pay = basic_pay - loss_of_pay_amount

    kwargs = {
        "employee": employee,
        "start_date": start_date,
        "end_date": end_date,
        "basic_pay": basic_pay,
        "day_dict": working_days_details,
    }
    # basic pay will be basic_pay = basic_pay - update_compensation_amount
    allowances = calculate_allowance(**kwargs)

    # finding the total allowance
    total_allowance = sum(allowance["amount"] for allowance in allowances["allowances"])

    kwargs["allowances"] = allowances
    kwargs["total_allowance"] = total_allowance
    updated_gross_pay_data = calculate_gross_pay(**kwargs)
    gross_pay = updated_gross_pay_data["gross_pay"]
    gross_pay_deductions = updated_gross_pay_data["deductions"]

    kwargs["gross_pay"] = gross_pay
    pretax_deductions = calculate_pre_tax_deduction(**kwargs)
    post_tax_deductions = calculate_post_tax_deduction(**kwargs)

    installments = (
        pretax_deductions["installments"] | post_tax_deductions["installments"]
    )

    taxable_gross_pay = calculate_taxable_gross_pay(**kwargs)
    tax_deductions = calculate_tax_deduction(**kwargs)
    federal_tax = calculate_taxable_amount(**kwargs)

    total_allowance = sum(item["amount"] for item in allowances["allowances"])
    total_pretax_deduction = sum(
        item["amount"] for item in pretax_deductions["pretax_deductions"]
    )
    total_post_tax_deduction = sum(
        item["amount"] for item in post_tax_deductions["post_tax_deductions"]
    )
    total_tax_deductions = sum(
        item["amount"] for item in tax_deductions["tax_deductions"]
    )

    total_deductions = (
        total_pretax_deduction
        + total_post_tax_deduction
        + total_tax_deductions
        + federal_tax
        + loss_of_pay_amount
    )

    net_pay = gross_pay - total_deductions
    # loss_of_pay        -> actual lop amount
    # loss_of_pay_amount -> actual lop if deduct from basic-
    #                       pay from contract is enabled
    net_pay = compute_net_pay(
        net_pay=net_pay,
        gross_pay=gross_pay,
        total_pretax_deduction=total_pretax_deduction,
        total_post_tax_deduction=total_post_tax_deduction,
        total_tax_deductions=total_tax_deductions,
        federal_tax=federal_tax,
        loss_of_pay_amount=loss_of_pay_amount,
        loss_of_pay=loss_of_pay,
    )
    updated_net_pay_data = update_compensation_deduction(
        employee, net_pay, "net_pay", start_date, end_date
    )
    net_pay = updated_net_pay_data["compensation_amount"]
    update_net_pay_deductions = updated_net_pay_data["deductions"]

    net_pay_deductions = calculate_net_pay_deduction(
        net_pay,
        post_tax_deductions["net_pay_deduction"],
        **kwargs,
    )
    net_pay_deduction_list = net_pay_deductions["net_pay_deductions"]
    for deduction in update_net_pay_deductions:
        net_pay_deduction_list.append(deduction)
    net_pay = net_pay - net_pay_deductions["net_deduction"]
    payslip_data = {
        "employee": employee,
        "contract_wage": contract_wage,
        "basic_pay": basic_pay,
        "gross_pay": gross_pay,
        "taxable_gross_pay": taxable_gross_pay["taxable_gross_pay"],
        "net_pay": net_pay,
        "allowances": allowances["allowances"],
        "paid_days": paid_days,
        "unpaid_days": unpaid_days,
        "basic_pay_deductions": basic_pay_deductions,
        "gross_pay_deductions": gross_pay_deductions,
        "pretax_deductions": pretax_deductions["pretax_deductions"],
        "post_tax_deductions": post_tax_deductions["post_tax_deductions"],
        "tax_deductions": tax_deductions["tax_deductions"],
        "net_deductions": net_pay_deduction_list,
        "total_deductions": total_deductions,
        "loss_of_pay": loss_of_pay,
        "federal_tax": federal_tax,
        "start_date": start_date,
        "end_date": end_date,
        "range": f"{start_date.strftime('%b %d %Y')} - {end_date.strftime('%b %d %Y')}",
    }
    data_to_json = payslip_data.copy()
    data_to_json["employee"] = employee.id
    data_to_json["start_date"] = start_date.strftime("%Y-%m-%d")
    data_to_json["end_date"] = end_date.strftime("%Y-%m-%d")
    json_data = json.dumps(data_to_json)

    payslip_data["json_data"] = json_data
    payslip_data["installments"] = installments
    return payslip_data


@login_required
@hx_request_required
def allowances_deductions_tab(request, emp_id):
    """
    Retrieve and render the allowances and deductions applicable to an employee.

    This view function retrieves the active contract, basic pay, allowances, and
    deductions for a specified employee. It filters allowances and deductions
    based on various conditions, including specific employee assignments and
    condition-based rules. The results are then rendered in the allowance and
    deduction tab template.
    """
    employee_deductions = []
    employee_allowances = []
    employee = Employee.objects.get(id=emp_id)
    active_contracts = employee.contract_set.filter(contract_status="active").first()
    basic_pay = active_contracts.wage if active_contracts else None
    if basic_pay:
        allowances = (
            Allowance.objects.filter(specific_employees=employee)
            | Allowance.objects.filter(is_condition_based=True).exclude(
                exclude_employees=employee
            )
            | Allowance.objects.filter(include_active_employees=True).exclude(
                exclude_employees=employee
            )
        )

        for allowance in allowances:
            applicable = True
            if allowance.is_condition_based:
                conditions = list(
                    allowance.other_conditions.values_list(
                        "field", "condition", "value"
                    )
                )
                conditions.append(
                    (
                        allowance.field,
                        allowance.condition,
                        allowance.value.lower().replace(" ", "_"),
                    )
                )
                for field, operator, value in conditions:
                    val = dynamic_attr(employee, field)
                    if val is None or not operator_mapping.get(operator)(
                        val, type(val)(value)
                    ):
                        applicable = False
                        break
            if applicable and allowance not in employee_allowances:
                employee_allowances.append(allowance)

        employee_allowances = [
            allowance
            for allowance in employee_allowances
            if operator_mapping.get(allowance.if_condition)(
                basic_pay if allowance.if_choice == "basic_pay" else 0,
                allowance.if_amount,
            )
        ]

        # Find the applicable deductions for the employee
        deductions = (
            Deduction.objects.filter(
                specific_employees=employee,
            )
            | Deduction.objects.filter(
                is_condition_based=True,
            ).exclude(exclude_employees=employee)
            | Deduction.objects.filter(
                include_active_employees=True,
            ).exclude(exclude_employees=employee)
        )
        for deduction in deductions:
            applicable = True
            if deduction.is_condition_based:
                conditions = list(
                    deduction.other_conditions.values_list(
                        "field", "condition", "value"
                    )
                )
                conditions.append(
                    (
                        deduction.field,
                        deduction.condition,
                        deduction.value.lower().replace(" ", "_"),
                    )
                )
                for field, operator, value in conditions:
                    val = dynamic_attr(employee, field)
                    if val is None or not operator_mapping.get(operator)(
                        val, type(val)(value)
                    ):
                        applicable = False
                        break
            if applicable:
                employee_deductions.append(deduction)

    allowance_ids = (
        json.dumps([instance.id for instance in employee_deductions])
        if employee_deductions
        else None
    )
    deduction_ids = (
        json.dumps([instance.id for instance in employee_deductions])
        if employee_deductions
        else None
    )
    context = {
        "active_contracts": active_contracts,
        "basic_pay": basic_pay,
        "allowances": employee_allowances if employee_allowances else None,
        "allowance_ids": allowance_ids,
        "deductions": employee_deductions if employee_deductions else None,
        "deduction_ids": deduction_ids,
        "employee": employee,
    }
    return render(request, "tabs/allowance_deduction-tab.html", context=context)


@login_required
@permission_required("payroll.add_allowance")
def create_allowance(request):
    """
    This method is used to create allowance condition template
    """
    form = forms.AllowanceForm()
    if request.method == "POST":
        form = forms.AllowanceForm(request.POST)
        if form.is_valid():
            form.save()
            form = forms.AllowanceForm()
            messages.success(request, _("Allowance created."))
            return redirect(view_allowance)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@permission_required("payroll.view_allowance")
def view_allowance(request):
    """
    This method is used render template to view all the allowance instances
    """
    allowances = Allowance.objects.exclude(only_show_under_employee=True)
    allowance_filter = AllowanceFilter(request.GET)
    allowances = paginator_qry(allowances, request.GET.get("page"))
    allowance_ids = json.dumps([instance.id for instance in allowances.object_list])
    return render(
        request,
        "payroll/allowance/view_allowance.html",
        {
            "allowances": allowances,
            "f": allowance_filter,
            "allowance_ids": allowance_ids,
        },
    )


@login_required
@hx_request_required
def view_single_allowance(request, allowance_id):
    """
    This method is used render template to view the selected allowance instances
    """
    previous_data = get_urlencode(request)
    allowance = Allowance.find(allowance_id)
    allowance_ids_json = request.GET.get("instances_ids")
    context = {
        "allowance": allowance,
    }
    if allowance_ids_json:
        allowance_ids = json.loads(allowance_ids_json)
        previous_id, next_id = closest_numbers(allowance_ids, allowance_id)
        context["next"] = next_id
        context["previous"] = previous_id
        context["allowance_ids"] = allowance_ids
    context["pd"] = previous_data
    return render(
        request,
        "payroll/allowance/view_single_allowance.html",
        context,
    )


@login_required
@hx_request_required
@permission_required("payroll.view_allowance")
def filter_allowance(request):
    """
    Filter and retrieve a list of allowances based on the provided query parameters.
    """
    query_string = request.GET.urlencode()
    allowances = AllowanceFilter(request.GET).qs.exclude(only_show_under_employee=True)
    list_view = "payroll/allowance/list_allowance.html"
    card_view = "payroll/allowance/card_allowance.html"
    template = card_view
    if request.GET.get("view") == "list":
        template = list_view
    allowances = sortby(request, allowances, "sortby")
    allowances = paginator_qry(allowances, request.GET.get("page"))
    allowance_ids = json.dumps([instance.id for instance in allowances.object_list])
    data_dict = parse_qs(query_string)
    get_key_instances(Allowance, data_dict)
    return render(
        request,
        template,
        {
            "allowances": allowances,
            "pd": query_string,
            "filter_dict": data_dict,
            "allowance_ids": allowance_ids,
        },
    )


@login_required
@permission_required("payroll.change_allowance")
def update_allowance(request, allowance_id, **kwargs):
    """
    This method is used to update the allowance
    Args:
        id : allowance instance id
    """
    instance = Allowance.objects.get(id=allowance_id)
    form = forms.AllowanceForm(instance=instance)
    if request.method == "POST":
        form = forms.AllowanceForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, _("Allowance updated."))
            return redirect(view_allowance)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@hx_request_required
@permission_required("payroll.delete_allowance")
def delete_allowance(request, allowance_id):
    """
    This method is used to delete the allowance instance
    """
    previous_data = get_urlencode(request)
    try:
        allowance = Allowance.objects.filter(id=allowance_id).first()
        if allowance:
            allowance.delete()
            messages.success(request, _("Allowance deleted successfully"))
        else:
            messages.error(request, _("Allowance not found"))
    except Exception as e:
        messages.error(request, _("An error occurred while deleting the allowance"))
        messages.error(request, str(e))

    if (
        request.path.split("/")[2] == "delete-employee-allowance"
        or not Allowance.objects.exists()
    ):
        return HttpResponse("<script>window.location.reload();</script>")

    instances_ids = request.GET.get("instances_ids")
    if instances_ids:
        instances_list = json.loads(instances_ids)
        previous_instance, next_instance = closest_numbers(instances_list, allowance_id)
        if allowance_id in instances_list:
            instances_list.remove(allowance_id)
            url = f"/payroll/single-allowance-view/{next_instance}"
            params = f"?{previous_data}&instances_ids={instances_list}"
            return redirect(url + params)

    return redirect(f"/payroll/filter-allowance?{previous_data}")


@login_required
@permission_required("payroll.add_deduction")
def create_deduction(request):
    """
    This method is used to create deduction
    """
    form = forms.DeductionForm()
    if request.method == "POST":
        form = forms.DeductionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _("Deduction created."))
            return redirect(view_deduction)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@permission_required("payroll.view_allowance")
def view_deduction(request):
    """
    This method is used render template to view all the deduction instances
    """

    deductions = Deduction.objects.exclude(only_show_under_employee=True)
    deduction_filter = DeductionFilter(request.GET)
    deductions = paginator_qry(deductions, request.GET.get("page"))
    deduction_ids = json.dumps([instance.id for instance in deductions.object_list])
    return render(
        request,
        "payroll/deduction/view_deduction.html",
        {
            "deductions": deductions,
            "f": deduction_filter,
            "deduction_ids": deduction_ids,
        },
    )


@login_required
@hx_request_required
def view_single_deduction(request, deduction_id):
    """
    Render template to view a single deduction instance with navigation.
    """
    previous_data = get_urlencode(request)
    deduction = Deduction.objects.filter(id=deduction_id).first()
    context = {"deduction": deduction, "pd": previous_data}

    # Handle deduction IDs and navigation
    deduction_ids_json = request.GET.get("instances_ids")
    if deduction_ids_json:
        deduction_ids = json.loads(deduction_ids_json)
        context["previous"], context["next"] = closest_numbers(
            deduction_ids, deduction_id
        )
        context["deduction_ids"] = deduction_ids

    # Determine htmx load URL and target
    HTTP_REFERER = request.META.get("HTTP_REFERER", "")
    referer_parts = HTTP_REFERER.rstrip("/").split("/")

    if "view-deduction" in referer_parts:
        context.update(
            {
                "load_hx_url": f"/payroll/filter-deduction?{previous_data}",
                "load_hx_target": "#payroll-deduction-container",
            }
        )
    elif referer_parts[-2:] == ["employee-view", str(referer_parts[-1])]:
        try:
            context.update(
                {
                    "load_hx_url": f"/payroll/allowances-deductions-tab/{int(referer_parts[-1])}",
                    "load_hx_target": "#allowance_deduction",
                }
            )
        except ValueError:
            pass
    elif HTTP_REFERER.endswith("employee-profile/"):
        context.update(
            {
                "load_hx_url": f"/payroll/allowances-deductions-tab/{request.user.employee_get.id}",
                "load_hx_target": "#allowance_deduction",
            }
        )
    else:
        context.update({"load_hx_url": None, "load_hx_target": None})

    return render(request, "payroll/deduction/view_single_deduction.html", context)


@login_required
@hx_request_required
@permission_required("payroll.view_allowance")
def filter_deduction(request):
    """
    This method is used search the deduction
    """
    query_string = request.GET.urlencode()
    deductions = DeductionFilter(request.GET).qs.exclude(only_show_under_employee=True)
    list_view = "payroll/deduction/list_deduction.html"
    card_view = "payroll/deduction/card_deduction.html"
    template = card_view
    if request.GET.get("view") == "list":
        template = list_view
    deductions = sortby(request, deductions, "sortby")
    deductions = paginator_qry(deductions, request.GET.get("page"))
    deduction_ids = json.dumps([instance.id for instance in deductions.object_list])
    data_dict = parse_qs(query_string)
    get_key_instances(Deduction, data_dict)
    return render(
        request,
        template,
        {
            "deductions": deductions,
            "pd": query_string,
            "filter_dict": data_dict,
            "deduction_ids": deduction_ids,
        },
    )


@login_required
@permission_required("payroll.change_deduction")
def update_deduction(request, deduction_id, **kwargs):
    """
    This method is used to update the deduction instance
    """
    instance = Deduction.objects.get(id=deduction_id)
    form = forms.DeductionForm(instance=instance)
    if request.method == "POST":
        form = forms.DeductionForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, _("Deduction updated."))
            return redirect(view_deduction)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@hx_request_required
@permission_required("payroll.delete_deduction")
def delete_deduction(request, deduction_id, emp_id=None):
    instances_ids = request.GET.get("instances_ids")
    next_instance = None
    instances_list = None
    previous_data = ""
    if instances_ids:
        previous_data = get_urlencode(request)
        instances_list = json.loads(instances_ids)
        previous_instance, next_instance = closest_numbers(instances_list, deduction_id)
        instances_list.remove(deduction_id)
    deduction = Deduction.objects.filter(id=deduction_id).first()
    if deduction:
        deduction.delete()
        messages.success(request, _("Deduction deleted successfully"))
    else:
        messages.error(request, _("Deduction not found"))

    paths = {
        "payroll-deduction-container": f"/payroll/filter-deduction?{request.GET.urlencode()}",
        "allowance_deduction": f"/payroll/allowances-deductions-tab/{emp_id}",
        "objectDetailsModalTarget": f"/payroll/single-deduction-view/{next_instance}?{previous_data}&instances_ids={instances_list}",
    }
    http_hx_target = request.META.get("HTTP_HX_TARGET")
    redirected_path = paths.get(http_hx_target)
    if http_hx_target:
        if (
            http_hx_target == "payroll-deduction-container"
            and not Deduction.objects.filter()
        ):
            return HttpResponse("<script>window.location.reload();</script>")
        if redirected_path:
            return redirect(redirected_path)
    default_redirect = (
        request.path if http_hx_target else request.META.get("HTTP_REFERER", "/")
    )
    return HttpResponseRedirect(default_redirect)


def get_month_start_end(year):
    start_end_dates = []
    for month in range(1, 13):
        # Start date is the first day of the month
        start_date = date(year, month, 1)

        # Calculate the last day of the month
        if month == 12:  # December
            end_date = date(year, 12, 31)
        else:
            next_month = date(year, month + 1, 1)
            end_date = next_month - timedelta(days=1)

        start_end_dates.append((start_date, end_date))
    return start_end_dates


@login_required
@permission_required("payroll.add_payslip")
def generate_payslip(request):
    """
    Generate payslips for selected employees within a specified date range.

    Requires the user to be logged in and have the 'payroll.add_payslip' permission.

    """
    if (
        request.META.get("HTTP_HX_REQUEST")
        and request.META.get("HTTP_HX_TARGET") == "objectCreateModalTarget"
    ):
        bulk_form = forms.GeneratePayslipForm()
        return render(
            request,
            "payroll/payslip/bulk_create_payslip.html",
            {"bulk_form": bulk_form},
        )
    payslips = []
    json_data = []
    form = forms.GeneratePayslipForm()
    if request.method == "POST":
        form = forms.GeneratePayslipForm(request.POST)
        if form.is_valid():
            instances = []
            employees = form.cleaned_data["employee_id"]
            start_date = form.cleaned_data["start_date"]
            end_date = form.cleaned_data["end_date"]

            group_name = form.cleaned_data["group_name"]
            for employee in employees:
                contract = Contract.objects.filter(
                    employee_id=employee, contract_status="active"
                ).first()
                if start_date < contract.contract_start_date:
                    start_date = contract.contract_start_date
                payslip = payroll_calculation(employee, start_date, end_date)
                payslips.append(payslip)
                json_data.append(payslip["json_data"])

                payslip["payslip"] = payslip
                data = {}
                data["employee"] = employee
                data["group_name"] = group_name
                data["start_date"] = payslip["start_date"]
                data["end_date"] = payslip["end_date"]
                data["status"] = "draft"
                data["contract_wage"] = payslip["contract_wage"]
                data["basic_pay"] = payslip["basic_pay"]
                data["gross_pay"] = payslip["gross_pay"]
                data["deduction"] = payslip["total_deductions"]
                data["net_pay"] = payslip["net_pay"]
                data["pay_data"] = json.loads(payslip["json_data"])
                calculate_employer_contribution(data)
                data["installments"] = payslip["installments"]
                instance = save_payslip(**data)
                instances.append(instance)
                notify.send(
                    request.user.employee_get,
                    recipient=employee.employee_user_id,
                    verb="Payslip has been generated for you.",
                    verb_ar="تم إصدار كشف راتب لك.",
                    verb_de="Gehaltsabrechnung wurde für Sie erstellt.",
                    verb_es="Se ha generado la nómina para usted.",
                    verb_fr="La fiche de paie a été générée pour vous.",
                    redirect=reverse(
                        "view-created-payslip", kwargs={"payslip_id": instance.id}
                    ),
                    icon="close",
                )
            messages.success(request, f"{employees.count()} payslip saved as draft")
            return redirect(
                f"/payroll/view-payslip?group_by=group_name&active_group={group_name}"
            )

    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@hx_request_required
def check_contract_start_date(request):
    """
    Check if the employee's contract start date is after the provided payslip start date.
    """
    employee_id = request.GET.get("employee_id")
    start_date = request.GET.get("start_date")

    contract = Contract.objects.filter(
        employee_id=employee_id, contract_status="active"
    ).first()

    if not contract or start_date >= str(contract.contract_start_date):
        return HttpResponse("")

    title_message = _(
        "When this payslip is run, the payslip start date will be updated to match the employee contract start date."
    )
    text_content = _("Employee Contract Start Date")

    return HttpResponse(
        format_html(
            """
        <div id='messageDiv' style='background-color: hsl(48, 100%, 94%);
            border: 1px solid hsl(46, 97%, 88%);
            border-radius: 18px; padding:5px; font-weight: bold; display: flex;'>
            {text_content}: {contract_start_date}
            <img style='width: 20px; height: 20px; cursor: pointer;'
                src='/static/images/ui/info.png' class='ml-2' title='{title_message}'>
        </div>
        """,
            text_content=text_content,
            contract_start_date=contract.contract_start_date,
            title_message=title_message,
        )
    )


@login_required
@permission_required("payroll.add_payslip")
def create_payslip(request, new_post_data=None):
    """
    Create a payslip for an employee.

    This method is used to create a payslip for an employee based on the provided form data.

    Args:
        request: The HTTP request object.

    Returns:
        A rendered HTML template for the payslip creation form.
    """
    if new_post_data:
        request.POST = new_post_data

    form = forms.PayslipForm()

    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        start_date = (
            datetime.strptime(request.POST.get("start_date"), "%Y-%m-%d").date()
            if isinstance(request.POST.get("start_date"), str)
            else request.POST.get("start_date")
        )

        if employee_id and start_date:
            contract = Contract.objects.filter(
                employee_id=employee_id, contract_status="active"
            ).first()

            if contract and start_date < contract.contract_start_date:
                new_post_data = request.POST.copy()
                new_post_data["start_date"] = contract.contract_start_date
                request.POST = new_post_data
        form = forms.PayslipForm(request.POST)
        if form.is_valid():
            employee = form.cleaned_data["employee_id"]
            start_date = form.cleaned_data["start_date"]
            end_date = form.cleaned_data["end_date"]
            payslip = Payslip.objects.filter(
                employee_id=employee, start_date=start_date, end_date=end_date
            ).first()

            if form.is_valid():
                employee = form.cleaned_data["employee_id"]
                start_date = form.cleaned_data["start_date"]
                end_date = form.cleaned_data["end_date"]
                payslip_data = payroll_calculation(employee, start_date, end_date)
                payslip_data["payslip"] = payslip
                data = {}
                data["employee"] = employee
                data["start_date"] = payslip_data["start_date"]
                data["end_date"] = payslip_data["end_date"]
                data["status"] = (
                    "draft"
                    if request.GET.get("status") is None
                    else request.GET["status"]
                )
                data["contract_wage"] = payslip_data["contract_wage"]
                data["basic_pay"] = payslip_data["basic_pay"]
                data["gross_pay"] = payslip_data["gross_pay"]
                data["deduction"] = payslip_data["total_deductions"]
                data["net_pay"] = payslip_data["net_pay"]
                data["pay_data"] = json.loads(payslip_data["json_data"])
                calculate_employer_contribution(data)
                data["installments"] = payslip_data["installments"]
                payslip_data["instance"] = save_payslip(**data)
                form = forms.PayslipForm()
                messages.success(request, _("Payslip Saved"))
                payslip = payslip_data["instance"]
                notify.send(
                    request.user.employee_get,
                    recipient=employee.employee_user_id,
                    verb="Payslip has been generated for you.",
                    verb_ar="تم إصدار كشف راتب لك.",
                    verb_de="Gehaltsabrechnung wurde für Sie erstellt.",
                    verb_es="Se ha generado la nómina para usted.",
                    verb_fr="La fiche de paie a été générée pour vous.",
                    redirect=reverse(
                        "view-created-payslip", kwargs={"payslip_id": payslip.pk}
                    ),
                    icon="close",
                )
                return HttpResponse(
                    f'<script>window.location.href = "/payroll/view-payslip/{payslip_data["instance"].id}/"</script>'
                )
    return render(
        request,
        "payroll/payslip/create_payslip.html",
        {"individual_form": form},
    )


@login_required
@permission_required("payroll.add_payslip")
def validate_start_date(request):
    """
    This method to validate the contract start date and the pay period start date
    """
    end_datetime = None
    start_datetime = None
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    employee_id = request.GET.getlist("employee_id")
    if start_date:
        start_datetime = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_datetime = datetime.strptime(end_date, "%Y-%m-%d").date()
    error_message = ""
    response = {"valid": True, "message": error_message}
    for emp_id in employee_id:
        contract = Contract.objects.filter(
            employee_id__id=emp_id, contract_status="active"
        ).first()

        if start_datetime is not None and start_datetime < contract.contract_start_date:
            error_message = f"<ul class='errorlist'><li>The {contract.employee_id}'s \
                contract start date is smaller than pay period start date</li></ul>"
            response["message"] = error_message
            response["valid"] = False

    if (
        start_datetime is not None
        and end_datetime is not None
        and start_datetime > end_datetime
    ):
        error_message = "<ul class='errorlist'><li>The end date must be greater than \
                or equal to the start date.</li></ul>"
        response["message"] = error_message
        response["valid"] = False

    if end_datetime is not None:
        if end_datetime > datetime.today().date():
            error_message = '<ul class="errorlist"><li>The end date cannot be in the future.</li></ul>'
            response["message"] = error_message
            response["valid"] = False
    return JsonResponse(response)


@login_required
@permission_required("payroll.view_payslip")
def view_individual_payslip(request, employee_id, start_date, end_date):
    """
    This method is used to render the template for viewing a payslip.
    """

    payslip_data = payroll_calculation(employee_id, start_date, end_date)
    return render(
        request,
        "payroll/payslip/individual_payslip.html",
        payslip_data,
    )


@login_required
@never_cache
def view_payslip(request):
    """
    This method is used to render the template for viewing a payslip.
    """
    if request.user.has_perm("payroll.view_payslip"):
        payslips = Payslip.objects.all()
    else:
        payslips = Payslip.objects.filter(employee_id__employee_user_id=request.user)
    export_column = forms.PayslipExportColumnForm()
    filter_form = PayslipFilter(request.GET, payslips)
    payslips = filter_form.qs
    bulk_form = forms.GeneratePayslipForm()
    field = request.GET.get("group_by")
    if field in Payslip.__dict__.keys():
        payslips = payslips.filter(group_name__isnull=False).order_by(field)
    payslips = paginator_qry(payslips, request.GET.get("page"))
    previous_data = request.GET.urlencode()
    data_dict = parse_qs(previous_data)
    get_key_instances(Payslip, data_dict)
    return render(
        request,
        "payroll/payslip/view_payslips.html",
        {
            "payslips": payslips,
            "f": filter_form,
            "export_column": export_column,
            "export_filter": PayslipFilter(request.GET),
            "bulk_form": bulk_form,
            "filter_dict": data_dict,
            "gp_fields": PayslipReGroup.fields,
        },
    )


@login_required
@hx_request_required
def filter_payslip(request):
    """
    Filter and retrieve a list of payslips based on the provided query parameters.
    """
    query_string = request.GET.urlencode()
    if request.user.has_perm("payroll.view_payslip"):
        payslips = PayslipFilter(request.GET).qs
    else:
        emp_request = request.GET.copy()
        employee = Employee.objects.filter(employee_user_id=request.user.id).first()
        employee_id = employee.id
        emp_request["employee_id"] = str(employee_id)
        payslips = PayslipFilter(emp_request).qs
    template = "payroll/payslip/payslip_table.html"
    view = request.GET.get("view")
    if view == "card":
        template = "payroll/payslip/group_payslips.html"
        payslips = payslips.filter(group_name__isnull=False).order_by("-group_name")
    payslips = sortby(request, payslips, "sortby")
    data_dict = []
    if not request.GET.get("dashboard"):
        data_dict = parse_qs(query_string)
        get_key_instances(Payslip, data_dict)
    if "status" in data_dict:
        status_list = data_dict["status"]
        if len(status_list) > 1:
            data_dict["status"] = [status_list[-1]]
    field = request.GET.get("field")
    if field != "" and field is not None:
        payslips = group_by_queryset(payslips, field, request.GET.get("page"), "page")
        template = "payroll/payslip/group_by.html"
    else:
        payslips = paginator_qry(payslips, request.GET.get("page"))
    return render(
        request,
        template,
        {
            "payslips": payslips,
            "pd": query_string,
            "filter_dict": data_dict,
        },
    )


@login_required
@permission_required("payroll.change_payslip")
def payslip_export(request):
    """
    This view exports payslip data based on selected fields and filters,
    and generates an Excel file for download.
    """
    if request.META.get("HTTP_HX_REQUEST"):
        return render(
            request,
            "payroll/payslip/payslip_export_filter.html",
            {
                "export_column": forms.PayslipExportColumnForm(),
                "export_filter": PayslipFilter(request.GET),
            },
        )

    choices_mapping = {
        "draft": _("Draft"),
        "review_ongoing": _("Review Ongoing"),
        "confirmed": _("Confirmed"),
        "paid": _("Paid"),
    }
    selected_columns = []
    payslips_data = {}
    payslips = PayslipFilter(request.GET).qs
    today_date = date.today().strftime("%Y-%m-%d")
    file_name = f"Payslip_excel_{today_date}.xlsx"
    selected_fields = request.GET.getlist("selected_fields")
    form = forms.PayslipExportColumnForm()

    if not selected_fields:
        selected_fields = form.fields["selected_fields"].initial
        ids = request.GET.get("ids")
        id_list = json.loads(ids)
        payslips = Payslip.objects.filter(id__in=id_list)

    for field in forms.excel_columns:
        value = field[0]
        key = field[1]
        if value in selected_fields:
            selected_columns.append((value, key))

    for column_value, column_name in selected_columns:
        nested_attributes = column_value.split("__")
        payslips_data[column_name] = []
        for payslip in payslips:
            value = payslip
            for attr in nested_attributes:
                value = getattr(value, attr, None)
                if value is None:
                    break
            data = str(value) if value is not None else ""
            if column_name == "Status":
                data = choices_mapping.get(value, "")

            if type(value) == date:
                date_format = request.user.employee_get.get_date_format()
                start_date = datetime.strptime(str(value), "%Y-%m-%d").date()

                for format_name, format_string in HORILLA_DATE_FORMATS.items():
                    if format_name == date_format:
                        data = start_date.strftime(format_string)
            else:
                data = str(value) if value is not None else ""
            payslips_data[column_name].append(data)

    data_frame = pd.DataFrame(data=payslips_data)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    writer = pd.ExcelWriter(response, engine="xlsxwriter")
    data_frame.style.applymap(lambda x: "text-align: center").to_excel(
        writer, index=False, sheet_name="Sheet1"
    )
    worksheet = writer.sheets["Sheet1"]
    worksheet.set_column("A:Z", 20)
    writer.close()
    return response


@login_required
@permission_required("payroll.add_allowance")
def hx_create_allowance(request):
    """
    This method is used to render htmx allowance form
    """
    form = forms.AllowanceForm()
    return render(request, "payroll/htmx/form.html", {"form": form})


@login_required
@permission_required("payroll.add_payslip")
def send_slip(request):
    """
    Send payslip method
    """
    email_backend = ConfiguredEmailBackend()
    view = request.GET.get("view")
    payslip_ids = request.GET.getlist("id")
    payslips = Payslip.objects.filter(id__in=payslip_ids)
    if not getattr(
        email_backend, "dynamic_from_email_with_display_name", None
    ) or not len(email_backend.dynamic_from_email_with_display_name):
        messages.error(request, "Email server is not configured")
        if view:
            return HttpResponse("<script>window.location.reload()</script>")
        else:
            return redirect(filter_payslip)

    result_dict = defaultdict(
        lambda: {"employee_id": None, "instances": [], "count": 0}
    )
    for payslip in payslips:
        employee_id = payslip.employee_id
        result_dict[employee_id]["employee_id"] = employee_id
        result_dict[employee_id]["instances"].append(payslip)
        result_dict[employee_id]["count"] += 1
    mail_thread = MailSendThread(request, result_dict=result_dict, ids=payslip_ids)
    mail_thread.start()
    messages.info(request, "Mail processing")
    if view:
        return HttpResponse("<script>window.location.reload()</script>")
    else:
        return redirect(filter_payslip)


@login_required
@permission_required("payroll.add_allowance")
def add_bonus(request):
    employee_id = request.GET["employee_id"]
    payslip_id = request.GET.get("payslip_id")
    if payslip_id != "None" and payslip_id:
        instance = Payslip.objects.get(id=payslip_id)
        form = forms.PayslipAllowanceForm(
            initial={"employee_id": employee_id, "date": instance.start_date}
        )
    else:
        form = forms.BonusForm(initial={"employee_id": employee_id})
    if request.method == "POST":
        form = forms.BonusForm(request.POST, initial={"employee_id": employee_id})
        contract = Contract.objects.filter(
            employee_id=employee_id, contract_status="active"
        ).first()
        employee = Employee.objects.filter(id=employee_id).first()
        if form.is_valid():
            form.save()
            messages.success(request, _("Bonus Added"))
            if payslip_id != "None" and payslip_id:
                if contract and contract.contract_start_date <= instance.start_date:

                    new_post_data = QueryDict(mutable=True)
                    new_post_data.update(
                        {
                            "employee_id": instance.employee_id,
                            "start_date": instance.start_date,
                            "end_date": instance.end_date,
                        }
                    )
                    instance.delete()
                    create_payslip(request, new_post_data)
                    payslip = Payslip.objects.filter(
                        employee_id=instance.employee_id,
                        start_date=instance.start_date,
                        end_date=instance.end_date,
                    ).first()
                    return HttpResponse(
                        f"<script>window.location.href='/payroll/view-payslip/{payslip.id}'</script>"
                    )
                else:
                    messages.warning(
                        request,
                        _(
                            "No active contract found for  {} during this payslip period"
                        ).format(employee),
                    )
            return HttpResponse("<script>window.location.reload()</script>")
    return render(
        request,
        "payroll/bonus/form.html",
        {"form": form, "employee_id": employee_id, "payslip_id": payslip_id},
    )


@login_required
@permission_required("payroll.add_deduction")
def add_deduction(request):
    employee_id = request.GET["employee_id"]
    payslip_id = request.GET.get("payslip_id")
    instance = Payslip.objects.get(id=payslip_id)

    if request.method == "POST":
        form = forms.PayslipDeductionForm(
            request.POST,
            initial={"employee_id": employee_id, "one_time_date": instance.start_date},
        )
        if form.is_valid():
            # Save the form to create the Deduction instance
            deduction_instance = form.save(commit=False)
            deduction_instance.only_show_under_employee = True
            deduction_instance.save()

            # Now that the Deduction instance is saved, add the related employees
            deduction_instance.specific_employees.set([employee_id])
            deduction_instance.include_active_employees = False
            deduction_instance.save()

            # Now create new payslip by deleting existing payslip
            new_post_data = QueryDict(mutable=True)
            new_post_data.update(
                {
                    "employee_id": instance.employee_id,
                    "start_date": instance.start_date,
                    "end_date": instance.end_date,
                }
            )
            instance.delete()
            create_payslip(request, new_post_data)
            payslip = Payslip.objects.filter(
                employee_id=instance.employee_id,
                start_date=instance.start_date,
                end_date=instance.end_date,
            ).first()
            return HttpResponse(
                f"<script>window.location.href='/payroll/view-payslip/{payslip.id}'</script>"
            )

    else:
        form = forms.PayslipDeductionForm(
            initial={"employee_id": employee_id, "one_time_date": instance.start_date}
        )

    return render(
        request,
        "payroll/deduction/payslip_deduct.html",
        {"form": form, "employee_id": employee_id, "payslip_id": payslip_id},
    )


@login_required
@permission_required("payroll.view_loanaccount")
def view_loans(request):
    """
    This method is used to render template to disply all the loan records
    """
    records = LoanAccount.objects.all()
    loan = records.filter(type="loan")
    adv_salary = records.filter(type="advanced_salary")
    fine = records.filter(type="fine")

    fine_ids = json.dumps(list(fine.values_list("id", flat=True)))
    loan_ids = json.dumps(list(loan.values_list("id", flat=True)))
    adv_salary_ids = json.dumps(list(adv_salary.values_list("id", flat=True)))
    loan = sortby(request, loan, "sortby")
    adv_salary = sortby(request, adv_salary, "sortby")
    fine = sortby(request, fine, "sortby")
    filter_instance = LoanAccountFilter()
    return render(
        request,
        "payroll/loan/view_loan.html",
        {
            "records": paginator_qry(records, request.GET.get("page")),
            "loan": paginator_qry(loan, request.GET.get("lpage")),
            "adv_salary": paginator_qry(adv_salary, request.GET.get("apage")),
            "fine_ids": fine_ids,
            "loan_ids": loan_ids,
            "adv_salary_ids": adv_salary_ids,
            "fine": paginator_qry(fine, request.GET.get("fpage")),
            "f": filter_instance,
        },
    )


@login_required
@hx_request_required
def create_loan(request):
    """
    This method is used to create and update the loan instance
    """
    instance_id = eval_validate(str(request.GET.get("instance_id")))
    instance = LoanAccount.objects.filter(id=instance_id).first()
    form = forms.LoanAccountForm(instance=instance)
    if request.method == "POST":
        form = forms.LoanAccountForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Loan created/updated")
            return HttpResponse("<script>window.location.reload()</script>")
    return render(
        request, "payroll/loan/form.html", {"form": form, "instance_id": instance_id}
    )


@login_required
@permission_required("payroll.view_loanaccount")
def view_installments(request):
    """
    View install ments
    """
    loan_id = request.GET["loan_id"]
    loan = LoanAccount.objects.get(id=loan_id)
    installments = loan.deduction_ids.all()

    requests_ids_json = request.GET.get("instances_ids")
    if requests_ids_json:
        requests_ids = json.loads(requests_ids_json)
        previous_id, next_id = closest_numbers(requests_ids, int(loan_id))
    return render(
        request,
        "payroll/loan/installments.html",
        {
            "installments": installments,
            "loan": loan,
            "instances_ids": requests_ids_json,
            "previous": previous_id,
            "next": next_id,
        },
    )


@login_required
@permission_required("payroll.delete_loanaccount")
def delete_loan(request):
    """
    Delete loan
    """
    ids = request.GET.getlist("ids")
    loans = LoanAccount.objects.filter(id__in=ids)
    # This 👇 would'nt trigger the delete method in the model
    # loans.delete()
    for loan in loans:
        if (
            not loan.settled
            and not Payslip.objects.filter(
                installment_ids__in=list(
                    loan.deduction_ids.values_list("id", flat=True)
                )
            ).exists()
        ):
            loan.delete()
            messages.success(request, "Loan account deleted")
        else:
            messages.error(request, "Loan account cannot be deleted")
    return redirect(view_loans)


@login_required
@permission_required("payroll.view_loanaccount")
def edit_installment_amount(request):
    loan_id = request.GET.get("loan_id")
    ded_id = request.GET.get("ded_id")
    value = float(request.POST.get("amount")) if request.POST.get("amount") else 0

    loan = LoanAccount.objects.filter(id=loan_id).first()
    deductions = loan.deduction_ids.all().order_by("one_time_date")
    deduction = deductions.filter(id=ded_id).first()
    deductions_before = deductions.filter(one_time_date__lt=deduction.one_time_date)
    deductions_after = deductions.filter(one_time_date__gt=deduction.one_time_date)
    total_sum = deductions_before.aggregate(Sum("amount"))["amount__sum"] or 0

    balance_instalment = len(deductions_after) if len(deductions_after) != 0 else 1

    new_installment = (loan.loan_amount - total_sum - value) / balance_instalment
    new_installment = round(new_installment, 2)
    if total_sum + value > loan.loan_amount:
        value = loan.loan_amount - total_sum
        new_installment = 0

    if not deduction.installment_payslip():
        deduction.amount = value
        deduction.save()

        for item in deductions.filter(one_time_date__gt=deduction.one_time_date):
            item.amount = new_installment
            item.save()

        if len(deductions_after) == 0 and new_installment != 0:
            date = get_next_month_same_date(deduction.one_time_date)
            installment = create_deductions(loan, new_installment, date)
            loan.deduction_ids.add(installment)

        messages.success(request, "Installment amount updated successfully")
    else:
        messages.error(request, "Cannot change paid installments ")

    return render(
        request,
        "payroll/loan/installments.html",
        {
            "installments": loan.deduction_ids.all(),
            "loan": loan,
        },
    )


@login_required
@hx_request_required
@permission_required("payroll.view_loanaccount")
def search_loan(request):
    """
    Search loan method
    """
    records = LoanAccountFilter(request.GET).qs
    loan = records.filter(type="loan")
    adv_salary = records.filter(type="advanced_salary")
    fine = records.filter(type="fine")

    fine_ids = json.dumps(list(fine.values_list("id", flat=True)))
    loan_ids = json.dumps(list(loan.values_list("id", flat=True)))
    adv_salary_ids = json.dumps(list(adv_salary.values_list("id", flat=True)))
    loan = sortby(request, loan, "sortby")
    adv_salary = sortby(request, adv_salary, "sortby")
    fine = sortby(request, fine, "sortby")

    data_dict = parse_qs(request.GET.urlencode())
    get_key_instances(LoanAccount, data_dict)
    view = request.GET.get("view")
    template = "payroll/loan/records_card.html"
    if view == "list":
        template = "payroll/loan/records_list.html"
    return render(
        request,
        template,
        {
            "records": paginator_qry(records, request.GET.get("page")),
            "loan": paginator_qry(loan, request.GET.get("lpage")),
            "adv_salary": paginator_qry(adv_salary, request.GET.get("apage")),
            "fine": paginator_qry(fine, request.GET.get("fpage")),
            "fine_ids": fine_ids,
            "loan_ids": loan_ids,
            "adv_salary_ids": adv_salary_ids,
            "filter_dict": data_dict,
            "pd": request.GET.urlencode(),
        },
    )


@login_required
@permission_required("payroll.add_loanaccount")
def asset_fine(request):
    """
    Add asset fine method
    """
    if apps.is_installed("asset"):
        Asset = get_horilla_model_class(app_label="asset", model="asset")
    asset_id = request.GET["asset_id"]
    employee_id = request.GET["employee_id"]
    asset = Asset.objects.get(id=asset_id)
    employee = Employee.objects.get(id=employee_id)
    form = forms.AssetFineForm()
    if request.method == "POST":
        form = forms.AssetFineForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.employee_id = employee
            instance.type = "fine"
            instance.provided_date = date.today()
            instance.asset_id = asset
            instance.save()
            messages.success(request, _("Asset fine added"))
            return HttpResponse(
                "<script>$('#dynamicCreateModal').toggleClass('oh-modal--show'); $('#reloadMessagesButton').click();</script>"
            )  # 880
    return render(
        request,
        "payroll/asset_fine/form.html",
        {"form": form, "asset_id": asset_id, "employee_id": employee_id},
    )


@login_required
def view_reimbursement(request):
    """
    This method is used to render template to view reimbursements
    """
    reimbursement_exists = False
    if Reimbursement.objects.exists():
        reimbursement_exists = True
    if request.GET:
        filter_object = ReimbursementFilter(request.GET)
    else:
        filter_object = ReimbursementFilter({"status": "requested"})
    requests = filter_own_records(
        request, filter_object.qs, "payroll.view_reimbursement"
    )
    reimbursements = requests.filter(type="reimbursement")
    leave_encashments = requests.filter(type="leave_encashment")
    bonus_encashment = requests.filter(type="bonus_encashment")
    data_dict = {"status": ["requested"]}
    view = request.GET.get("view")
    template = "payroll/reimbursement/view_reimbursement.html"

    return render(
        request,
        template,
        {
            "requests": paginator_qry(requests, request.GET.get("page")),
            "reimbursements": paginator_qry(reimbursements, request.GET.get("rpage")),
            "leave_encashments": paginator_qry(
                leave_encashments, request.GET.get("lpage")
            ),
            "bonus_encashments": paginator_qry(
                bonus_encashment, request.GET.get("bpage")
            ),
            "f": filter_object,
            "pd": request.GET.urlencode(),
            "filter_dict": data_dict,
            "view": view,
            "reimbursement_exists": reimbursement_exists,
        },
    )


@login_required
@hx_request_required
def create_reimbursement(request):
    """
    Create or update a reimbursement entry.
    """
    instance = None
    instance_id = request.GET.get("instance_id")

    if instance_id:
        instance = Reimbursement.objects.filter(id=instance_id).first()

    if request.method == "POST":
        form = forms.ReimbursementForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Reimbursement saved successfully")
            return HttpResponse(status=204, headers={"HX-Refresh": "true"})
    else:
        form = forms.ReimbursementForm(instance=instance)

    return render(request, "payroll/reimbursement/form.html", {"form": form})


@login_required
@hx_request_required
def search_reimbursement(request):
    """
    This method is used to search/filter reimbursement
    """
    requests = ReimbursementFilter(request.GET).qs
    requests = filter_own_records(request, requests, "payroll.view_reimbursement")
    data_dict = parse_qs(request.GET.urlencode())
    reimbursements = requests.filter(type="reimbursement")
    leave_encashments = requests.filter(type="leave_encashment")
    bonus_encashment = requests.filter(type="bonus_encashment")
    reimbursements_ids = json.dumps(list(reimbursements.values_list("id", flat=True)))
    leave_encashments_ids = json.dumps(
        list(leave_encashments.values_list("id", flat=True))
    )
    bonus_encashment_ids = json.dumps(
        list(bonus_encashment.values_list("id", flat=True))
    )
    reimbursements = sortby(request, reimbursements, "sortby")
    leave_encashments = sortby(request, leave_encashments, "sortby")
    bonus_encashment = sortby(request, bonus_encashment, "sortby")
    view = request.GET.get("view")
    template = "payroll/reimbursement/request_cards.html"
    if view == "list":
        template = "payroll/reimbursement/reimbursement_list.html"
    get_key_instances(Reimbursement, data_dict)

    return render(
        request,
        template,
        {
            "requests": paginator_qry(requests, request.GET.get("page")),
            "reimbursements": paginator_qry(reimbursements, request.GET.get("rpage")),
            "leave_encashments": paginator_qry(
                leave_encashments, request.GET.get("lpage")
            ),
            "bonus_encashments": paginator_qry(
                bonus_encashment, request.GET.get("bpage")
            ),
            "filter_dict": data_dict,
            "pd": request.GET.urlencode(),
            "reimbursements_ids": reimbursements_ids,
            "leave_encashments_ids": leave_encashments_ids,
            "bonus_encashment_ids": bonus_encashment_ids,
        },
    )


@login_required
def get_assigned_leaves(request):
    """
    This method is used to return assigned leaves of the employee
    in Json
    """
    if apps.is_installed("leave"):
        AvailableLeave = get_horilla_model_class(
            app_label="leave", model="availableleave"
        )

    assigned_leaves = (
        AvailableLeave.objects.filter(
            employee_id__id=request.GET["employeeId"],
            total_leave_days__gte=1,
            leave_type_id__is_encashable=True,
        )
        .values(
            "leave_type_id__name",
            "available_days",
            "carryforward_days",
            "leave_type_id__id",
        )
        .distinct()
    )
    return JsonResponse(list(assigned_leaves), safe=False)


@login_required
@permission_required("payroll.change_reimbursement")
def approve_reimbursements(request):
    """
    This method is used to approve or reject the reimbursement request
    """
    ids = request.GET.getlist("ids")
    status = request.GET["status"]
    if status == "canceled":
        status = "rejected"
    amount = (
        eval_validate(request.GET.get("amount")) if request.GET.get("amount") else 0
    )
    amount = max(0, amount)
    reimbursements = Reimbursement.objects.filter(id__in=ids)
    if status and len(status):
        for reimbursement in reimbursements:
            if reimbursement.type == "leave_encashment":
                reimbursement.amount = amount
            elif reimbursement.type == "bonus_encashment":
                reimbursement.amount = amount

            emp = reimbursement.employee_id
            reimbursement.status = status
            reimbursement.save()
            if reimbursement.status == "requested":
                if not (messages.get_messages(request)._queued_messages):
                    messages.info(request, _("Please check the data you provided."))
            else:
                messages.success(
                    request,
                    _(f"Request {reimbursement.get_status_display()} successfully"),
                )
        if status == "rejected":
            notify.send(
                request.user.employee_get,
                recipient=emp.employee_user_id,
                verb="Your reimbursement request has been rejected.",
                verb_ar="تم رفض طلب استرداد النفقات الخاص بك.",
                verb_de="Ihr Erstattungsantrag wurde abgelehnt.",
                verb_es="Su solicitud de reembolso ha sido rechazada.",
                verb_fr="Votre demande de remboursement a été rejetée.",
                redirect=reverse("view-reimbursement") + f"?id={reimbursement.id}",
                icon="checkmark",
            )
        else:
            notify.send(
                request.user.employee_get,
                recipient=emp.employee_user_id,
                verb="Your reimbursement request has been approved.",
                verb_ar="تمت الموافقة على طلب استرداد نفقاتك.",
                verb_de="Ihr Rückerstattungsantrag wurde genehmigt.",
                verb_es="Se ha aprobado tu solicitud de reembolso.",
                verb_fr="Votre demande de remboursement a été approuvée.",
                redirect=reverse("view-reimbursement") + f"?id={reimbursement.id}",
                icon="checkmark",
            )
    return redirect(view_reimbursement)


@login_required
@permission_required("payroll.delete_reimbursement")
def delete_reimbursements(request):
    """
    This method is used to delete the reimbursements
    """
    ids = request.GET.getlist("ids")
    reimbursements = Reimbursement.objects.filter(id__in=ids)
    for reimbursement in reimbursements:
        user = reimbursement.employee_id.employee_user_id
    reimbursements.delete()
    messages.success(request, "Reimbursements deleted")
    notify.send(
        request.user.employee_get,
        recipient=user,
        verb="Your reimbursement request has been deleted.",
        verb_ar="تم حذف طلب استرداد نفقاتك.",
        verb_de="Ihr Rückerstattungsantrag wurde gelöscht.",
        verb_es="Tu solicitud de reembolso ha sido eliminada.",
        verb_fr="Votre demande de remboursement a été supprimée.",
        redirect="/",
        icon="trash",
    )

    return redirect(view_reimbursement)


@login_required
@owner_can_enter("payroll.view_reimbursement", Reimbursement, True)
def reimbursement_individual_view(request, instance_id):
    """
    This method is used to render the individual view of reimbursement object
    """
    reimbursement = Reimbursement.objects.get(id=instance_id)
    requests_ids_json = request.GET.get("instances_ids")
    if requests_ids_json:
        requests_ids = json.loads(requests_ids_json)
        previous_id, next_id = closest_numbers(requests_ids, instance_id)
    context = {
        "reimbursement": reimbursement,
        "instances_ids": requests_ids_json,
        "previous": previous_id,
        "next": next_id,
    }
    return render(
        request,
        "payroll/reimbursement/reimbursenent_individual.html",
        context,
    )


@login_required
@owner_can_enter("payroll.view_reimbursement", Reimbursement, True)
def reimbursement_attachments(request, instance_id):
    """
    This method is used to render all the attachements under the reimbursement object
    """
    reimbursement = Reimbursement.objects.get(id=instance_id)
    return render(
        request,
        "payroll/reimbursement/attachments.html",
        {"reimbursement": reimbursement},
    )


@login_required
@owner_can_enter("payroll.delete_reimbursement", Reimbursement, True)
def delete_attachments(request, _reimbursement_id):
    """
    This mehtod is used to delete the attachements
    """
    ids = request.GET.getlist("ids")
    ReimbursementMultipleAttachment.objects.filter(id__in=ids).delete()
    messages.success(request, "Attachment deleted")
    return redirect(view_reimbursement)


@login_required
@permission_required("payroll.view_payslip")
def get_contribution_report(request):
    """
    This method is used to get the contribution report
    """
    employee_id = request.GET.get("employee_id")
    contribution_deductions = []
    if employee_id:
        pay_heads = Payslip.objects.filter(employee_id__id=employee_id).values_list(
            "pay_head_data", flat=True
        )
        deductions = []
        for head in pay_heads:
            for deduction in head["gross_pay_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["basic_pay_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["pretax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["post_tax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["tax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["net_deductions"]:
                deductions.append(deduction)

        deductions.sort(key=lambda x: x["deduction_id"])
        grouped_deductions = {
            key: list(group)
            for key, group in groupby(deductions, key=lambda x: x["deduction_id"])
        }

        for deduction_id, group in grouped_deductions.items():
            title = group[0]["title"]
            employee_contribution = sum(item.get("amount", 0) for item in group)
            employer_contribution = sum(
                item.get("employer_contribution_amount", 0) for item in group
            )
            total_contribution = employee_contribution + employer_contribution
            if employer_contribution > 0:
                contribution_deductions.append(
                    {
                        "deduction_id": deduction_id,
                        "title": title,
                        "employee_contribution": employee_contribution,
                        "employer_contribution": employer_contribution,
                        "total_contribution": total_contribution,
                    }
                )
    return render(
        request,
        "payroll/dashboard/contribution.html",
        {"contribution_deductions": contribution_deductions},
    )


def all_deductions(pay_head):

    extracted_items = []

    potential_lists = [
        "basic_pay_deductions",
        "gross_pay_deductions",
        "pretax_deductions",
        "post_tax_deductions",
        "tax_deductions",
        "net_deductions",
    ]

    for list_name in potential_lists:
        if list_name in pay_head.keys():
            for item in pay_head[list_name]:
                if "deduction_id" in item:
                    extracted_items.append(item)

    return extracted_items


@login_required
def payslip_detailed_export_data(request):
    """
    This view create the data for exporting payslip data based on selected fields and filters,
    """
    choices_mapping = {
        "draft": _("Draft"),
        "review_ongoing": _("Review Ongoing"),
        "confirmed": _("Confirmed"),
        "paid": _("Paid"),
    }
    selected_columns = []
    payslips_data = []
    totals = {}
    payslips = PayslipFilter(request.GET).qs
    selected_fields = request.GET.getlist("selected_fields")
    form = forms.PayslipExportColumnForm()

    allowances = Allowance.objects.all()
    deductions = Deduction.objects.all()

    if not selected_fields:
        selected_fields = form.fields["selected_fields"].initial

    for field in forms.excel_columns:
        value, key = field

        if value in selected_fields:
            selected_columns.append((value, key))

    selected_columns += [
        (value.title, value.title)
        for value in allowances.filter(
            one_time_date__isnull=True, include_active_employees=True
        )
    ]
    selected_columns += [
        ("other_allowances", "Other Allowances"),
        ("total_allowances", "Total Allowances"),
    ]

    selected_columns += [
        (value.title, value.title)
        for value in deductions.filter(
            one_time_date__isnull=True,
            include_active_employees=True,
            update_compensation__isnull=True,
        )
    ]
    selected_columns += [
        ("federal_tax", "Federal Tax"),
        ("other_deductions", "Other Deductions"),
        ("total_deductions", "Total Deductions"),
    ]

    allowance_totals = {
        column_name.title: 0
        for column_name in allowances.filter(
            one_time_date__isnull=True,
            include_active_employees=True,
        )
    }

    deduction_totals = {
        column_name.title: 0
        for column_name in deductions.filter(
            one_time_date__isnull=True,
            include_active_employees=True,
            update_compensation__isnull=True,
        )
    }

    other_totals = {
        "Other Allowances": 0,
        "Other Deductions": 0,
        "Total Allowances": 0,
        "Total Deductions": 0,
        "Net Pay": 0,
        "Gross Pay": 0,
        "Federal Tax": 0,
    }

    totals.update(allowance_totals)
    totals.update(deduction_totals)
    totals.update(other_totals)
    for payslip in payslips:
        payslip_data = {}
        other_allowances_sum = 0
        other_deductions_sum = 0
        total_allowance = 0
        total_deduction = 0
        total_federal_tax = 0

        federal_tax = payslip.pay_head_data["federal_tax"]
        total_federal_tax += federal_tax

        allos = payslip.pay_head_data["allowances"]
        deducts = all_deductions(payslip.pay_head_data)

        if allos:
            for allowance in allos:
                if not any(
                    str(allowance["title"]) == str(column_name)
                    for item, column_name in selected_columns
                ):
                    other_allowances_sum += (
                        allowance["amount"] if allowance["amount"] is not None else 0
                    )
                total_allowance += allowance["amount"]

        if deducts:
            for deduction in deducts:
                if not any(
                    str(deduction["title"]) == str(column_name)
                    for item, column_name in selected_columns
                ):
                    other_deductions_sum += (
                        deduction["amount"] if deduction["amount"] is not None else 0
                    )
                total_deduction += deduction["amount"]

        for column_value, column_name in selected_columns:
            nested_attributes = column_value.split("__")
            value = payslip
            for attr in nested_attributes:
                value = getattr(value, attr, None)
                if value is None:
                    break
            data = str(value) if value is not None else ""
            if column_name == "Status":
                data = choices_mapping.get(value, "")

            if isinstance(value, date):
                date_format = request.user.employee_get.get_date_format()
                start_date = datetime.strptime(str(value), "%Y-%m-%d").date()

                for format_name, format_string in HORILLA_DATE_FORMATS.items():
                    if format_name == date_format:
                        data = start_date.strftime(format_string)
            else:
                data = str(value) if value is not None else ""

            if allos:
                for allowance in allos:
                    if str(allowance["title"]) == str(column_name):
                        data = (
                            float(allowance["amount"])
                            if allowance["title"] is not None
                            else 0
                        )

            if deducts:
                for deduction in deducts:
                    if str(deduction["title"]) == str(column_name):
                        data = (
                            float(deduction["amount"])
                            if deduction["title"] is not None
                            else 0
                        )

            payslip_data[column_name] = data
            if column_name in totals:
                try:
                    totals[column_name] += float(data)
                except ValueError:
                    pass

        payslip_data["Other Allowances"] = other_allowances_sum
        payslip_data["Other Deductions"] = other_deductions_sum
        payslip_data["Total Allowances"] = total_allowance
        payslip_data["Total Deductions"] = total_deduction
        payslip_data["Federal Tax"] = federal_tax

        totals["Other Allowances"] += other_allowances_sum
        totals["Other Deductions"] += other_deductions_sum
        totals["Total Allowances"] += total_allowance
        totals["Total Deductions"] += total_deduction
        totals["Federal Tax"] += federal_tax

        payslips_data.append(payslip_data)

    totals_row = {}

    for item, column_name in selected_columns:
        if column_name in totals:
            totals_row[column_name] = totals[column_name]
        else:
            totals_row[column_name] = "-"

    totals_row["Other Allowances"] = totals["Other Allowances"]
    totals_row["Other Deductions"] = totals["Other Deductions"]
    totals_row["Total Allowances"] = totals["Total Allowances"]
    totals_row["Total Deductions"] = totals["Total Deductions"]
    totals_row["Employee"] = "Total"

    payslips_data.append(totals_row)

    return {
        "payslips_data": payslips_data,
        "selected_columns": selected_columns,
        "allowances": list(
            allowances.filter(
                one_time_date__isnull=True,
                include_active_employees=True,
            ).values_list("title", flat=True)
        ),
        "deductions": list(
            deductions.filter(
                one_time_date__isnull=True,
                include_active_employees=True,
                update_compensation__isnull=True,
            ).values_list("title", flat=True)
        ),
    }


@login_required
@permission_required("payroll.change_payslip")
def payslip_detailed_export(request):
    """
    Generate an Excel file for download containing detailed payslip data based on
    filters.

    Args:
        request (HttpRequest): The incoming HTTP request object.

    Returns:
        HttpResponse: A response object with the Excel file as an attachment.
    """

    if request.META.get("HTTP_HX_REQUEST"):
        return render(
            request,
            "payroll/payslip/payslip_export_filter.html",
            {
                "export_column": forms.PayslipExportColumnForm(),
                "export_filter": PayslipFilter(request.GET),
                "report": True,
            },
        )

    export_data = payslip_detailed_export_data(request)
    payslips_data = export_data["payslips_data"]
    selected_columns = export_data["selected_columns"]
    allowances = export_data["allowances"]
    deductions = export_data["deductions"]
    today_date = date.today().strftime("%Y-%m-%d")
    file_name = f"Payslip_excel_{today_date}.xlsx"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    right_border = Border(right=Side(style="thin"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Payslips"

    header_row = [col_name for _, col_name in selected_columns]
    allowances_header = allowances + ["Other Allowances", "Total Allowances"]
    deductions_header = deductions + [
        "Federal Tax",
        "Other Deductions",
        "Total Deductions",
    ]

    basic_cols = len(header_row) - len(allowances_header) - len(deductions_header)
    allowance_cols = len(allowances_header)
    deduction_cols = len(deductions_header)

    merged_sections = [
        (1, basic_cols, "Employee Details", "0000FF"),
        (basic_cols + 1, basic_cols + allowance_cols, "Allowances", "008000"),
        (
            basic_cols + allowance_cols + 1,
            basic_cols + allowance_cols + deduction_cols,
            "Deductions",
            "FF0000",
        ),
    ]

    bold_cols = [
        1,
        basic_cols + allowance_cols,
        basic_cols + allowance_cols + deduction_cols,
    ]

    for start_col, end_col, title, color in merged_sections:
        ws.merge_cells(
            start_row=1, start_column=start_col, end_row=1, end_column=end_col
        )
        cell = ws.cell(row=1, column=start_col, value=title)
        cell.font = Font(color=color, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

        if end_col <= len(header_row):
            ws.cell(row=1, column=end_col).border = thin_border + right_border
    last_row = len(payslips_data) + 2
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[last_row].height = 25

    subheaders = [
        (header_row[:basic_cols], Font(bold=True, color="0000FF")),
        (allowances_header, Font(bold=True, color="008000")),
        (deductions_header, Font(bold=True, color="FF0000")),
    ]

    col_num = 1
    for subheader, font in subheaders:
        for header in subheader:
            cell = ws.cell(row=2, column=col_num, value=str(header))
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            col_num += 1

    for row_num, payslip_data in enumerate(payslips_data, 3):
        for col_num, header in enumerate(header_row, 1):
            cell = ws.cell(
                row=row_num, column=col_num, value=payslip_data.get(header, "")
            )
            if row_num == last_row:
                cell.font = Font(bold=True, color="800080")
                cell.alignment = Alignment(horizontal="right")
            elif col_num in bold_cols:
                cell.font = Font(bold=True)
            cell.border = thin_border

    for col_num, _ in enumerate(header_row, 1):
        max_length = max(
            len(str(cell.value))
            for cell in ws[get_column_letter(col_num)]
            if cell.value is not None
        )
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    ws.freeze_panes = ws["B3"]

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={file_name}.xlsx"
    wb.save(response)

    return response
