import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path, PurePosixPath
from zipfile import BadZipFile, ZipFile

from django.core.exceptions import PermissionDenied, ValidationError
from django.core.validators import validate_email
from django.db import IntegrityError, transaction
from django.db.models.functions import Lower
from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import load_workbook

from hydra_imports.forms import MAX_IMPORT_BYTES
from hydra_imports.models import CandidateImportRow, CandidateImportSession
from hydra_people.models import Person
from hydra_people.recruitment_selectors import recruitments_for_user
from hydra_people.services import create_candidate_application, save_person
from recruitment.models import Candidate, Recruitment


MAX_IMPORT_ROWS = 500
MAX_ZIP_ENTRIES = 200
MAX_UNCOMPRESSED_BYTES = 20 * 1024 * 1024
WORKSHEET_NAME = "Candidates"
HEADERS = (
    "passport_name",
    "first_name",
    "last_name",
    "date_of_birth",
    "gender",
    "citizenship",
    "preferred_language",
    "email",
    "phone",
    "whatsapp_viber",
    "candidate_mobile",
)
IMPORT_PERMISSIONS = (
    "hydra_imports.view_candidateimportsession",
    "hydra_imports.import_candidate",
    "hydra_people.add_person",
    "hydra_people.view_person",
    "hydra_people.change_person",
    "hydra_people.link_candidate",
    "recruitment.add_candidate",
    "recruitment.view_candidate",
    "recruitment.view_recruitment",
)


@dataclass
class ParsedCandidateRow:
    row_number: int
    passport_name: str = ""
    first_name: str = ""
    last_name: str = ""
    date_of_birth: date | None = None
    gender: str = ""
    citizenship: str = ""
    preferred_language: str = ""
    email: str = ""
    phone: str = ""
    whatsapp_viber: str = ""
    candidate_mobile: str = ""
    errors: list[str] = field(default_factory=list)
    duplicate_reasons: list[str] = field(default_factory=list)

    @property
    def identity_key(self):
        if self.errors or not self.date_of_birth:
            return None
        return (
            _identity_name(self.passport_name),
            self.date_of_birth,
            self.citizenship,
        )

    @property
    def source_row_hash(self):
        payload = {
            header: (
                getattr(self, header).isoformat()
                if isinstance(getattr(self, header), date)
                else getattr(self, header)
            )
            for header in HEADERS
        }
        return hashlib.sha256(
            json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
        ).hexdigest()

    @property
    def outcome(self):
        if self.errors:
            return CandidateImportRow.Outcome.ERROR
        if self.duplicate_reasons:
            return CandidateImportRow.Outcome.DUPLICATE
        return CandidateImportRow.Outcome.VALID

    def as_model(self, *, session):
        return CandidateImportRow(
            session=session,
            row_number=self.row_number,
            outcome=self.outcome,
            error_message="; ".join(self.errors),
            duplicate_reason="; ".join(dict.fromkeys(self.duplicate_reasons)),
            source_row_hash=self.source_row_hash,
            **{header: getattr(self, header) for header in HEADERS},
        )


def _require_permissions(actor, permissions=IMPORT_PERMISSIONS):
    if not actor.is_authenticated or not actor.has_perms(permissions):
        raise PermissionDenied


def _validate_target(*, actor, recruitment, job_position):
    scoped = recruitments_for_user(
        user=actor,
        permission="view_recruitment",
    ).filter(pk=recruitment.pk, closed=False, is_active=True)
    if not scoped.exists():
        raise PermissionDenied
    if not recruitment.open_positions.filter(pk=job_position.pk).exists():
        raise ValidationError(
            {"job_position": _("Choose a position from this recruitment.")}
        )


def _safe_text(value, *, label, required=False, max_length=None, collapse=False):
    if value is None:
        text = ""
    elif isinstance(value, str):
        text = value.strip()
    elif isinstance(value, (int, float)) and label in {
        "phone",
        "whatsapp_viber",
        "candidate_mobile",
    }:
        text = str(value).strip()
        if text.endswith(".0"):
            text = text[:-2]
    else:
        raise ValueError(_("%(field)s must be text.") % {"field": label})
    if collapse:
        text = " ".join(text.split())
    if required and not text:
        raise ValueError(_("%(field)s is required.") % {"field": label})
    if max_length and len(text) > max_length:
        raise ValueError(
            _("%(field)s exceeds %(limit)s characters.")
            % {"field": label, "limit": max_length}
        )
    return text


def _safe_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError as error:
            raise ValueError(_("date_of_birth must use YYYY-MM-DD.")) from error
    raise ValueError(_("date_of_birth must be an Excel date or YYYY-MM-DD."))


def _validate_choice(value, *, field_name, choices):
    normalized = value.strip().lower()
    if normalized not in choices:
        raise ValueError(
            _("%(field)s must be one of: %(choices)s.")
            % {"field": field_name, "choices": ", ".join(sorted(choices))}
        )
    return normalized


def _validation_messages(error):
    if hasattr(error, "messages"):
        return [str(message) for message in error.messages]
    return [str(error)]


def _parse_row(row_number, values):
    row = ParsedCandidateRow(row_number=row_number)
    raw = dict(zip(HEADERS, values))

    specifications = (
        ("passport_name", True, 255, True),
        ("first_name", True, 100, True),
        ("last_name", True, 100, True),
        ("email", True, 254, False),
        ("phone", False, 25, False),
        ("whatsapp_viber", False, 25, False),
        ("candidate_mobile", False, 15, False),
    )
    for name, required, max_length, collapse in specifications:
        try:
            setattr(
                row,
                name,
                _safe_text(
                    raw[name],
                    label=name,
                    required=required,
                    max_length=max_length,
                    collapse=collapse,
                ),
            )
        except ValueError as error:
            row.errors.append(str(error))

    try:
        row.date_of_birth = _safe_date(raw["date_of_birth"])
    except ValueError as error:
        row.errors.append(str(error))

    try:
        row.gender = _validate_choice(
            _safe_text(raw["gender"], label="gender", required=True),
            field_name="gender",
            choices={choice[0] for choice in Person.Gender.choices},
        )
    except ValueError as error:
        row.errors.append(str(error))

    try:
        citizenship = _safe_text(
            raw["citizenship"],
            label="citizenship",
            required=True,
            max_length=2,
        ).upper()
        if not re.fullmatch(r"[A-Z]{2}", citizenship):
            raise ValueError(_("citizenship must be a two-letter ISO country code."))
        row.citizenship = citizenship
    except ValueError as error:
        row.errors.append(str(error))

    try:
        row.preferred_language = _validate_choice(
            _safe_text(
                raw["preferred_language"],
                label="preferred_language",
                required=True,
                max_length=3,
            ),
            field_name="preferred_language",
            choices={choice[0] for choice in Person.PreferredLanguage.choices},
        )
    except ValueError as error:
        row.errors.append(str(error))

    if row.email:
        row.email = row.email.lower()
        try:
            validate_email(row.email)
        except ValidationError as error:
            row.errors.extend(
                f"email: {message}" for message in _validation_messages(error)
            )

    person = Person(
        passport_name=row.passport_name,
        first_name=row.first_name,
        last_name=row.last_name,
        date_of_birth=row.date_of_birth,
        gender=row.gender or Person.Gender.UNSPECIFIED,
        citizenship=row.citizenship,
        preferred_language=row.preferred_language or Person.PreferredLanguage.POLISH,
        phone=row.phone,
        whatsapp_viber=row.whatsapp_viber,
        email=row.email,
    )
    try:
        person.full_clean(
            exclude=(
                "passport_name",
                "first_name",
                "last_name",
                "date_of_birth",
                "gender",
                "citizenship",
                "preferred_language",
                "email",
            ),
            validate_unique=False,
        )
    except ValidationError as error:
        row.errors.extend(_validation_messages(error))

    if row.candidate_mobile:
        mobile_field = Candidate._meta.get_field("mobile")
        try:
            mobile_field.clean(row.candidate_mobile, None)
        except ValidationError as error:
            row.errors.extend(
                f"candidate_mobile: {message}"
                for message in _validation_messages(error)
            )
    return row


def _validate_xlsx_container(content):
    if len(content) > MAX_IMPORT_BYTES:
        raise ValidationError(_("The workbook exceeds the 5 MB limit."))
    if not content.startswith(b"PK"):
        raise ValidationError(_("The uploaded file is not a valid .xlsx workbook."))
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ZIP_ENTRIES:
                raise ValidationError(_("The workbook contains too many internal files."))
            total_size = 0
            names = set()
            for entry in entries:
                path = PurePosixPath(entry.filename)
                if path.is_absolute() or ".." in path.parts or entry.flag_bits & 0x1:
                    raise ValidationError(_("The workbook container is not supported."))
                total_size += entry.file_size
                if total_size > MAX_UNCOMPRESSED_BYTES:
                    raise ValidationError(_("The expanded workbook exceeds the safety limit."))
                names.add(entry.filename)
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ValidationError(_("The uploaded file is not a valid .xlsx workbook."))
    except BadZipFile as error:
        raise ValidationError(_("The uploaded file is not a valid .xlsx workbook.")) from error


def parse_candidate_workbook(content):
    _validate_xlsx_container(content)
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise ValidationError(_("The uploaded file is not a readable .xlsx workbook.")) from error
    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise ValidationError(
                _("The workbook must contain a worksheet named Candidates.")
            )
        worksheet = workbook[WORKSHEET_NAME]
        if worksheet.max_row and worksheet.max_row > MAX_IMPORT_ROWS + 1:
            raise ValidationError(
                _("The Candidates worksheet exceeds the 500-row data area.")
            )
        if worksheet.max_column and worksheet.max_column > len(HEADERS):
            raise ValidationError(
                _("The Candidates worksheet contains columns outside the template.")
            )
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1), ())
        header_values = [
            cell.value.strip() if isinstance(cell.value, str) else cell.value
            for cell in first_row
        ]
        while header_values and header_values[-1] is None:
            header_values.pop()
        if tuple(header_values) != HEADERS:
            raise ValidationError(
                _("Candidate headers must exactly match the current Hydra template.")
            )

        parsed_rows = []
        for excel_row in worksheet.iter_rows(min_row=2):
            if any(cell.data_type == "f" for cell in excel_row):
                raise ValidationError(
                    _("Formulas are not accepted (row %(row)s).")
                    % {"row": excel_row[0].row}
                )
            values = [cell.value for cell in excel_row]
            if not any(value not in (None, "") for value in values):
                continue
            if any(value not in (None, "") for value in values[len(HEADERS) :]):
                raise ValidationError(
                    _("Unexpected data appears after the final column (row %(row)s).")
                    % {"row": excel_row[0].row}
                )
            if len(parsed_rows) >= MAX_IMPORT_ROWS:
                raise ValidationError(
                    _("A workbook can contain at most %(limit)s candidate rows.")
                    % {"limit": MAX_IMPORT_ROWS}
                )
            padded = (values[: len(HEADERS)] + [None] * len(HEADERS))[: len(HEADERS)]
            parsed_rows.append(_parse_row(excel_row[0].row, padded))
        if not parsed_rows:
            raise ValidationError(_("The Candidates worksheet contains no candidate rows."))
        return parsed_rows
    finally:
        workbook.close()


def _identity_name(value):
    return " ".join(value.split()).casefold()


def detect_duplicates(*, rows, recruitment):
    clean_rows = [row for row in rows if not row.errors]
    identity_counts = Counter(row.identity_key for row in clean_rows)
    email_counts = Counter(row.email for row in clean_rows)
    for row in clean_rows:
        if identity_counts[row.identity_key] > 1:
            row.duplicate_reasons.append(
                _("Duplicate identity appears more than once in this workbook.")
            )
        if email_counts[row.email] > 1:
            row.duplicate_reasons.append(
                _("Duplicate email appears more than once in this workbook.")
            )

    dates = {row.date_of_birth for row in clean_rows}
    countries = {row.citizenship for row in clean_rows}
    existing_identity_keys = {
        (
            _identity_name(person.passport_name),
            person.date_of_birth,
            person.citizenship.upper(),
        )
        for person in Person.objects.filter(
            date_of_birth__in=dates,
            citizenship__in=countries,
        ).only("passport_name", "date_of_birth", "citizenship")
    }
    existing_emails = set(
        Candidate._base_manager.filter(recruitment_id=recruitment)
        .annotate(normalized_email=Lower("email"))
        .filter(normalized_email__in={row.email for row in clean_rows})
        .values_list("normalized_email", flat=True)
    )
    for row in clean_rows:
        if row.identity_key in existing_identity_keys:
            row.duplicate_reasons.append(
                _("A Hydra Person with this passport identity already exists.")
            )
        if row.email in existing_emails:
            row.duplicate_reasons.append(
                _("This email already has an application in the selected recruitment.")
            )
    return rows


def _fingerprint(*, content, actor, recruitment, job_position):
    file_sha256 = hashlib.sha256(content).hexdigest()
    material = ":".join(
        (
            file_sha256,
            str(actor.pk),
            str(recruitment.pk),
            str(job_position.pk),
        )
    ).encode("ascii")
    return file_sha256, hashlib.sha256(material).hexdigest()


def preview_candidate_import(*, workbook, recruitment, job_position, actor):
    _require_permissions(actor)
    _validate_target(
        actor=actor,
        recruitment=recruitment,
        job_position=job_position,
    )
    content = workbook.read(MAX_IMPORT_BYTES + 1)
    if len(content) > MAX_IMPORT_BYTES:
        raise ValidationError(_("The workbook exceeds the 5 MB limit."))
    file_sha256, fingerprint = _fingerprint(
        content=content,
        actor=actor,
        recruitment=recruitment,
        job_position=job_position,
    )
    existing = CandidateImportSession.objects.filter(fingerprint=fingerprint).first()
    if existing:
        return existing

    rows = detect_duplicates(
        rows=parse_candidate_workbook(content),
        recruitment=recruitment,
    )
    error_count = sum(row.outcome == CandidateImportRow.Outcome.ERROR for row in rows)
    duplicate_count = sum(
        row.outcome == CandidateImportRow.Outcome.DUPLICATE for row in rows
    )
    valid_count = sum(row.outcome == CandidateImportRow.Outcome.VALID for row in rows)
    status = (
        CandidateImportSession.Status.READY
        if not error_count and not duplicate_count
        else CandidateImportSession.Status.BLOCKED
    )
    source_filename = Path(workbook.name).name[:255] or "candidates.xlsx"
    try:
        with transaction.atomic():
            session = CandidateImportSession.objects.create(
                recruitment=recruitment,
                job_position=job_position,
                source_filename=source_filename,
                file_sha256=file_sha256,
                fingerprint=fingerprint,
                status=status,
                row_count=len(rows),
                valid_count=valid_count,
                duplicate_count=duplicate_count,
                error_count=error_count,
                created_by=actor,
                modified_by=actor,
            )
            CandidateImportRow.objects.bulk_create(
                [row.as_model(session=session) for row in rows]
            )
            return session
    except IntegrityError:
        return CandidateImportSession.objects.get(fingerprint=fingerprint)


@transaction.atomic
def apply_candidate_import(*, session_uuid, actor):
    _require_permissions(actor)
    session = CandidateImportSession.objects.select_for_update().get(
        uuid=session_uuid
    )
    if not actor.is_superuser and session.created_by_id != actor.pk:
        raise PermissionDenied
    _validate_target(
        actor=actor,
        recruitment=session.recruitment,
        job_position=session.job_position,
    )
    if session.status == CandidateImportSession.Status.APPLIED:
        return session
    if session.status != CandidateImportSession.Status.READY:
        raise ValidationError(
            _("Resolve every validation error and duplicate before applying the import.")
        )

    rows = list(
        CandidateImportRow.objects.select_for_update()
        .filter(session=session)
        .order_by("row_number")
    )
    if len(rows) != session.valid_count or any(
        row.outcome != CandidateImportRow.Outcome.VALID for row in rows
    ):
        raise ValidationError(_("The stored preview is not ready to apply."))

    for row in rows:
        person = save_person(
            person=Person(
                passport_name=row.passport_name,
                first_name=row.first_name,
                last_name=row.last_name,
                date_of_birth=row.date_of_birth,
                gender=row.gender,
                citizenship=row.citizenship,
                preferred_language=row.preferred_language,
                phone=row.phone,
                whatsapp_viber=row.whatsapp_viber,
                email=row.email,
            ),
            actor=actor,
        )
        candidate, _link = create_candidate_application(
            person=person,
            candidate=Candidate(
                recruitment_id=session.recruitment,
                job_position_id=session.job_position,
                email=row.email,
                mobile=row.candidate_mobile,
                source="software",
                resume="",
            ),
            actor=actor,
        )
        row.created_person = person
        row.created_candidate = candidate
        row.save(update_fields=("created_person", "created_candidate"))

    session.status = CandidateImportSession.Status.APPLIED
    session.applied_at = timezone.now()
    session.applied_by = actor
    session.modified_by = actor
    session.save(
        update_fields=("status", "applied_at", "applied_by", "modified_by")
    )
    return session
