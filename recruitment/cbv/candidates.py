"""
This module used for recruitment candidates
"""

import ast
import io
import json
import re
from typing import Any

from bs4 import BeautifulSoup
from django import forms
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils.decorators import method_decorator
from django.utils.translation import gettext_lazy as _
from django.views import View
from import_export import fields, resources
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

from employee.forms import BulkUpdateFieldForm
from horilla.horilla_middlewares import _thread_locals
from horilla_views.cbv_methods import export_xlsx, login_required, permission_required
from horilla_views.forms import DynamicBulkUpdateForm
from horilla_views.generic.cbv.views import (
    HorillaCardView,
    HorillaDetailedView,
    HorillaFormView,
    HorillaListView,
    HorillaNavView,
    TemplateView,
)
from horilla_views.templatetags.generic_template_filters import getattribute
from recruitment.cbv.candidate_reject_reason import DynamicRejectReasonFormView
from recruitment.cbv_decorators import all_manager_can_enter, manager_can_enter
from recruitment.filters import CandidateFilter
from recruitment.forms import (
    CandidateExportForm,
    RejectedCandidateForm,
    ToSkillZoneForm,
)
from recruitment.models import (
    Candidate,
    RecruitmentSurvey,
    RecruitmentSurveyAnswer,
    RejectedCandidate,
    SkillZoneCandidate,
)

_getattribute = getattribute


def clean_column_name(question):
    """
    Convert the question text into a safe attribute name by:
    - Replacing spaces with underscores
    - Removing special characters except underscores
    """
    return re.sub(r"[^\w\s]", "", question).replace(" ", "_")


@method_decorator(
    permission_required(perm="recruitment.view_candidate"), name="dispatch"
)
@method_decorator(login_required, name="dispatch")
class CandidatesView(TemplateView):
    """
    For page view

    """

    template_name = "cbv/candidates/candidates.html"

    def get_context_data(self, **kwargs: Any) -> dict:
        context = super().get_context_data(**kwargs)
        update_fields = BulkUpdateFieldForm()
        context["update_fields_form"] = update_fields
        return context


@method_decorator(login_required, name="dispatch")
@method_decorator(manager_can_enter(perm="recruitment.view_candidate"), name="dispatch")
class ListCandidates(HorillaListView):
    """
    List view of candidates
    """

    model = Candidate
    filter_class = CandidateFilter
    bulk_template = "cbv/employees_view/bulk_update_page.html"
    bulk_update_fields = [
        "gender",
        "job_position_id",
        "hired_date",
        "referral",
        "country",
        "state",
        "city",
        "zip",
        "joining_date",
        "probation_end",
    ]

    def get_bulk_form(self):
        """
        Bulk from generating method
        """

        form = DynamicBulkUpdateForm(
            root_model=Candidate, bulk_update_fields=self.bulk_update_fields
        )

        form.fields["country"] = forms.ChoiceField(
            required=False,
            widget=forms.Select(
                attrs={
                    "class": "oh-select oh-select-2",
                    "required": False,
                    "style": "width: 100%; height:45px;",
                }
            ),
        )

        form.fields["state"] = forms.ChoiceField(
            required=False,
            widget=forms.Select(
                attrs={
                    "class": "oh-select oh-select-2",
                    "required": False,
                    "style": "width: 100%; height:45px;",
                },
            ),
        )

        return form

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        self.export_fields = []
        self.search_url = reverse("list-candidate")
        if self.request.user.has_perm("recruitment.change_candidate"):
            self.option_method = "options"
        else:
            self.option_method = None

        unique_questions = RecruitmentSurvey.objects.values_list(
            "question", flat=True
        ).distinct()
        self.survey_question_mapping = {}
        for question in unique_questions:
            survey_question = (question, f"question_{clean_column_name(question)}")
            self.survey_question_mapping[f"question_{clean_column_name(question)}"] = (
                question
            )
            if not survey_question in self.export_fields:
                self.export_fields.append(survey_question)

    columns = [
        ("Candidates", "name", "get_avatar"),
        ("Email", "email"),
        ("Phone", "mobile"),
        ("Rating", "rating"),
        ("Recruitment", "recruitment_id"),
        ("Job Position", "job_position_id"),
        ("Hired Date", "hired_date"),
        ("Resume", "resume_pdf"),
    ]
    default_columns = columns

    header_attrs = {
        "option": """
                   style ="width : 230px !important;"
                   """,
        "action": """
                   style ="width : 200px !important;"
                   """,
        "email": """
                   style ="width : 200px !important;"
                   """,
        "rating": """
                   style ="width : 170px !important;"
                   """,
    }

    actions = [
        {
            "action": "Edit",
            "icon": "create-outline",
            "attrs": """class="oh-btn oh-btn--light-bkg w-100"
            onclick="event.stopPropagation()
            window.location.href='{get_update_url}' "
             """,
        },
        {
            "action": "Archive",
            "accessibility": "recruitment.cbv.accessibility.archive_status",
            "icon": "archive",
            "attrs": """
                class="oh-btn oh-btn--danger-outline oh-btn--light-bkg w-100"
                onclick="event.stopPropagation()
                archiveCandidate({get_archive_url});  "
            """,
        },
        {
            "action": "Un-archive",
            "accessibility": "recruitment.cbv.accessibility.unarchive_status",
            "icon": "archive",
            "attrs": """
                class="oh-btn oh-btn--danger-outline oh-btn--light-bkg w-100"
                onclick="event.stopPropagation()
                archiveCandidate({get_archive_url});  "
            """,
        },
        {
            "action": _("Delete"),
            "icon": "trash-outline",
            "attrs": """
                    class="oh-btn oh-btn--danger-outline oh-btn--light-bkg w-100"
                    hx-get="{get_delete_url}?model=recruitment.candidate&pk={pk}"
                    data-toggle="oh-modal-toggle"
                    data-target="#deleteConfirmation"
                    hx-target="#deleteConfirmationBody"
                """,
        },
    ]

    sortby_mapping = [
        ("Candidates", "name", "get_avatar"),
    ]
    row_status_indications = [
        (
            "canceled--dot",
            "Canceled",
            """
            onclick="
                $('#applyFilter').closest('form').find('[name=canceled]').val('true');
                $('#applyFilter').click();
            "
            """,
        ),
        (
            "nothired--dot",
            "Not Hired",
            """
            onclick="
                $('#applyFilter').closest('form').find('[name=hired]').val('false');
                $('#applyFilter').click();
            "
            """,
        ),
        (
            "hired--dot",
            "Hired",
            """
            onclick="$('#applyFilter').closest('form').find('[name=hired]').val('true');
                $('#applyFilter').click();
            "
            """,
        ),
    ]

    records_per_page = 10

    row_status_class = "hired-{hired} canceled-{canceled}"

    # row_attrs = """
    #             {is_employee_converted}
    #             hx-get='{get_details_candidate}'
    #             data-toggle="oh-modal-toggle"
    #             data-target="#genericModal"
    #             hx-target="#genericModalBody"
    #             """
    row_attrs = """
                {is_employee_converted}
                onclick="window.location.href='{get_individual_url}?instance_ids={ordered_ids}'"
                """

    def export_data(self, *args, **kwargs):
        """
        Export with survey answer and question
        """

        request = getattr(_thread_locals, "request", None)
        ids = ast.literal_eval(request.POST["ids"])
        _columns = ast.literal_eval(request.POST["columns"])
        queryset = self.model.objects.filter(id__in=ids)
        question_mapping = self.survey_question_mapping
        export_format = request.POST.get("format", "xlsx")

        _model = self.model

        class HorillaListViewResorce(resources.ModelResource):
            """
            Instant Resource class
            """

            id = fields.Field(column_name="ID")
            question = {}

            class Meta:
                """
                Meta class for additional option
                """

                model = _model
                fields = [field[1] for field in _columns]  # 773

            def __init__(self, **kwargs):
                super().__init__(**kwargs)

                for field_tuple in _columns:
                    if field_tuple[1].startswith("question_"):
                        safe_field_name = field_tuple[1]
                        self.fields[safe_field_name] = fields.Field(
                            column_name=question_mapping[safe_field_name],
                            attribute=safe_field_name,
                            readonly=True,
                        )

            def export_field(self, field, obj):
                """
                Override this method to fetch the candidate's answers dynamically.
                """

                if field.attribute:
                    # Get the stored JSON field containing answers
                    survey_answers = RecruitmentSurveyAnswer.objects.filter(
                        candidate_id=obj
                    ).first()
                    if survey_answers and field.attribute.startswith("question_"):
                        survey_answers = survey_answers.answer_json
                        if isinstance(survey_answers, str):
                            try:
                                survey_answers = ast.literal_eval(
                                    survey_answers
                                )  # Convert string to dict
                            except Exception:
                                survey_answers = {}

                        # Extract the actual question text

                        original_question = question_mapping[field.attribute]
                        # Retrieve answer from JSON if available
                        answer = survey_answers.get(original_question, "")
                        if not answer:
                            answer = survey_answers.get(
                                "rating_" + original_question, ""
                            )
                        if not answer:
                            answer = survey_answers.get(
                                "percentage_" + original_question, ""
                            )
                        if not answer:
                            answer = survey_answers.get("file_" + original_question, "")
                        if not answer:
                            answer = survey_answers.get("date_" + original_question, "")
                        if not answer:
                            answer = survey_answers.get(
                                "multiple_choices_" + original_question, ""
                            )
                        return answer

                return super().export_field(field, obj)

            def dehydrate_id(self, instance):
                """
                Dehydrate method for id field
                """
                return instance.pk

            for field_tuple in _columns:
                if not field_tuple[1].startswith("question_"):
                    dynamic_fn_str = f"def dehydrate_{field_tuple[1]}(self, instance):return self.remove_extra_spaces(getattribute(instance, '{field_tuple[1]}'))"
                    exec(dynamic_fn_str)
                    dynamic_fn = locals()[f"dehydrate_{field_tuple[1]}"]
                    locals()[field_tuple[1]] = fields.Field(column_name=field_tuple[0])

            def remove_extra_spaces(self, text):
                """
                Remove blank space but keep line breaks and add new lines for <li> tags.
                """
                soup = BeautifulSoup(str(text), "html.parser")
                for li in soup.find_all("li"):
                    li.insert_before("\n")
                    li.unwrap()
                text = soup.get_text()
                lines = text.splitlines()
                non_blank_lines = [line.strip() for line in lines if line.strip()]
                cleaned_text = "\n".join(non_blank_lines)
                return cleaned_text

        book_resource = HorillaListViewResorce()

        # Export the data using the resource
        dataset = book_resource.export(queryset)

        # Set the response headers
        # file_name = self.export_file_name
        if export_format == "json":
            json_data = json.loads(dataset.export("json"))
            response = HttpResponse(
                json.dumps(json_data, indent=4), content_type="application/json"
            )
            response["Content-Disposition"] = (
                f'attachment; filename="{self.export_file_name}.json"'
            )
            return response

        # CSV
        elif export_format == "csv":
            csv_data = dataset.export("csv")
            response = HttpResponse(csv_data, content_type="text/csv")
            response["Content-Disposition"] = (
                f'attachment; filename="{self.export_file_name}.csv"'
            )
            return response
        elif export_format == "pdf":

            headers = dataset.headers
            rows = dataset.dict

            # Render to HTML using a template
            html_string = render_to_string(
                "generic/export_pdf.html",
                {
                    "headers": headers,
                    "rows": rows,
                },
            )

            # Convert HTML to PDF using xhtml2pdf
            result = io.BytesIO()
            pisa_status = pisa.CreatePDF(html_string, dest=result)

            if pisa_status.err:
                return HttpResponse("PDF generation failed", status=500)

            # Return response
            response = HttpResponse(result.getvalue(), content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="{self.export_file_name}.pdf"'
            )
            return response

        # response = HttpResponse(
        #     dataset.export("xlsx"), content_type="application/vnd.ms-excel"
        # )
        # response["Content-Disposition"] = (
        #     f'attachment; filename="{self.export_file_name}.xls"'
        # )
        json_data = json.loads(dataset.export("json"))
        headers = list(json_data[0].keys()) if json_data else []

        wb = Workbook()
        ws = wb.active
        ws.title = "Exported Data"

        # Styling
        header_fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        wrap_alignment = Alignment(vertical="top", wrap_text=True)

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Write data rows
        for row_idx, item in enumerate(json_data, start=2):
            for col_idx, key in enumerate(headers, start=1):
                value = item.get(key, "")
                # Convert lists to newline-separated string
                if isinstance(value, list):
                    value = "\n".join(str(v) for v in value)
                elif isinstance(value, dict):
                    value = json.dumps(
                        value, ensure_ascii=False
                    )  # or format it as needed
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = wrap_alignment

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 5, 50)

        # Output to Excel
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="exported_data.xlsx"'
        return response


@method_decorator(login_required, name="dispatch")
@method_decorator(manager_can_enter(perm="recruitment.view_candidate"), name="dispatch")
class CardCandidates(HorillaCardView):
    """
    For card view
    """

    model = Candidate
    filter_class = CandidateFilter

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        self.search_url = reverse("card-candidate")

    details = {
        "image_src": "get_avatar",
        "title": "{get_full_name}",
        "subtitle": "{email} <br> {get_job_position}",
    }

    actions = [
        {
            "action": "Convert to Employee",
            "accessibility": "recruitment.cbv.accessibility.convert_emp",
            "attrs": """
                onclick="event.stopPropagation()
                return confirm('Are you sure you want to convert this candidate into an employee?')"
                href='{get_convert_to_emp}'
                class="oh-dropdown__link"

            """,
        },
        {
            "action": "Add to Skill Zone",
            "accessibility": "recruitment.cbv.accessibility.add_skill_zone",
            "attrs": """
                data-toggle="oh-modal-toggle"
                data-target="#genericModal"
                hx-get="{get_add_to_skill}"
                hx-target="#genericModalBody"
                class="oh-dropdown__link"

            """,
        },
        {
            "action": "View candidate self tracking",
            "accessibility": "recruitment.cbv.accessibility.check_candidate_self_tracking",
            "attrs": """
                href="{get_self_tracking_url}"
                class="oh-dropdown__link"
            """,
        },
        {
            "action": "Request Document",
            "accessibility": "recruitment.cbv.accessibility.request_document",
            "attrs": """
                data-toggle="oh-modal-toggle"
                data-target="#genericModal"
                hx-get="{get_document_request_doc}"
                hx-target="#genericModalBody"
                class="oh-dropdown__link"
            """,
        },
        {
            "action": "Add to Rejected",
            "accessibility": "recruitment.cbv.accessibility.add_reject",
            "attrs": """
                hx-target="#genericModalBody"
                hx-swap="innerHTML"
                data-toggle="oh-modal-toggle"
                data-target="#genericModal"
                hx-get="{get_add_to_reject}"
                class="oh-dropdown__link"

            """,
        },
        {
            "action": "Edit Rejected Candidate",
            "accessibility": "recruitment.cbv.accessibility.edit_reject",
            "attrs": """
                hx-target="#genericModalBody"
                hx-swap="innerHTML"
                data-toggle="oh-modal-toggle"
                data-target="#genericModal"
                hx-get="{get_add_to_reject}"
                class="oh-dropdown__link"

            """,
        },
        {
            "action": "Edit Profile",
            "attrs": """
                onclick="event.stopPropagation()
                window.location.href='{get_update_url}' "
                class="oh-dropdown__link"

            """,
        },
        {
            "action": "archive_status",
            "attrs": """
                class="oh-dropdown__link"
                onclick="archiveCandidate({get_archive_url});"


            """,
        },
        {
            "action": "Delete",
            "attrs": """
                class="oh-dropdown__link oh-dropdown__link--danger"
                onclick="event.stopPropagation();
                deleteCandidate('{get_delete_url}'); "

            """,
        },
    ]
    card_status_indications = [
        (
            "canceled--dot",
            "Canceled",
            """
            onclick="
                $('#applyFilter').closest('form').find('[name=canceled]').val('true');
                $('#applyFilter').click();
            "
            """,
        ),
        (
            "nothired--dot",
            "Not Hired",
            """
            onclick="
                $('#applyFilter').closest('form').find('[name=hired]').val('false');
                $('#applyFilter').click();
            "
            """,
        ),
        (
            "hired--dot",
            "Hired",
            """
            onclick="$('#applyFilter').closest('form').find('[name=hired]').val('true');
                $('#applyFilter').click();
            "
            """,
        ),
    ]
    card_status_class = "hired-{hired} canceled-{canceled}"
    card_attrs = """
                onclick="window.location.href='{get_individual_url}?instance_ids={ordered_ids}'"
                """

    records_per_page = 30


@method_decorator(login_required, name="dispatch")
@method_decorator(manager_can_enter(perm="recruitment.view_candidate"), name="dispatch")
class CandidateNav(HorillaNavView):
    """
    For nav bar
    """

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        self.search_url = reverse("list-candidate")
        self.create_attrs = f"""
                            href='{reverse_lazy('candidate-create')}'"
                            """
        self.actions = [
            {
                "action": "Export",
                "attrs": f"""
                 data-toggle="oh-modal-toggle"
                 data-target="#genericModal"
                 hx-get="{reverse('export')}"
                 hx-target="#genericModalBody"
                 """,
            },
            {
                "action": "Bulk mail",
                "attrs": f"""
                data-toggle="oh-modal-toggle"
                data-target="#genericModal"
                hx-get="{reverse('send-mail')}"
                hx-target="#genericModalBody"
                """,
            },
            {
                "action": "Create document request",
                "attrs": f"""
                data-toggle="oh-modal-toggle"
                data-target="#objectCreateModal"
                hx-get="{reverse('candidate-document-request')}"
                hx-target="#objectCreateModalTarget"
                """,
            },
            {
                "action": "Archive",
                "attrs": """
                id="archiveCandidates"

                """,
            },
            {
                "action": "Un archive",
                "attrs": """
                id="unArchiveCandidates"

                """,
            },
            {
                "action": "Delete",
                "attrs": """
                data-action = "delete"
                id="deleteCandidates"
                 new_init
                """,
            },
        ]

        self.view_types = [
            {
                "type": "list",
                "icon": "list-outline",
                "url": reverse("list-candidate"),
                "attrs": """
                            title='List'
                            """,
            },
            {
                "type": "card",
                "icon": "grid-outline",
                "url": reverse("card-candidate"),
                "attrs": """
                            title='Card'
                            """,
            },
        ]
        self.filter_instance = CandidateFilter()

    nav_title = "Candidates"
    filter_body_template = "cbv/candidates/filter.html"
    filter_form_context_name = "form"
    search_swap_target = "#listContainer"
    group_by_fields = [
        ("recruitment_id", "Recruitment"),
        ("job_position_id", "Job Position"),
        ("hired", "Hired"),
        ("country", "Country"),
        ("stage_id", "Stage"),
        ("joining_date", "Date joining"),
        ("probation_end", "Probation End"),
        ("offer_letter_status", "offer Letter Status"),
        ("rejected_candidate__reject_reason_id", "Reject reason"),
        ("skillzonecandidate_set", "Skill zone"),
    ]


@method_decorator(login_required, name="dispatch")
@method_decorator(manager_can_enter(perm="recruitment.view_candidate"), name="dispatch")
class ExportView(TemplateView):
    """
    For candidate export
    """

    template_name = "cbv/candidates/export.html"

    def get_context_data(self, **kwargs: Any):
        """
        Adds export fields and filter object to the context.
        """
        context = super().get_context_data(**kwargs)
        candidates = Candidate.objects.filter(is_active=True)
        export_column = CandidateExportForm()
        export_filter = CandidateFilter(queryset=candidates)
        context["export_column"] = export_column
        context["export_filter"] = export_filter
        return context


@method_decorator(login_required, name="dispatch")
@method_decorator(manager_can_enter(perm="recruitment.view_candidate"), name="dispatch")
class AddToRejectedCandidatesView(View):
    """
    Class for Add to reject candidate
    """

    template_name = "onboarding/rejection/form.html"

    def get(self, request):
        """
        get method
        """
        candidate_id = request.GET.get("candidate_id")
        instance = None
        if candidate_id:
            instance = RejectedCandidate.objects.filter(
                candidate_id=candidate_id
            ).first()
        form = RejectedCandidateForm(
            initial={"candidate_id": candidate_id}, instance=instance
        )
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        """
        post method
        """
        candidate_id = request.GET.get("candidate_id")
        instance = None
        if candidate_id:
            instance = RejectedCandidate.objects.filter(
                candidate_id=candidate_id
            ).first()
        form = RejectedCandidateForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Candidate reject reason saved")
            return HttpResponse("<script>window.location.reload()</script>")
        return render(request, self.template_name, {"form": form})


@method_decorator(login_required, name="dispatch")
@method_decorator(
    all_manager_can_enter(perm="recruitment.view_candidate"), name="dispatch"
)
class CandidateDetail(HorillaDetailedView):
    """
    Candidate detail
    """

    title = "Candidate Details"

    model = Candidate

    header = {"title": "get_full_name", "subtitle": "get_email", "avatar": "get_avatar"}

    body = [
        ("Gender", "gender"),
        ("Phone", "mobile"),
        ("Stage", "stage_drop_down"),
        ("Rating", "rating_bar"),
        ("Recruitment", "recruitment_id"),
        ("Job Position", "job_position_id"),
        ("Interview Table", "candidate_interview_view", True),
    ]

    cols = {
        "candidate_interview_view": 12,
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.actions = [
            {
                "action": "Edit",
                "icon": "create-outline",
                "attrs": """
            class="oh-btn oh-btn--info w-50"
            onclick="window.location.href='{get_update_url}' "
            """,
            },
            {
                "action": "View",
                "icon": "eye-outline",
                "attrs": """
            class="oh-btn oh-btn--success w-50"
            onclick="window.location.href='{get_individual_url}'"
            """,
            },
        ]

        if self.request.user.has_perm("recruitment.delete_candidate"):
            self.actions.append(
                {
                    "action": "Delete",
                    "icon": "trash-outline",
                    "accessibility": "recruitment.cbv.candidates.delete_cand",
                    "attrs": f"""
            class="oh-btn oh-btn--danger w-50"
            hx-get="{reverse_lazy("generic-delete")}?model=recruitment.Candidate&pk={{pk}}"
            hx-target="#deleteConfirmationBody"
            data-toggle="oh-modal-toggle"
            data-target="#deleteConfirmation"
            onclick="event.stopPropagation()
            deleteCandidate('{{get_delete_url}}'); "
            """,
                }
            )


class ToSkillZoneFormView(HorillaFormView):
    """
    Form View
    """

    model = SkillZoneCandidate
    form_class = ToSkillZoneForm
    new_display_title = "Add To Skill Zone"

    def get_context_data(self, **kwargs):
        """
        Returns context with form and candidate data.
        """
        context = super().get_context_data(**kwargs)
        candidate_id = self.kwargs.get("cand_id")
        candidate = Candidate.objects.get(id=candidate_id)
        form = self.form_class(
            initial={
                "candidate_id": candidate,
                "skill_zone_ids": SkillZoneCandidate.objects.filter(
                    candidate_id=candidate
                ).values_list("skill_zone_id", flat=True),
            }
        )
        context["form"] = form
        return context

    def form_invalid(self, form: Any) -> HttpResponse:
        """
        Handles and renders form errors or defers to superclass.
        """
        form = self.form_class(self.request.POST)
        if not form.is_valid():
            errors = form.errors.as_data()
            return render(
                self.request, self.template_name, {"form": form, "errors": errors}
            )
        return super().form_invalid(form)

    def form_valid(self, form: ToSkillZoneForm) -> HttpResponse:
        """
        Handles valid form submission and saves rejected candidate reason.
        """
        if form.is_valid():
            candidate_id = self.kwargs.get("cand_id")
            candidate = Candidate.objects.get(id=candidate_id)
            self.form_class(
                initial={
                    "candidate_id": candidate,
                    "skill_zone_ids": SkillZoneCandidate.objects.filter(
                        candidate_id=candidate
                    ).values_list("skill_zone_id", flat=True),
                }
            )
            skill_zones = self.form.cleaned_data["skill_zone_ids"]
            for zone in skill_zones:
                if not SkillZoneCandidate.objects.filter(
                    candidate_id=candidate_id, skill_zone_id=zone
                ).exists():
                    zone_candidate = SkillZoneCandidate()
                    zone_candidate.candidate_id = candidate
                    zone_candidate.skill_zone_id = zone
                    zone_candidate.reason = self.form.cleaned_data["reason"]
                    zone_candidate.save()
            message = "Candidate Added to skill zone successfully"
            messages.success(self.request, _(message))
            return self.HttpResponse()
        return super().form_valid(form)


class RejectReasonFormView(HorillaFormView):
    """
    Form View
    """

    model = RejectedCandidate
    form_class = RejectedCandidateForm
    new_display_title = "Rejected Candidate"
    dynamic_create_fields = [("reject_reason_id", DynamicRejectReasonFormView)]
    template_name = "candidate/candidate_rejection_form.html"

    def get_initial(self) -> dict:
        initial = super().get_initial()
        initial["candidate_id"] = self.request.GET.get("candidate_id")
        return initial

    def init_form(self, *args, data={}, files={}, instance=None, **kwargs):
        candidate_id = self.request.GET.get("candidate_id")
        instance = RejectedCandidate.objects.filter(candidate_id=candidate_id).first()
        return super().init_form(
            *args, data=data, files=files, instance=instance, **kwargs
        )

    def form_valid(self, form: RejectedCandidateForm) -> HttpResponse:
        """
        Handles valid form submission and saves rejected candidate reason.
        """
        if form.is_valid():
            message = _("Candidate reject reason saved")
            messages.success(self.request, _(message))
            form.save()
            return self.HttpResponse()
        return super().form_valid(form)
