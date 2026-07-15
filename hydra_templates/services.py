import hashlib
from io import BytesIO

from django.core.exceptions import PermissionDenied, ValidationError
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from hydra_coordination.selectors import company_ids_for_user
from hydra_templates.models import MessageTemplate, TemplateDataExport
from hydra_templates.placeholders import PLACEHOLDERS, render_template_text, sample_values
from hydra_templates.selectors import export_people_for_user


EXPORT_PERMISSIONS = (
    "hydra_templates.export_template_data",
    "hydra_people.view_person",
)
MAX_EXPORT_ROWS = 10000


def save_message_template(*, template, actor):
    permission = (
        "hydra_templates.add_messagetemplate"
        if template._state.adding
        else "hydra_templates.change_messagetemplate"
    )
    if not actor.has_perm(permission):
        raise PermissionDenied
    if template.company_id not in company_ids_for_user(user=actor):
        raise PermissionDenied
    template.full_clean()
    if template._state.adding:
        template.created_by = actor
    template.modified_by = actor
    template.save()
    return template


def preview_message_template(*, subject, body):
    values = sample_values()
    return {
        "subject": render_template_text(subject, values),
        "body": render_template_text(body, values),
    }


def _safe_excel_text(value):
    text = "" if value is None else str(value)
    return ILLEGAL_CHARACTERS_RE.sub("", text)


def _person_values(person):
    assignments = getattr(person, "current_export_assignments", ())
    assignment = assignments[0] if assignments else None
    if assignment:
        team = assignment.team
        section = team.section
        location = section.location
        company = location.company
    else:
        team = section = location = company = None
    return {
        "HYDRA_ID": person.hydra_id,
        "PASSPORT_NAME": person.passport_name,
        "FIRST_NAME": person.first_name,
        "LAST_NAME": person.last_name,
        "DATE_OF_BIRTH": person.date_of_birth.isoformat(),
        "CITIZENSHIP": person.citizenship,
        "PREFERRED_LANGUAGE": person.preferred_language,
        "PHONE": person.phone,
        "WHATSAPP_VIBER": person.whatsapp_viber,
        "EMAIL": person.email,
        "LIFECYCLE_STATE": person.lifecycle_state,
        "COMPANY_NAME": company.company if company else "",
        "LOCATION_NAME": location.name if location else "",
        "SECTION_NAME": section.name if section else "",
        "TEAM_NAME": team.name if team else "",
    }


def build_template_data_workbook(*, people):
    workbook = Workbook()
    workbook.properties.creator = "Hydra"
    workbook.properties.title = "Szablonizator data export"
    data_sheet = workbook.active
    data_sheet.title = "Dane"
    data_sheet.sheet_view.showGridLines = False
    headers = [placeholder.name for placeholder in PLACEHOLDERS]
    data_sheet.append(headers)
    for person in people:
        values = _person_values(person)
        data_sheet.append([_safe_excel_text(values[header]) for header in headers])

    navy = "17324D"
    pale_blue = "EAF2F8"
    white = "FFFFFF"
    border = Border(bottom=Side(style="thin", color="B8C6D1"))
    for cell in data_sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    data_sheet.row_dimensions[1].height = 28
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions
    for column_index, placeholder in enumerate(PLACEHOLDERS, start=1):
        column = get_column_letter(column_index)
        max_length = max(
            len(placeholder.name),
            *(
                len(str(data_sheet.cell(row=row, column=column_index).value or ""))
                for row in range(2, data_sheet.max_row + 1)
            ),
        )
        data_sheet.column_dimensions[column].width = min(max(max_length + 2, 12), 36)
        for cell in data_sheet[column][1:]:
            cell.data_type = "s"
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="top")

    instruction_sheet = workbook.create_sheet("Instrukcja")
    instruction_sheet.sheet_view.showGridLines = False
    instruction_sheet["A1"] = "Hydra → Szablonizator"
    instruction_sheet["A1"].font = Font(size=16, bold=True, color=navy)
    instruction_sheet["A3"] = "Contract"
    instruction_sheet["A3"].font = Font(bold=True, color=white)
    instruction_sheet["A3"].fill = PatternFill("solid", fgColor=navy)
    instructions = (
        "Use the Dane sheet as the data source.",
        "Headers are stable identifiers from the Hydra placeholder registry.",
        "Every cell is a value; formulas and hidden worksheets are intentionally absent.",
        "Dates use the ISO YYYY-MM-DD format.",
        "The workbook is an authorized point-in-time export. Do not upload it to public storage.",
        "Szablonizator remains a separate desktop application; Hydra does not run WPF, .NET or .exe code.",
    )
    for row, instruction in enumerate(instructions, start=4):
        instruction_sheet.cell(row=row, column=1, value=f"{row - 3}.")
        instruction_sheet.cell(row=row, column=2, value=instruction)
    header_row = 12
    instruction_sheet.cell(row=header_row, column=1, value="Placeholder")
    instruction_sheet.cell(row=header_row, column=2, value="Meaning")
    for cell in instruction_sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
    for row, placeholder in enumerate(PLACEHOLDERS, start=header_row + 1):
        instruction_sheet.cell(row=row, column=1, value=placeholder.token)
        instruction_sheet.cell(row=row, column=2, value=placeholder.label)
        if row % 2:
            instruction_sheet.cell(row=row, column=1).fill = PatternFill("solid", fgColor=pale_blue)
            instruction_sheet.cell(row=row, column=2).fill = PatternFill("solid", fgColor=pale_blue)
    instruction_sheet.column_dimensions["A"].width = 26
    instruction_sheet.column_dimensions["B"].width = 88
    instruction_sheet.freeze_panes = "A13"
    instruction_sheet.sheet_state = "visible"
    data_sheet.sheet_state = "visible"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def create_template_data_export(*, actor, company=None):
    if not actor.has_perms(EXPORT_PERMISSIONS):
        raise PermissionDenied
    scoped_company_ids = sorted(company_ids_for_user(user=actor))
    if company is not None and company.pk not in scoped_company_ids:
        raise PermissionDenied
    people = list(export_people_for_user(user=actor, company=company))
    if len(people) > MAX_EXPORT_ROWS:
        raise ValidationError(
            _("Export exceeds the %(limit)s-row safety limit.")
            % {"limit": MAX_EXPORT_ROWS}
        )
    payload = build_template_data_workbook(people=people)
    digest = hashlib.sha256(payload).hexdigest()
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    filename = f"Hydra_Szablonizator_Dane_{timestamp}.xlsx"
    audit = TemplateDataExport.objects.create(
        actor=actor,
        filename=filename,
        row_count=len(people),
        sha256=digest,
        filters={"company_id": company.pk if company else None},
        scope_company_ids=scoped_company_ids,
    )
    return payload, audit
