import hashlib
from io import BytesIO

from django.core.exceptions import ValidationError
from django.urls import reverse
from openpyxl import load_workbook

from base.models import HydraMailTemplate
from hydra_people.tests.test_recruitment import HydraRecruitmentTestCase
from hydra_shell.templatetags.hydra_shell_tags import hydra_nav_is_active
from hydra_templates.models import MessageTemplate, TemplateDataExport
from hydra_templates.placeholders import placeholder_names, render_template_text
from hydra_templates.services import save_message_template


class TemplateModuleTestCase(HydraRecruitmentTestCase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.template_a = save_message_template(
            template=MessageTemplate(
                company=cls.company_a,
                code="WELCOME",
                name="Welcome A",
                language=MessageTemplate.Language.POLISH,
                subject="Welcome {{FIRST_NAME}}",
                body="Hydra ID: {{ HYDRA_ID }}",
            ),
            actor=cls.admin,
        )
        cls.template_b = save_message_template(
            template=MessageTemplate(
                company=cls.company_b,
                code="WELCOME",
                name="Welcome B",
                language=MessageTemplate.Language.POLISH,
                subject="Welcome {{FIRST_NAME}}",
                body="Team: {{TEAM_NAME}}",
            ),
            actor=cls.admin,
        )

    def grant_template_read(self):
        self.grant(("hydra_templates", "view_messagetemplate"))

    def grant_template_write(self):
        self.grant_template_read()
        self.grant(
            ("hydra_templates", "add_messagetemplate"),
            ("hydra_templates", "change_messagetemplate"),
        )

    def grant_export(self):
        self.grant_template_read()
        self.grant(
            ("hydra_templates", "export_template_data"),
            ("hydra_people", "view_person"),
            ("hydra_templates", "view_templatedataexport"),
        )


class PlaceholderContractTests(TemplateModuleTestCase):
    def test_parser_accepts_contract_spacing_and_renders_known_values(self):
        source = "Hello {{ FIRST_NAME }} / {{HYDRA_ID}} / {{FIRST_NAME}}"

        names = placeholder_names(source)
        rendered = render_template_text(
            source,
            {"FIRST_NAME": "Anna", "HYDRA_ID": "HYD-1"},
        )

        self.assertEqual(names, ("FIRST_NAME", "HYDRA_ID"))
        self.assertEqual(rendered, "Hello Anna / HYD-1 / Anna")

    def test_parser_rejects_unknown_and_malformed_placeholders(self):
        with self.assertRaisesMessage(ValidationError, "Unknown placeholders"):
            placeholder_names("{{PERSON.LABEL_FROM_UI}}")
        for malformed in ("{{ FIRST NAME }}", "{{BROKEN}", "{HYDRA_ID}}"):
            with self.subTest(malformed=malformed):
                with self.assertRaisesMessage(ValidationError, "Malformed placeholder"):
                    placeholder_names(malformed)


class TemplateScopePermissionAndPreviewTests(TemplateModuleTestCase):
    def test_list_and_direct_update_intersect_permission_and_company_scope(self):
        self.grant_template_write()
        self.login()

        response = self.client.get(reverse("hydra-template-list"))
        denied = self.client.get(
            reverse("hydra-template-update", args=(self.template_b.uuid,))
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Welcome A")
        self.assertNotContains(response, "Welcome B")
        self.assertContains(response, 'aria-current="page"')
        self.assertEqual(denied.status_code, 404)

    def test_missing_model_permission_returns_403(self):
        self.login()

        response = self.client.get(reverse("hydra-template-list"))

        self.assertEqual(response.status_code, 403)

    def test_create_form_rejects_company_outside_active_scope(self):
        self.grant_template_write()
        self.login()

        response = self.client.post(
            reverse("hydra-template-create"),
            {
                "company": self.company_b.pk,
                "code": "OUTSIDE",
                "name": "Outside",
                "language": "en",
                "subject": "Hello {{FIRST_NAME}}",
                "body": "{{HYDRA_ID}}",
                "is_active": "on",
                "action": "save",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select a valid choice")
        self.assertFalse(MessageTemplate.objects.filter(code="OUTSIDE").exists())

    def test_preview_is_plain_text_escaped_and_does_not_save(self):
        self.grant_template_write()
        self.login()
        count = MessageTemplate.objects.count()

        response = self.client.post(
            reverse("hydra-template-create"),
            {
                "company": self.company_a.pk,
                "code": "PREVIEW",
                "name": "Preview",
                "language": "uk",
                "subject": "Hello {{FIRST_NAME}}",
                "body": "2 < 3 & 4 > 1\n{{ HYDRA_ID }}",
                "is_active": "on",
                "action": "preview",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["form"].errors, response.context["form"].errors)
        self.assertContains(response, "Hello Anna")
        self.assertContains(response, "2 &lt; 3 &amp; 4 &gt; 1")
        self.assertNotContains(response, "2 < 3 & 4 > 1")
        self.assertEqual(MessageTemplate.objects.count(), count)

    def test_unknown_placeholder_is_a_field_error(self):
        self.grant_template_write()
        self.login()

        response = self.client.post(
            reverse("hydra-template-create"),
            {
                "company": self.company_a.pk,
                "code": "INVALID",
                "name": "Invalid",
                "language": "pl",
                "subject": "{{UNKNOWN}}",
                "body": "Body",
                "is_active": "on",
                "action": "save",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Unknown placeholders: UNKNOWN")
        self.assertFalse(MessageTemplate.objects.filter(code="INVALID").exists())


class TemplateDataExportTests(TemplateModuleTestCase):
    def test_missing_export_permission_returns_403(self):
        self.grant_template_read()
        self.grant(("hydra_people", "view_person"))
        self.login()

        response = self.client.post(reverse("hydra-template-data-export"), {})

        self.assertEqual(response.status_code, 403)

    def test_export_is_scoped_values_only_and_records_checksum(self):
        self.grant_export()
        self.login()
        type(self.person_a).objects.filter(pk=self.person_a.pk).update(
            passport_name="=2+2"
        )

        response = self.client.post(
            reverse("hydra-template-data-export"),
            {"company": self.company_a.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(response["Cache-Control"], "no-store, private")
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Dane", "Instrukcja"])
        self.assertTrue(all(sheet.sheet_state == "visible" for sheet in workbook))
        data_sheet = workbook["Dane"]
        headers = tuple(cell.value for cell in data_sheet[1])
        self.assertEqual(
            headers,
            (
                "HYDRA_ID",
                "PASSPORT_NAME",
                "FIRST_NAME",
                "LAST_NAME",
                "DATE_OF_BIRTH",
                "CITIZENSHIP",
                "PREFERRED_LANGUAGE",
                "PHONE",
                "WHATSAPP_VIBER",
                "EMAIL",
                "LIFECYCLE_STATE",
                "COMPANY_NAME",
                "LOCATION_NAME",
                "SECTION_NAME",
                "TEAM_NAME",
            ),
        )
        rows = list(data_sheet.iter_rows(min_row=2, values_only=True))
        exported_ids = {row[0] for row in rows}
        self.assertEqual(exported_ids, {self.person_a.hydra_id, self.person_c.hydra_id})
        self.assertNotIn(self.person_b.hydra_id, exported_ids)
        self.assertIn("=2+2", {row[1] for row in rows})
        self.assertTrue(all(row[4] == "1992-04-05" for row in rows))
        self.assertTrue(
            all(cell.data_type != "f" for sheet in workbook for row in sheet.iter_rows() for cell in row)
        )
        workbook.close()
        audit = TemplateDataExport.objects.get()
        self.assertEqual(audit.actor, self.user)
        self.assertEqual(audit.row_count, 2)
        self.assertEqual(audit.sha256, hashlib.sha256(response.content).hexdigest())
        self.assertEqual(audit.filters, {"company_id": self.company_a.pk})

    def test_out_of_scope_company_is_rejected_without_audit(self):
        self.grant_export()
        self.login()

        response = self.client.post(
            reverse("hydra-template-data-export"),
            {"company": self.company_b.pk},
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(TemplateDataExport.objects.exists())

    def test_export_audit_is_append_only(self):
        audit = TemplateDataExport.objects.create(
            actor=self.admin,
            filename="audit.xlsx",
            row_count=0,
            sha256="0" * 64,
            filters={},
            scope_company_ids=[],
        )

        audit.row_count = 2
        with self.assertRaisesMessage(TypeError, "append-only"):
            audit.save()
        with self.assertRaisesMessage(TypeError, "append-only"):
            audit.delete()
        with self.assertRaisesMessage(TypeError, "append-only"):
            TemplateDataExport.objects.update(row_count=2)
        with self.assertRaisesMessage(TypeError, "append-only"):
            TemplateDataExport.objects.all().delete()


class TemplateCompatibilityTests(TemplateModuleTestCase):
    def test_shell_state_and_legacy_templates_remain_operational(self):
        self.grant_template_read()
        self.login()
        legacy = HydraMailTemplate.objects.create(
            title="Legacy template",
            body="Legacy body",
            company_id=self.company_a,
        )

        response = self.client.get(reverse("hydra-template-list"))
        self.client.force_login(self.admin)
        legacy_response = self.client.get(reverse("view-mail-templates"))

        self.assertTrue(hydra_nav_is_active(response.context, "templates"))
        self.assertEqual(legacy_response.status_code, 200)
        self.assertContains(legacy_response, legacy.title)
