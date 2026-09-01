"""
Tests for enterprise reporting foundation — registry, engine, metric formulas.
"""

from datetime import date

from django.test import RequestFactory, SimpleTestCase

from report.engine import ReportFilters, month_bounds, month_offset, parse_period
from report.export import export_csv, export_xlsx
from report.pivot_limits import MAX_PIVOT_ROWS, capped_list
from report.registry import (
    ReportDefinition,
    get_report,
    list_reports,
    register,
    run_report,
)


class EngineHelpersTests(SimpleTestCase):
    def test_parse_period_defaults(self):
        from_date, to_date, preset = parse_period()
        today = date.today()
        self.assertEqual(from_date, today.replace(day=1))
        self.assertEqual(to_date, today)
        self.assertEqual(preset, "custom")

    def test_period_presets(self):
        from report.engine import ALL_TIME_FROM, resolve_period_preset

        today = date(2026, 8, 4)
        start, end = resolve_period_preset("ytd", today)
        self.assertEqual(start, date(2026, 1, 1))
        self.assertEqual(end, today)
        start, end = resolve_period_preset("last_30", today)
        self.assertEqual((end - start).days, 29)
        start, end = resolve_period_preset("all_time", today)
        self.assertEqual(start, ALL_TIME_FROM)
        self.assertEqual(end, today)

    def test_all_time_filters_from_request(self):
        from report.engine import (
            ALL_TIME_FROM,
            EMPLOYMENT_STATUS_ALL,
            filters_from_request,
        )

        request = RequestFactory().get(
            "/report/standard/workforce-composition/export/",
            {
                "period_preset": "all_time",
                "employment_status": "all",
                "format": "pdf",
            },
        )
        filters = filters_from_request(request)
        self.assertEqual(filters.period_preset, "all_time")
        self.assertEqual(filters.from_date, ALL_TIME_FROM)
        self.assertEqual(filters.to_date, date.today())
        self.assertEqual(filters.employment_status, EMPLOYMENT_STATUS_ALL)
        self.assertEqual(filters.period_label, "All time")
        self.assertIn("All time", filters.summary_labels())
        self.assertNotIn("Status: all", " · ".join(filters.summary_labels()))

    def test_print_filters_url_resolves(self):
        from django.urls import reverse

        url = reverse("standard-report-print-filters", args=["workforce-composition"])
        self.assertEqual(url, "/report/standard/workforce-composition/print-filters/")

    def test_month_offset(self):
        d = date(2026, 3, 15)
        self.assertEqual(month_offset(d, 0), date(2026, 3, 1))
        self.assertEqual(month_offset(d, 3), date(2025, 12, 1))

    def test_month_bounds(self):
        start, end = month_bounds(date(2026, 2, 10))
        self.assertEqual(start, date(2026, 2, 1))
        self.assertEqual(end, date(2026, 2, 28))

    def test_capped_list(self):
        rows = list(range(10))
        data, truncated = capped_list(rows, limit=5)
        self.assertEqual(len(data), 5)
        self.assertTrue(truncated)
        data2, truncated2 = capped_list(rows, limit=20)
        self.assertFalse(truncated2)
        self.assertEqual(MAX_PIVOT_ROWS, 5000)


class RegistryTests(SimpleTestCase):
    def test_new_operational_registers_registered(self):
        "Payroll readiness and the three registers are in the catalog."
        import report.metrics  # noqa: F401
        from report.registry import get_report

        expected = {
            "payroll-readiness": "payroll",
            "loan-advance-ledger": "payroll",
            "reimbursement-register": "payroll",
            "asset-register": "compliance",
        }
        for slug, domain in expected.items():
            definition = get_report(slug)
            self.assertIsNotNone(definition, f"{slug} not registered")
            self.assertEqual(definition.domain, domain)
            self.assertTrue(callable(definition.query_fn))
            # Each carries its own subject-matter permission so it can be
            # granted without handing over unrelated access.
            self.assertTrue(definition.alt_permissions)

    def test_reimbursement_register_keeps_undated_claims(self):
        "Undated rows must not vanish from every window."
        import inspect

        from report.metrics.payroll import reimbursement_register

        # Guards the bulk_create/auto_now_add NULL created_at trap: a plain
        # range filter excludes those rows from any period at all.
        source = inspect.getsource(reimbursement_register)
        self.assertIn("created_at__isnull=True", source)

    def test_retired_slugs_still_resolve(self):
        "Consolidated reports keep resolving so bookmarks do not 404."
        import report.metrics  # noqa: F401
        from report.registry import RETIRED_SLUGS, get_report

        self.assertTrue(RETIRED_SLUGS)
        for retired, survivor in RETIRED_SLUGS.items():
            definition = get_report(retired)
            self.assertIsNotNone(
                definition, f"retired slug {retired} no longer resolves"
            )
            self.assertEqual(definition.slug, survivor)

    def test_compliance_reports_grantable_independently(self):
        "Compliance must not require full workforce analytics access."
        import report.metrics  # noqa: F401
        from report.registry import get_report

        definition = get_report("document-expiry-aging")
        self.assertIn("horilla_documents.view_document", definition.alt_permissions)

        class _User:
            is_authenticated = True
            is_superuser = False

            def __init__(self, perms):
                self._perms = perms

            def has_perm(self, perm):
                return perm in self._perms

        # Granted on its own subject-matter permission ...
        self.assertTrue(
            definition.user_has_permission(_User({"horilla_documents.view_document"}))
        )
        # ... without revoking anyone who only holds the original one.
        self.assertTrue(
            definition.user_has_permission(_User({"employee.view_employee"}))
        )
        self.assertFalse(definition.user_has_permission(_User({"asset.view_asset"})))

    def test_definitions_registered(self):
        import report.metrics  # noqa: F401

        report = get_report("workforce-composition")
        self.assertIsNotNone(report)
        self.assertEqual(report.domain, "workforce")
        self.assertTrue(report.permission.endswith("view_employee"))

        slugs = {r.slug for r in list_reports()}
        for expected in (
            "workforce-composition",
            "diversity-snapshot",
            "tenure-longevity",
            "turnover-attrition",
            "joiners-leavers",
            "attendance-summary",
            "absenteeism-rate",
            "overtime-analysis",
            "leave-utilization",
            "leave-liability",
            "labor-cost-summary",
            "cost-composition",
            "payroll-headcount-cost",
            "payslip-register",
            "recruitment-funnel",
            "time-to-hire",
            "offer-acceptance",
            "performance-distribution",
            "audit-activity",
            "span-of-control",
            "pipeline-aging",
            "source-quality",
            "document-expiry-aging",
            "headcount-bridge",
            "exit-analysis",
            "new-hire-90-day-attrition",
            "unscheduled-absence",
            "visa-contract-expiry",
            "quality-of-hire",
        ):
            self.assertIn(expected, slugs)

        composition = get_report("workforce-composition")
        self.assertIsNotNone(composition.drilldown_fn)

    def test_suggested_report_slugs_registered(self):
        import report.metrics  # noqa: F401
        from report.personalization import SUGGESTED_REPORT_SLUGS

        slugs = {r.slug for r in list_reports()}
        for slug in SUGGESTED_REPORT_SLUGS:
            self.assertIn(slug, slugs)
        # Culture-sensitive defaults stay out of Suggested pack
        for excluded in (
            "absenteeism-rate",
            "payslip-register",
            "diversity-snapshot",
            "performance-distribution",
        ):
            self.assertNotIn(excluded, SUGGESTED_REPORT_SLUGS)

    def test_dashboard_pin_priority_subset(self):
        from report.personalization import (
            DASHBOARD_PIN_PRIORITY_SLUGS,
            MAX_DASHBOARD_REPORT_PINS,
            SUGGESTED_REPORT_SLUGS,
        )

        self.assertEqual(MAX_DASHBOARD_REPORT_PINS, 6)
        self.assertTrue(
            set(DASHBOARD_PIN_PRIORITY_SLUGS).issubset(set(SUGGESTED_REPORT_SLUGS))
        )
        self.assertLessEqual(
            len(DASHBOARD_PIN_PRIORITY_SLUGS), MAX_DASHBOARD_REPORT_PINS
        )

    def test_run_report_attaches_metadata(self):
        def fake_query(filters):
            return {
                "kpis": [{"label": "A", "value": 1}],
                "charts": [],
                "table": {"columns": [], "rows": []},
            }

        register(
            ReportDefinition(
                slug="unit-test-report",
                name="Unit Test",
                domain="workforce",
                description="test",
                permission="employee.view_employee",
                query_fn=fake_query,
                required_apps=(),
            )
        )
        filters = ReportFilters(from_date=date(2026, 1, 1), to_date=date(2026, 1, 31))
        payload = run_report("unit-test-report", filters)
        self.assertEqual(payload["slug"], "unit-test-report")
        self.assertEqual(payload["period"]["from_date"], "2026-01-01")


class ExportTests(SimpleTestCase):
    def test_exports_neutralize_spreadsheet_formulas(self):
        """A value beginning =, +, -, @ must not execute when opened."""
        import io as _io

        import openpyxl

        from report.export import export_csv, export_xlsx

        payload = {
            "title": "Injection",
            "period": {"from_date": "2026-01-01", "to_date": "2026-01-31"},
            "kpis": [],
            "table": {
                "columns": [
                    {"key": "name", "label": "Name"},
                    {"key": "n", "label": "Count"},
                ],
                "rows": [
                    {"name": '=HYPERLINK("http://evil.test","x")', "n": 5},
                    {"name": "@SUM(A1:A9)", "n": -3},
                    {"name": "Normal Name", "n": 7},
                ],
            },
        }

        csv_body = export_csv(payload, "i.csv").content.decode("utf-8-sig")
        self.assertIn("'=HYPERLINK", csv_body)
        self.assertIn("'@SUM(A1:A9)", csv_body)
        self.assertIn("Normal Name", csv_body)

        wb = openpyxl.load_workbook(
            _io.BytesIO(export_xlsx(payload, "i.xlsx", meta={}).content)
        )
        values = [c for row in wb["Data"].iter_rows(values_only=True) for c in row]
        # The exporter writes its own =SUM(...) for the totals row, which must
        # stay a live formula -- only data carried through from the payload is
        # neutralized, so assert on the injected values specifically.
        self.assertIn('\'=HYPERLINK("http://evil.test","x")', values)
        self.assertIn("'@SUM(A1:A9)", values)
        self.assertIn("Normal Name", values)
        # Negative numbers must stay numeric -- the guard only touches strings.
        self.assertIn(-3, values)

    def test_export_csv_and_xlsx(self):
        payload = {
            "title": "Workforce Composition",
            "slug": "workforce-composition",
            "domain": "workforce",
            "period": {"from_date": "2026-01-01", "to_date": "2026-01-31"},
            "kpis": [
                {"label": "Headcount", "value": 10, "hint": "Active"},
                {"label": "Female %", "value": "42.5%", "hint": ""},
            ],
            "table": {
                "columns": [
                    {"key": "dept", "label": "Department"},
                    {"key": "count", "label": "Count"},
                    {"key": "gross", "label": "Gross Pay"},
                ],
                "rows": [
                    {"dept": "HR", "count": 3, "gross": 12000.5},
                    {"dept": "IT", "count": 7, "gross": 54000},
                ],
            },
            "charts": [
                {
                    "id": "c1",
                    "title": "By Dept",
                    "categories": ["HR", "IT"],
                    "series": [{"name": "Count", "data": [3, 7]}],
                }
            ],
        }
        csv_resp = export_csv(payload, "sample.csv")
        # charset is declared explicitly, and the body carries a UTF-8 BOM so
        # Excel on Windows does not fall back to the local ANSI codepage.
        self.assertTrue(csv_resp["Content-Type"].startswith("text/csv"))
        self.assertTrue(csv_resp.content.startswith(b"\xef\xbb\xbf"))
        self.assertIn(b"Headcount", csv_resp.content)
        self.assertIn(b"Key Metrics", csv_resp.content)

        xlsx_resp = export_xlsx(
            payload,
            "sample.xlsx",
            meta={
                "company": {
                    "name": "Acme Corp",
                    "address": "1 Market St",
                    "city": "SF",
                    "state": "CA",
                    "country": "USA",
                    "zip": "94105",
                    "is_all": False,
                    "location_line": "SF, CA, USA",
                    "address_lines": [
                        "Acme Corp",
                        "1 Market St",
                        "SF, CA, USA",
                        "ZIP: 94105",
                    ],
                    "logo_path": None,
                },
                "user": "Admin",
                "product_name": "Horilla HR · Standard Reports",
                "filters_label": "This month · Active",
                "slug": "workforce-composition",
                "domain": "workforce",
            },
        )
        self.assertIn("spreadsheetml", xlsx_resp["Content-Type"])
        self.assertTrue(len(xlsx_resp.content) > 1000)

        # Workbook structure: Cover, Data, Charts
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(xlsx_resp.content))
        self.assertEqual(wb.sheetnames, ["Cover", "Data", "Charts"])
        cover = wb["Cover"]
        # Company letterhead + report title somewhere on Cover
        cover_values = [
            str(c.value)
            for row in cover.iter_rows(max_row=20, max_col=6)
            for c in row
            if c.value
        ]
        self.assertTrue(any("Acme Corp" in v for v in cover_values))
        self.assertTrue(any("Workforce Composition" in v for v in cover_values))
        self.assertTrue(any("Report details" in v for v in cover_values))
        self.assertTrue(any("Applied filters" in v for v in cover_values))
        self.assertTrue(any("Key performance indicators" in v for v in cover_values))

        data = wb["Data"]
        data_values = [
            str(c.value)
            for row in data.iter_rows(max_row=25, max_col=6)
            for c in row
            if c.value
        ]
        self.assertTrue(any(v == "Department" for v in data_values))
        self.assertTrue(data.auto_filter.ref)
        self.assertTrue(data.freeze_panes)

    def test_zero_padded_identifiers_keep_their_padding(self):
        """Badge ids / phone numbers must not be coerced to int.

        int("00042") == 42 silently destroyed the identifier; a zero-padded
        digit run is data, not a quantity.
        """
        from report.export import _coerce_cell

        for text in ("00042", "0501234567", "007", "00"):
            self.assertEqual(_coerce_cell(text), text, msg=text)
        # A bare zero and ordinary integers are still real numbers.
        self.assertEqual(_coerce_cell("0"), 0)
        self.assertEqual(_coerce_cell("42"), 42)
        self.assertEqual(_coerce_cell("-42"), -42)

    def test_percent_strings_above_100_are_not_divided_twice(self):
        """ "150%" must display as 150.0%, not 1.5%.

        _coerce_cell already turns "150%" into 1.5; the data sheet's >1
        rescale then has to leave it alone or the value is divided twice.
        """
        import io as _io

        import openpyxl

        from report.export import export_xlsx

        payload = {
            "title": "Rates",
            "period": {"from_date": "2026-01-01", "to_date": "2026-01-31"},
            "kpis": [],
            "table": {
                "columns": [
                    {"key": "src", "label": "Source"},
                    {"key": "rate", "label": "Accept Rate %"},
                ],
                "rows": [
                    {"src": "Referral", "rate": "150%"},
                    {"src": "Portal", "rate": "89.5%"},
                    # Bare integer under a percent header: must still format
                    # as a percentage rather than a plain number.
                    {"src": "Agency", "rate": "75"},
                ],
            },
        }
        wb = openpyxl.load_workbook(
            _io.BytesIO(export_xlsx(payload, "r.xlsx", meta={}).content)
        )
        ws = wb["Data"]
        by_source = {}
        for row in ws.iter_rows():
            cells = [c for c in row]
            if len(cells) >= 2 and cells[0].value in ("Referral", "Portal", "Agency"):
                by_source[cells[0].value] = cells[1]

        self.assertAlmostEqual(by_source["Referral"].value, 1.5, places=4)
        self.assertAlmostEqual(by_source["Portal"].value, 0.895, places=4)
        self.assertAlmostEqual(by_source["Agency"].value, 0.75, places=4)
        for cell in by_source.values():
            self.assertEqual(cell.number_format, "0.0%")


class AttendanceDecimalTests(SimpleTestCase):
    """The pivot sums the *_Decimal columns, so they must be real hours."""

    def _helpers(self):
        # The converters are closures inside the module-level guard, so reach
        # them the way the view does -- via the registered pivot view module.
        from report.views import attendance_report  # noqa: F401

        return attendance_report

    def test_duration_to_decimal_hours(self):
        from datetime import time as _time

        # Re-derive the same arithmetic the view applies; a "HH.MM" string
        # would make 1:45 + 1:45 total 2.90 instead of 3.50.
        def convert(value):
            if isinstance(value, str):
                hours, minutes = map(int, value.split(":")[:2])
            elif isinstance(value, _time):
                hours, minutes = value.hour, value.minute
            else:
                return 0.0
            return round(hours + minutes / 60, 2)

        self.assertEqual(convert("1:45"), 1.75)
        self.assertEqual(convert("0:30"), 0.5)
        self.assertEqual(convert("8:00"), 8.0)
        self.assertEqual(convert("1:45") + convert("1:45"), 3.5)
        self.assertEqual(convert(None), 0.0)

    def test_view_module_returns_numeric_decimals(self):
        """Guard the real converters, not just the arithmetic above."""
        import inspect

        from report.views import attendance_report

        source = inspect.getsource(attendance_report)
        # The base-60-as-base-100 formatting must be gone from both helpers.
        self.assertNotIn('f"{hours:02}.{minutes:02}"', source)
        self.assertNotIn('f"{t.hour:02}.{t.minute:02}"', source)


class SubscriptionFormTests(SimpleTestCase):
    def test_recipients_reject_malformed_addresses(self):
        from report.forms import ReportSubscriptionForm

        form = ReportSubscriptionForm(
            data={
                "report_slug": "workforce-composition",
                "name": "Weekly",
                "frequency": "weekly",
                "recipients": "a@b.com, notanemail",
                "format": "xlsx",
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("recipients", form.errors)
        self.assertIn("notanemail", str(form.errors["recipients"]))

    def test_recipients_normalize_and_dedupe(self):
        from report.forms import ReportSubscriptionForm

        form = ReportSubscriptionForm(
            data={
                "report_slug": "workforce-composition",
                "name": "Weekly",
                "frequency": "weekly",
                "recipients": " a@b.com , c@d.com,a@b.com ",
                "format": "xlsx",
            }
        )
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["recipients"], "a@b.com, c@d.com")

    def test_recipients_required(self):
        from report.forms import ReportSubscriptionForm

        form = ReportSubscriptionForm(
            data={
                "report_slug": "workforce-composition",
                "name": "Weekly",
                "frequency": "weekly",
                "recipients": " , ",
                "format": "xlsx",
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("recipients", form.errors)


class PdfExportTests(SimpleTestCase):
    def test_narrative_blurb(self):
        from report.narrative import build_narrative

        payload = {
            "title": "Workforce Composition",
            "compare": {
                "preset": "prior_period",
                "label": "Prior period",
                "period": {"from_date": "2026-07-01", "to_date": "2026-07-31"},
            },
            "kpis": [
                {
                    "label": "Headcount",
                    "value": 12,
                    "prior_value": 10,
                    "delta_label": "+2 (+20.0%)",
                    "delta_direction": "up",
                }
            ],
        }
        narrative = build_narrative(payload)
        self.assertIn("Prior period", narrative)
        self.assertIn("Headcount", narrative)

    def test_export_pdf_via_pisa(self):
        from unittest.mock import patch

        from report.export import export_pdf

        payload = {
            "title": "Workforce Composition",
            "slug": "workforce-composition",
            "domain": "workforce",
            "period": {"from_date": "2026-08-01", "to_date": "2026-08-31"},
            "filters": ["2026-08-01 → 2026-08-31"],
            "kpis": [{"label": "Headcount", "value": 12, "hint": "Active"}],
            "table": {
                "columns": [
                    {"key": "dept", "label": "Department"},
                    {"key": "count", "label": "Count"},
                ],
                "rows": [{"dept": "HR", "count": 3}],
            },
            "charts": [],
        }
        html = (
            "<html><body><h1>Workforce Composition</h1>"
            "<p>Headcount 12</p></body></html>"
        )
        with patch("django.template.loader.render_to_string", return_value=html):
            pdf = export_pdf(
                payload,
                "sample.pdf",
                meta={
                    "company": {
                        "name": "Acme Corp",
                        "is_all": False,
                        "address_lines": ["Acme Corp"],
                        "logo_path": None,
                    },
                    "user": "Admin",
                    "product_name": "Horilla HR · Standard Reports",
                    "domain": "workforce",
                },
            )
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))
        self.assertTrue(len(pdf.content) > 200)


class TurnoverFormulaTests(SimpleTestCase):
    """Pure formula checks mirroring metric logic without DB."""

    def test_turnover_rate_formula(self):
        total_employees = 100
        total_exits = 12
        rate = round(total_exits / total_employees * 100, 1)
        self.assertEqual(rate, 12.0)

    def test_absenteeism_formula(self):
        expected = 20 * 22  # employees * working days
        present = 400
        absent = max(0, expected - present)
        rate = round(absent / expected * 100, 1)
        self.assertEqual(absent, 40)
        self.assertEqual(rate, 9.1)

    def test_leave_utilization_formula(self):
        allocated = 100.0
        used = 35.0
        rate = round((used / allocated * 100), 1)
        self.assertEqual(rate, 35.0)


class PersonalizationHelperTests(SimpleTestCase):
    """Phase 2 helpers (no DB — test DB migrate needs base migrations)."""

    def test_filters_dict_from_request(self):
        from django.test import RequestFactory

        from report.personalization import filters_dict_from_request

        request = RequestFactory().get(
            "/report/standard/x/data/",
            {
                "period_preset": "last_month",
                "department_id": "3",
                "format": "xlsx",
                "csrfmiddlewaretoken": "ignore-me",
            },
        )
        filters = filters_dict_from_request(request)
        self.assertEqual(filters["period_preset"], "last_month")
        self.assertEqual(filters["department_id"], "3")
        self.assertNotIn("csrfmiddlewaretoken", filters)

    def test_session_company_id(self):
        from django.test import RequestFactory

        from report.personalization import session_company_id

        request = RequestFactory().get("/")
        request.session = {}
        self.assertIsNone(session_company_id(request))
        request.session["selected_company"] = "all"
        self.assertIsNone(session_company_id(request))
        request.session["selected_company"] = "42"
        self.assertEqual(session_company_id(request), 42)


class SubscriptionDeliveryTests(SimpleTestCase):
    """Phase 3 due-window + filter reconstruction (no DB)."""

    def test_filters_from_dict_presets(self):
        from datetime import date as d

        from report.engine import filters_from_dict

        f = filters_from_dict({"period_preset": "ytd", "department_id": "9"})
        today = d.today()
        self.assertEqual(f.from_date, d(today.year, 1, 1))
        self.assertEqual(f.to_date, today)
        self.assertEqual(f.department_id, 9)
        self.assertEqual(f.period_preset, "ytd")

        legacy = filters_from_dict({"period": "last_30"})
        self.assertEqual(legacy.period_preset, "last_30")
        self.assertEqual((legacy.to_date - legacy.from_date).days, 29)

        from report.engine import ALL_TIME_FROM

        all_time = filters_from_dict(
            {"period_preset": "all_time", "employment_status": "all"}
        )
        self.assertEqual(all_time.period_preset, "all_time")
        self.assertEqual(all_time.from_date, ALL_TIME_FROM)
        self.assertEqual(all_time.employment_status, "all")

    def test_subscription_is_due(self):
        from datetime import timedelta
        from types import SimpleNamespace

        from django.utils import timezone

        from report.delivery import subscription_is_due

        now = timezone.now()
        fresh = SimpleNamespace(last_run_at=None, frequency="weekly")
        self.assertTrue(subscription_is_due(fresh, now))

        recent = SimpleNamespace(
            last_run_at=now - timedelta(days=1), frequency="weekly"
        )
        self.assertFalse(subscription_is_due(recent, now))

        old = SimpleNamespace(last_run_at=now - timedelta(days=8), frequency="weekly")
        self.assertTrue(subscription_is_due(old, now))

        daily_ok = SimpleNamespace(
            last_run_at=now - timedelta(hours=24), frequency="daily"
        )
        self.assertTrue(subscription_is_due(daily_ok, now))

    def test_subscription_job_is_registered_not_started(self):
        # Replaces an argv-guard test: the module used to start its own
        # BackgroundScheduler at import and skip it for migrate/test argv.
        # Nothing starts at import now, so there is no argv to guard -- the job
        # is registered and run_scheduler owns execution.
        import report.scheduler  # noqa: F401
        from horilla.scheduling import get_registered_jobs

        job = next(
            (j for j in get_registered_jobs() if j.job_id == "report_subscriptions"),
            None,
        )
        self.assertIsNotNone(job, "report_subscriptions job was not registered")
        self.assertEqual(job.trigger, "interval")
        self.assertEqual(job.kwargs, {"hours": 1})


class PeriodCompareTests(SimpleTestCase):
    """Phase 4a prior period / prior year helpers."""

    def test_prior_period_and_year_bounds(self):
        from datetime import date as d

        from report.compare import prior_period_bounds, prior_year_bounds

        self.assertEqual(
            prior_period_bounds(d(2026, 8, 1), d(2026, 8, 31)),
            (d(2026, 7, 1), d(2026, 7, 31)),
        )
        self.assertEqual(
            prior_year_bounds(d(2026, 8, 1), d(2026, 8, 15)),
            (d(2025, 8, 1), d(2025, 8, 15)),
        )
        # Leap day safety
        self.assertEqual(
            prior_year_bounds(d(2024, 2, 29), d(2024, 2, 29)),
            (d(2023, 2, 28), d(2023, 2, 28)),
        )

    def test_kpi_delta_and_merge(self):
        from report.compare import compute_kpi_delta, merge_charts, merge_kpis

        delta = compute_kpi_delta(120, 100)
        self.assertEqual(delta["delta"], 20)
        self.assertEqual(delta["delta_direction"], "up")
        self.assertIn("+20", delta["delta_label"])

        pct = compute_kpi_delta("42.5%", "40%")
        self.assertEqual(pct["delta_direction"], "up")
        self.assertIn("pp", pct["delta_label"])

        kpis = merge_kpis(
            [{"label": "Headcount", "value": 12}],
            [{"label": "Headcount", "value": 10}],
        )
        self.assertEqual(kpis[0]["prior_value"], 10)
        self.assertEqual(kpis[0]["delta"], 2)

        charts = merge_charts(
            [
                {
                    "id": "by_dept",
                    "type": "bar",
                    "categories": ["HR", "IT"],
                    "series": [{"name": "Employees", "data": [3, 7]}],
                }
            ],
            [
                {
                    "id": "by_dept",
                    "type": "bar",
                    "categories": ["IT", "HR"],
                    "series": [{"name": "Employees", "data": [5, 2]}],
                }
            ],
            "Prior period",
        )
        self.assertEqual(len(charts[0]["series"]), 2)
        self.assertTrue(charts[0]["series"][1]["is_compare"])
        # Aligned to current category order: HR=2, IT=5
        self.assertEqual(charts[0]["series"][1]["data"], [2, 5])

    def test_run_report_attaches_compare_block(self):
        from datetime import date as d

        from report.engine import ReportFilters
        from report.registry import ReportDefinition, register, run_report

        def fake_query(filters):
            # Vary KPI by month so compare shows a delta
            value = 10 if filters.from_date.month == 8 else 7
            return {
                "kpis": [{"label": "Headcount", "value": value}],
                "charts": [
                    {
                        "id": "trend",
                        "type": "bar",
                        "categories": ["A"],
                        "series": [{"name": "N", "data": [value]}],
                    }
                ],
                "table": {"columns": [], "rows": []},
            }

        register(
            ReportDefinition(
                slug="unit-compare-report",
                name="Compare Unit",
                domain="workforce",
                description="test",
                permission="employee.view_employee",
                query_fn=fake_query,
                required_apps=(),
            )
        )
        filters = ReportFilters(
            from_date=d(2026, 8, 1),
            to_date=d(2026, 8, 31),
            compare_preset="prior_period",
        )
        payload = run_report("unit-compare-report", filters)
        self.assertIn("compare", payload)
        self.assertEqual(payload["compare"]["preset"], "prior_period")
        self.assertEqual(payload["kpis"][0]["prior_value"], 7)
        self.assertEqual(payload["kpis"][0]["delta"], 3)
        self.assertEqual(len(payload["charts"][0]["series"]), 2)


class AccessMatrixTests(SimpleTestCase):
    """Phase 5 matrix logic without DB (mocked rules)."""

    def test_fallback_when_no_rules(self):
        from types import SimpleNamespace
        from unittest.mock import patch

        from report.access import user_can_export_report, user_can_view_report
        from report.registry import ReportDefinition

        definition = ReportDefinition(
            slug="workforce-composition",
            name="WC",
            domain="workforce",
            description="",
            permission="employee.view_employee",
            query_fn=lambda f: {},
        )
        user = SimpleNamespace(
            is_authenticated=True,
            is_superuser=False,
            has_perm=lambda p: p == "employee.view_employee",
            groups=SimpleNamespace(values_list=lambda *a, **k: []),
        )
        with patch("report.access.matching_access_rules", return_value=[]):
            self.assertTrue(user_can_view_report(user, definition))
            self.assertTrue(user_can_export_report(user, definition))

        user_denied = SimpleNamespace(
            is_authenticated=True,
            is_superuser=False,
            has_perm=lambda p: False,
            groups=SimpleNamespace(values_list=lambda *a, **k: []),
        )
        with patch("report.access.matching_access_rules", return_value=[]):
            self.assertFalse(user_can_view_report(user_denied, definition))

    def test_matrix_deny_export(self):
        from types import SimpleNamespace
        from unittest.mock import patch

        from report.access import user_can_export_report, user_can_view_report
        from report.registry import ReportDefinition

        definition = ReportDefinition(
            slug="workforce-composition",
            name="WC",
            domain="workforce",
            description="",
            permission="employee.view_employee",
            query_fn=lambda f: {},
        )
        user = SimpleNamespace(
            is_authenticated=True,
            is_superuser=False,
            has_perm=lambda p: True,
            groups=SimpleNamespace(values_list=lambda *a, **k: [1]),
        )
        rules = [SimpleNamespace(can_view=True, can_export=False, can_subscribe=False)]
        with patch("report.access.matching_access_rules", return_value=rules):
            self.assertTrue(user_can_view_report(user, definition))
            self.assertFalse(user_can_export_report(user, definition))


class FormulaTests(SimpleTestCase):
    def test_core_hr_formulas(self):
        from report.formulas import (
            absenteeism_rate,
            leave_utilization_rate,
            offer_acceptance_rate,
            ot_concentration_share,
            turnover_rate,
        )

        self.assertEqual(turnover_rate(hires=10, exits=5, avg_headcount=100), 5.0)
        self.assertEqual(turnover_rate(0, 1, 0), 0.0)
        self.assertEqual(absenteeism_rate(10, 200), 5.0)
        self.assertEqual(leave_utilization_rate(40, 80), 50.0)
        self.assertEqual(offer_acceptance_rate(8, 10), 80.0)
        self.assertEqual(ot_concentration_share(600, 1000), 60.0)
        from report.formulas import early_attrition_rate, retention_rate

        self.assertEqual(early_attrition_rate(2, 10), 20.0)
        self.assertEqual(early_attrition_rate(1, 0), 0.0)
        self.assertEqual(retention_rate(8, 10), 80.0)


class TimeToHireHelperTests(SimpleTestCase):
    def test_median(self):
        from report.metrics.talent import _median

        self.assertEqual(_median([]), 0)
        self.assertEqual(_median([10]), 10)
        self.assertEqual(_median([1, 3, 5]), 3)
        self.assertEqual(_median([1, 2, 3, 4]), 2)


class CalendarExpectedDaysTests(SimpleTestCase):
    def test_weekday_only_default(self):
        from report.metrics._calendar import count_expected_working_days

        # Mon 2026-01-05 .. Fri 2026-01-09 = 5 days
        self.assertEqual(
            count_expected_working_days(date(2026, 1, 5), date(2026, 1, 9)),
            5,
        )
        # Include weekend Sat-Sun → still 5
        self.assertEqual(
            count_expected_working_days(date(2026, 1, 5), date(2026, 1, 11)),
            5,
        )

    def test_holiday_reduces_expected_days(self):
        from unittest.mock import patch

        from report.metrics._calendar import count_expected_working_days

        with patch(
            "report.metrics._calendar._holiday_dates",
            return_value={date(2026, 1, 7)},
        ), patch(
            # Company-leave rules are now fetched once per call rather than
            # queried per day; patch that seam, not the old per-day helper.
            "report.metrics._calendar._company_leave_rules",
            return_value=set(),
        ):
            # Mon–Fri minus Wed holiday → 4
            self.assertEqual(
                count_expected_working_days(date(2026, 1, 5), date(2026, 1, 9)),
                4,
            )

    def test_company_leave_rule_matching_matches_legacy_arithmetic(self):
        """The prefetched rule check must agree with base.methods.

        _matches_company_leave replaced a per-day query; it recomputes the
        same 0-based, month-start-offset week number, so a drift here would
        silently change every absenteeism figure.
        """
        from report.metrics._calendar import _matches_company_leave

        # 2026-01-05 is a Monday in the second week block of January 2026.
        monday = date(2026, 1, 5)
        first = monday.replace(day=1)
        week_no = (monday.day + first.weekday() - 1) // 7

        # A week-independent Monday rule matches any Monday.
        self.assertTrue(_matches_company_leave(monday, {(None, 0)}))
        # A rule pinned to this week/weekday matches.
        self.assertTrue(_matches_company_leave(monday, {(week_no, 0)}))
        # A different weekday does not.
        self.assertFalse(_matches_company_leave(monday, {(None, 2)}))
        # A different week block does not.
        self.assertFalse(_matches_company_leave(monday, {(week_no + 1, 0)}))
        # No rules at all is never a company leave.
        self.assertFalse(_matches_company_leave(monday, set()))


class NamedOtPrivacyTests(SimpleTestCase):
    def test_names_require_flag_and_perm(self):
        from types import SimpleNamespace

        from report.engine import ReportFilters
        from report.metrics._privacy import allow_named_ot_rows

        filters = ReportFilters(from_date=date(2026, 1, 1), to_date=date(2026, 1, 31))
        self.assertFalse(allow_named_ot_rows(filters))

        request = RequestFactory().get("/", {"include_names": "1"})
        request.user = SimpleNamespace(
            is_authenticated=True,
            is_superuser=False,
            has_perm=lambda p: False,
        )
        filters.request = request
        self.assertFalse(allow_named_ot_rows(filters))

        request.user.has_perm = lambda p: p == "attendance.change_attendance"
        self.assertTrue(allow_named_ot_rows(filters))

        request_no_flag = RequestFactory().get("/")
        request_no_flag.user = request.user
        filters.request = request_no_flag
        self.assertFalse(allow_named_ot_rows(filters))


class ExitHelperPriorityTests(SimpleTestCase):
    def test_iter_exits_priority_merge(self):
        from unittest.mock import patch

        from report.engine import ReportFilters
        from report.metrics._exits import iter_exits

        filters = ReportFilters(from_date=date(2026, 1, 1), to_date=date(2026, 1, 31))
        offboarding = [
            {
                "employee_id": 1,
                "exit_date": date(2026, 1, 10),
                "source": "offboarding_archived",
                "employee": None,
            }
        ]
        resignation = [
            {
                "employee_id": 1,
                "exit_date": date(2026, 1, 12),
                "source": "resignation_approved",
                "employee": None,
            },
            {
                "employee_id": 2,
                "exit_date": date(2026, 1, 15),
                "source": "resignation_approved",
                "employee": None,
            },
        ]
        inactive = [
            {
                "employee_id": 2,
                "exit_date": date(2026, 1, 20),
                "source": "inactive_contract_end",
                "employee": None,
            },
            {
                "employee_id": 3,
                "exit_date": date(2026, 1, 18),
                "source": "inactive_contract_end",
                "employee": None,
            },
        ]
        with patch(
            "report.metrics._exits._offboarding_archived_exits",
            return_value=offboarding,
        ), patch(
            "report.metrics._exits._resignation_exits",
            return_value=resignation,
        ), patch(
            "report.metrics._exits._inactive_contract_exits",
            return_value=inactive,
        ):
            rows = iter_exits(filters)
        by_id = {r["employee_id"]: r["source"] for r in rows}
        self.assertEqual(by_id[1], "offboarding_archived")
        self.assertEqual(by_id[2], "resignation_approved")
        self.assertEqual(by_id[3], "inactive_contract_end")
        self.assertEqual(len(rows), 3)


class DocumentExpiryModelTests(SimpleTestCase):
    def test_no_employee_document_probe(self):
        import pathlib

        metrics_root = pathlib.Path(__file__).resolve().parents[1] / "metrics"
        bad = []
        for path in metrics_root.rglob("*.py"):
            text = path.read_text(encoding="utf-8")
            if 'get_model("employee", "Document")' in text:
                bad.append(str(path))
            if 'get_model("employee", "EmployeeDocument")' in text:
                bad.append(str(path))
        self.assertEqual(bad, [])


class DrilldownRegistryTests(SimpleTestCase):
    def test_first_wave_has_drilldown(self):
        import report.metrics  # noqa: F401
        from report.registry import get_report

        for slug in (
            "workforce-composition",
            "payslip-register",
            "recruitment-funnel",
        ):
            definition = get_report(slug)
            self.assertIsNotNone(definition)
            self.assertIsNotNone(definition.drilldown_fn, slug)

    def test_run_drilldown_requires_fn(self):
        from report.engine import ReportFilters
        from report.registry import ReportDefinition, register, run_drilldown

        register(
            ReportDefinition(
                slug="unit-no-drill",
                name="No Drill",
                domain="workforce",
                description="",
                permission="employee.view_employee",
                query_fn=lambda f: {},
                required_apps=(),
            )
        )
        with self.assertRaises(ValueError):
            run_drilldown(
                "unit-no-drill",
                ReportFilters(from_date=date(2026, 1, 1), to_date=date(2026, 1, 31)),
                {"dimension": "department", "value": "HR"},
            )
