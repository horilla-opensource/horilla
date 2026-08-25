"""
Server-side Excel / CSV export for standard report payloads.

Workbook layout (aligned with Horilla explorer exports):
  1. Cover   — company letterhead, report title, metadata, filters, KPIs
  2. Data    — letterhead + primary tabular result set, print-ready
  3. Charts  — native openpyxl chart objects + their source data (when present)
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any, Optional

from django.http import HttpResponse
from django.utils import timezone

# ---------------------------------------------------------------------------
# Style tokens — Horilla primary (coral / #E54F38)
# ---------------------------------------------------------------------------
COLOR_PRIMARY = "E54F38"
COLOR_PRIMARY_DARK = "CE4732"
COLOR_PRIMARY_SOFT = "FCEDEB"
COLOR_SLATE = "595959"
COLOR_TEXT = "333333"
COLOR_ROW_ALT = "FEF6F5"
COLOR_WHITE = "FFFFFF"
COLOR_BORDER = "F7C8C1"
COLOR_KPI_VALUE = "AC3B2A"
COLOR_FILTER_BG = "FFF4CE"
COLOR_FILTER_BORDER = "E8B84B"


def _table_from_payload(
    payload: dict[str, Any]
) -> tuple[list[str], list[str], list[list[Any]]]:
    table = payload.get("table") or {}
    columns = table.get("columns") or []
    rows = table.get("rows") or []
    headers = [c.get("label") or c.get("key") for c in columns]
    keys = [c.get("key") for c in columns]
    data_rows = [[row.get(k, "") for k in keys] for row in rows]
    return headers, keys, data_rows


def _kpi_rows(payload: dict[str, Any]) -> list[list[Any]]:
    kpis = payload.get("kpis") or []
    return [
        [kpi.get("label", ""), kpi.get("value", ""), kpi.get("hint", "")]
        for kpi in kpis
    ]


def _period_label(period: dict) -> str:
    start = period.get("from_date") or ""
    end = period.get("to_date") or ""
    if start and end:
        return f"{start}  →  {end}"
    return start or end or "—"


def _company_from_meta(meta: Optional[dict]) -> dict[str, Any]:
    meta = meta or {}
    company = meta.get("company") or {}
    if isinstance(company, str):
        return {
            "name": company or "All companies",
            "address": "",
            "country": "",
            "state": "",
            "city": "",
            "zip": "",
            "logo_path": None,
            "is_all": company in ("", "All companies", "all"),
            "location_line": "",
            "address_lines": [company or "All companies"],
        }
    return {
        "name": company.get("name") or "All companies",
        "address": company.get("address") or "",
        "country": company.get("country") or "",
        "state": company.get("state") or "",
        "city": company.get("city") or "",
        "zip": company.get("zip") or "",
        "logo_path": company.get("logo_path"),
        "is_all": bool(company.get("is_all", False)),
        "location_line": company.get("location_line") or "",
        "address_lines": company.get("address_lines")
        or [company.get("name") or "All companies"],
    }


def _coerce_cell(value: Any) -> Any:
    """Normalize cell values for Excel (numbers stay numeric when possible)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (date, datetime)):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return ""
        if re.fullmatch(r"-?\d+(\.\d+)?%", text):
            try:
                return float(text[:-1]) / 100.0
            except ValueError:
                return text
        if re.fullmatch(r"-?\d+", text):
            try:
                return int(text)
            except ValueError:
                return text
        if re.fullmatch(r"-?\d+\.\d+", text):
            try:
                return float(text)
            except ValueError:
                return text
        return text
    return str(value)


def _looks_percent_header(header: str) -> bool:
    h = (header or "").lower()
    return "%" in h or "rate" in h or "pct" in h or "percent" in h


def _looks_money_header(header: str) -> bool:
    h = (header or "").lower()
    return any(
        token in h
        for token in (
            "gross",
            "net",
            "pay",
            "cost",
            "amount",
            "deduction",
            "allowance",
            "salary",
            "wage",
        )
    )


def _currency_format() -> tuple[str, str]:
    """(symbol, position) from PayrollSettings — same convention as
    base/templatetags/horillafilters.py::currency_symbol_position, with the
    same "$"/prefix fallback used app-wide when payroll isn't installed or
    no settings row exists yet."""
    try:
        from django.apps import apps as django_apps

        if django_apps.is_installed("payroll"):
            from payroll.models.tax_models import PayrollSettings

            settings_row = PayrollSettings.objects.first()
            if settings_row:
                return (
                    settings_row.currency_symbol or "$",
                    settings_row.position or "prefix",
                )
    except Exception:
        pass
    return "$", "prefix"


def _money_number_format(symbol: str, position: str) -> str:
    """Excel number_format string embedding the resolved currency symbol."""
    escaped = symbol.replace('"', '\\"')
    if position == "postfix":
        return f'#,##0.00"{escaped}"'
    return f'"{escaped}"#,##0.00'


def _format_money_text(value: float, symbol: str, position: str) -> str:
    """Plain-text money rendering for the PDF path (no Excel number_format)."""
    number = f"{value:,.2f}"
    return f"{number}{symbol}" if position == "postfix" else f"{symbol}{number}"


def _format_kpi_value(
    value: Any, label: str, symbol: str, position: str
) -> Optional[str]:
    """Plain-text KPI value/prior_value rendering for the PDF path — KPI
    payloads carry raw numbers (see report/metrics/*.py), so without this
    a "Total gross pay" KPI would print as a bare "128450.75"."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        if _looks_percent_header(label):
            pct = value * 100 if -1 <= value <= 1 else value
            return f"{pct:,.1f}%"
        if _looks_money_header(label):
            return _format_money_text(float(value), symbol, position)
        if isinstance(value, int):
            return f"{value:,}"
        return f"{value:,.2f}"
    return str(value)


def _format_table_cell(raw: Any, header: str, symbol: str, position: str) -> str:
    """Plain-text, presentation-ready rendering of a detail-table cell for
    the PDF export — money/percent/thousands formatting instead of a bare
    str(), matching the number_format treatment already applied on the
    Excel side."""
    # A source string that is already a percent ("101%", "89.5%") must round-trip
    # verbatim: _coerce_cell divides it by 100, and the re-multiply guard below
    # can't tell 1.01 (=101%) from a literal fraction — >100% values would
    # otherwise print as "1.0%".
    if isinstance(raw, str) and re.fullmatch(r"-?\d+(\.\d+)?%", raw.strip()):
        return raw.strip()
    value = _coerce_cell(raw)
    if value == "" or value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and _looks_percent_header(header):
        pct = value * 100 if -1 <= value <= 1 else value
        return f"{pct:,.1f}%"
    if isinstance(value, (int, float)) and _looks_money_header(header):
        return _format_money_text(float(value), symbol, position)
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        return f"{value:,.2f}"
    return str(value)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def export_csv(
    payload: dict[str, Any], filename: str = "report.csv", meta: Optional[dict] = None
) -> HttpResponse:
    """Export with company letterhead + KPIs + table (explorer-compatible)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    meta = meta or {}
    company = _company_from_meta(meta)
    title = payload.get("title") or "Report"
    period = payload.get("period") or {}

    if not company.get("is_all"):
        for line in company.get("address_lines") or [company.get("name")]:
            writer.writerow([line])
        writer.writerow([])
    else:
        writer.writerow([company.get("name") or "All companies"])
        writer.writerow([])

    writer.writerow([title])
    writer.writerow(["Period", _period_label(period)])
    writer.writerow(
        [
            "Generated",
            timezone.now().strftime("%Y-%m-%d %H:%M:%S %Z")
            or timezone.now().isoformat(),
        ]
    )
    if meta.get("user"):
        writer.writerow(["Prepared for", meta.get("user")])
    if payload.get("domain"):
        writer.writerow(
            ["Domain", str(payload.get("domain")).replace("_", " ").title()]
        )
    filters_label = meta.get("filters_label")
    if filters_label:
        writer.writerow(["Applied Filters", filters_label])
    writer.writerow([])

    writer.writerow(["=== Key Metrics ==="])
    writer.writerow(["Metric", "Value", "Notes"])
    writer.writerows(_kpi_rows(payload))
    writer.writerow([])

    headers, _keys, rows = _table_from_payload(payload)
    if headers:
        writer.writerow(["=== Detail Data ==="])
        writer.writerow(headers)
        writer.writerows(rows)

    response = HttpResponse(buffer.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color=COLOR_BORDER)
    filter_side = Side(style="thin", color=COLOR_FILTER_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "company_name_font": Font(name="Calibri", size=14, bold=True, color=COLOR_TEXT),
        "company_meta_font": Font(name="Calibri", size=10, color=COLOR_TEXT),
        "product_font": Font(name="Calibri", size=9, color=COLOR_SLATE),
        "title_font": Font(name="Calibri", size=16, bold=True, color=COLOR_PRIMARY),
        "subtitle_font": Font(name="Calibri", size=11, color=COLOR_SLATE),
        "section_font": Font(name="Calibri", size=11, bold=True, color=COLOR_PRIMARY),
        "label_font": Font(name="Calibri", size=10, bold=True, color=COLOR_SLATE),
        "meta_font": Font(name="Calibri", size=10, color=COLOR_TEXT),
        "header_font": Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE),
        "kpi_label_font": Font(name="Calibri", size=10, color=COLOR_SLATE),
        "kpi_value_font": Font(
            name="Calibri", size=13, bold=True, color=COLOR_KPI_VALUE
        ),
        "cell_font": Font(name="Calibri", size=10, color=COLOR_TEXT),
        "footer_font": Font(name="Calibri", size=8, italic=True, color=COLOR_SLATE),
        "filter_font": Font(name="Calibri", size=10, color=COLOR_TEXT),
        "filter_bold": Font(name="Calibri", size=10, bold=True, color=COLOR_TEXT),
        "section_fill": PatternFill("solid", fgColor=COLOR_PRIMARY_SOFT),
        "header_fill": PatternFill("solid", fgColor=COLOR_PRIMARY),
        "kpi_fill": PatternFill("solid", fgColor="FEF6F5"),
        "alt_fill": PatternFill("solid", fgColor=COLOR_ROW_ALT),
        "filter_fill": PatternFill("solid", fgColor=COLOR_FILTER_BG),
        "border": border,
        "filter_border": Border(
            left=filter_side, right=filter_side, top=filter_side, bottom=filter_side
        ),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "top_left": Alignment(horizontal="left", vertical="top", wrap_text=True),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center"),
    }


def _set_col_widths(ws, widths: dict[str, float]):
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def _autofit(ws, min_width: float = 10, max_width: float = 42, start_row: int = 1):
    from openpyxl.utils import get_column_letter

    for col_cells in ws.iter_cols(min_row=start_row, max_row=ws.max_row):
        lengths = []
        for cell in col_cells:
            if cell.value is None:
                continue
            lengths.append(len(str(cell.value)))
        if not lengths:
            continue
        width = min(max(max(lengths) + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _section_banner(ws, row: int, title: str, styles, cols: int = 6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = styles["section_font"]
    cell.alignment = styles["left"]
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).fill = styles["section_fill"]
    return row + 1


def _write_letterhead(ws, company: dict, styles, meta: dict, col_span: int = 6) -> int:
    """
    Write company letterhead at the top of a sheet.
    Returns the next free row index.
    """
    from openpyxl.drawing.image import Image as XLImage

    row = 1
    logo_path = company.get("logo_path")
    has_logo = False
    text_col = 1

    # Product label on the far right of row 1 (never overlapped by company merge)
    product = ws.cell(
        row=1,
        column=col_span,
        value=meta.get("product_name") or "Horilla HR · Standard Reports",
    )
    product.font = styles["product_font"]
    product.alignment = styles["right"]

    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 52
            img.height = 52
            ws.add_image(img, "A1")
            has_logo = True
            text_col = 2
            ws.column_dimensions["A"].width = 10
        except Exception:
            has_logo = False
            text_col = 1

    # Leave column col_span free for the product label
    end_col = max(text_col, col_span - 1)

    name_cell = ws.cell(
        row=row, column=text_col, value=company.get("name") or "All companies"
    )
    name_cell.font = styles["company_name_font"]
    name_cell.alignment = styles["top_left"]
    if end_col > text_col:
        ws.merge_cells(
            start_row=row, start_column=text_col, end_row=row, end_column=end_col
        )
    ws.row_dimensions[row].height = 18
    row += 1

    if not company.get("is_all"):
        detail_lines = []
        if company.get("address"):
            detail_lines.append(company["address"])
        if company.get("location_line"):
            detail_lines.append(company["location_line"])
        elif any(company.get(k) for k in ("city", "state", "country")):
            detail_lines.append(
                ", ".join(
                    p
                    for p in (
                        company.get("city"),
                        company.get("state"),
                        company.get("country"),
                    )
                    if p
                )
            )
        if company.get("zip"):
            detail_lines.append(f"ZIP: {company['zip']}")

        for line in detail_lines:
            cell = ws.cell(row=row, column=text_col, value=line)
            cell.font = styles["company_meta_font"]
            cell.alignment = styles["left"]
            if end_col > text_col:
                ws.merge_cells(
                    start_row=row,
                    start_column=text_col,
                    end_row=row,
                    end_column=end_col,
                )
            ws.row_dimensions[row].height = 14
            row += 1

        # Give the floating logo a little vertical room without large empty gaps
        if has_logo and row < 4:
            row = 4
    else:
        row += 1

    return row + 1  # blank spacer before title


def _write_cover(wb, payload: dict[str, Any], meta: Optional[dict] = None):
    from openpyxl.styles import Font as XLFont

    styles = _styles()
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True

    meta = meta or {}
    company = _company_from_meta(meta)
    title = str(payload.get("title") or "Standard Report")
    period = payload.get("period") or {}
    generated = meta.get("generated_at") or timezone.now()
    if hasattr(generated, "strftime"):
        generated_str = generated.strftime("%d %b %Y, %H:%M")
    else:
        generated_str = str(generated)

    row = _write_letterhead(ws, company, styles, meta, col_span=6)

    # Report title block
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value=title)
    title_cell.font = styles["title_font"]
    title_cell.alignment = styles["left"]
    ws.row_dimensions[row].height = 26
    row += 1

    domain = payload.get("domain") or meta.get("domain") or ""
    domain_label = str(domain).replace("_", " ").title() if domain else ""
    subtitle_bits = [f"Reporting period  {_period_label(period)}"]
    if domain_label:
        subtitle_bits.append(f"Domain  {domain_label}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sub = ws.cell(row=row, column=1, value="    ·    ".join(subtitle_bits))
    sub.font = styles["subtitle_font"]
    row += 2

    # ---- Report details (two-column grouped) ----
    row = _section_banner(ws, row, "Report details", styles, cols=6)

    left_details = [
        ("Generated on", generated_str),
        ("Generated by", meta.get("user") or "—"),
        ("Report ID", payload.get("slug") or meta.get("slug") or "—"),
    ]
    right_details = [
        ("Company", company.get("name") or "All companies"),
        ("Period", _period_label(period)),
        ("Domain", domain_label or "—"),
    ]

    start_details = row
    for idx, (label, value) in enumerate(left_details):
        r = start_details + idx
        ws.cell(row=r, column=1, value=label).font = styles["label_font"]
        ws.cell(row=r, column=2, value=value).font = styles["meta_font"]
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

    for idx, (label, value) in enumerate(right_details):
        r = start_details + idx
        ws.cell(row=r, column=4, value=label).font = styles["label_font"]
        ws.cell(row=r, column=5, value=value).font = styles["meta_font"]
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)

    row = start_details + max(len(left_details), len(right_details)) + 1

    # ---- Applied filters ----
    # Structured Filter | Value rows (audit-ready) when the caller provides
    # summary_pairs; single joined-label row as the legacy fallback.
    filters_pairs = meta.get("filters_pairs") or []
    filters_label = meta.get("filters_label") or ""
    row = _section_banner(ws, row, "Applied filters", styles, cols=6)
    if filters_pairs:
        for c_idx, heading in ((1, "Filter"), (3, "Selection")):
            cell = ws.cell(row=row, column=c_idx, value=heading)
            cell.font = styles["header_font"]
            cell.alignment = styles["left"]
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.fill = styles["header_fill"]
            c.border = styles["border"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1
        for label, value in filters_pairs:
            lcell = ws.cell(row=row, column=1, value=str(label))
            lcell.font = styles["filter_bold"]
            lcell.alignment = styles["left"]
            vcell = ws.cell(row=row, column=3, value=str(value))
            vcell.font = styles["filter_font"]
            vcell.alignment = styles["left"]
            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                c.fill = styles["filter_fill"]
                c.border = styles["filter_border"]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 16
            row += 1
        row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        filter_cell = ws.cell(
            row=row,
            column=1,
            value=filters_label or "Default period / no additional filters",
        )
        filter_cell.font = styles["filter_font"]
        filter_cell.alignment = styles["left"]
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.fill = styles["filter_fill"]
            c.border = styles["filter_border"]
        ws.row_dimensions[row].height = 22
        row += 2

    # ---- Narrative (compare / KPI summary) ----
    from report.narrative import build_narrative

    # Build the narrative from presentation-formatted KPI values so the prose
    # reads "Attendance rate: 93.4%", not the payload's raw "0.934".
    _sym, _pos = _currency_format()
    narrative = build_narrative(
        {
            **payload,
            "kpis": [
                {
                    **kpi,
                    "value": _format_kpi_value(
                        kpi.get("value"), str(kpi.get("label", "")), _sym, _pos
                    ),
                }
                for kpi in (payload.get("kpis") or [])
            ],
        }
    )
    if narrative:
        row = _section_banner(ws, row, "Summary", styles, cols=6)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ncell = ws.cell(row=row, column=1, value=narrative)
        ncell.font = styles["meta_font"]
        ncell.alignment = styles["left"]
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = styles["border"]
        ws.row_dimensions[row].height = 36
        row += 2

    # ---- KPIs ----
    kpis = payload.get("kpis") or []
    has_compare = any(
        k.get("prior_value") is not None or k.get("delta_label") for k in kpis
    )
    currency_symbol, currency_position = _currency_format()

    row = _section_banner(ws, row, "Key performance indicators", styles, cols=6)
    kpi_headers = ["#", "Metric", "Value"]
    if has_compare:
        kpi_headers += ["Prior", "Change"]
    kpi_headers.append("Notes")
    for col, text in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["center"]
    kpi_header_row = row
    row += 1

    notes_col = len(kpi_headers)
    if not kpis:
        cell = ws.cell(row=row, column=2, value="No KPI metrics for this report.")
        cell.font = styles["meta_font"]
        row += 1
    else:
        for idx, kpi in enumerate(kpis, start=1):
            label = str(kpi.get("label", "")).lower()
            values = [idx, kpi.get("label", ""), _coerce_cell(kpi.get("value", ""))]
            if has_compare:
                values.append(_coerce_cell(kpi.get("prior_value", "—")))
                values.append(kpi.get("delta_label") or "—")
            values.append(kpi.get("hint", "") or "")

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = styles["border"]
                cell.font = (
                    styles["kpi_value_font"] if col == 3 else styles["kpi_label_font"]
                )
                cell.fill = styles["kpi_fill"]
                cell.alignment = (
                    styles["center"] if col in (1, 3, notes_col - 1) else styles["left"]
                )
                is_value_col = col in (3, 4) if has_compare else col == 3
                if (
                    is_value_col
                    and isinstance(val, (int, float))
                    and not isinstance(val, bool)
                ):
                    if 0 <= val <= 1 and (
                        "%" in str(kpi.get("value", "")) or _looks_percent_header(label)
                    ):
                        cell.number_format = "0.0%"
                    elif _looks_money_header(label):
                        cell.number_format = _money_number_format(
                            currency_symbol, currency_position
                        )
                if has_compare and col == notes_col - 1:
                    direction = kpi.get("delta_direction")
                    if direction == "up":
                        cell.font = XLFont(
                            name="Calibri", size=10, bold=True, color="1B7A45"
                        )
                    elif direction == "down":
                        cell.font = XLFont(
                            name="Calibri", size=10, bold=True, color="B42318"
                        )
            row += 1

    message = payload.get("message")
    if message:
        row += 1
        row = _section_banner(ws, row, "Notes", styles, cols=6)
        note = ws.cell(row=row, column=1, value=str(message))
        note.font = styles["meta_font"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    row += 2
    footer = ws.cell(
        row=row,
        column=1,
        value=(
            "Confidential — for internal use only. "
            "Figures are aggregated server-side from Horilla HR source data."
        ),
    )
    footer.font = styles["footer_font"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    _set_col_widths(
        ws,
        {"A": 16, "B": 28, "C": 16, "D": 14, "E": 22, "F": 22},
    )
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    return ws, kpi_header_row


def _write_data_sheet(wb, payload: dict[str, Any], meta: Optional[dict] = None):
    from openpyxl.utils import get_column_letter

    styles = _styles()
    headers, _keys, rows = _table_from_payload(payload)
    ws = wb.create_sheet("Data")
    ws.sheet_view.showGridLines = False
    currency_symbol, currency_position = _currency_format()
    money_format = _money_number_format(currency_symbol, currency_position)

    meta = meta or {}
    company = _company_from_meta(meta)
    title = str(payload.get("title") or "Report")
    period = payload.get("period") or {}
    col_count = max(len(headers), 6)

    row = _write_letterhead(ws, company, styles, meta, col_span=col_count)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    banner = ws.cell(row=row, column=1, value=f"{title}  —  Detail")
    banner.font = styles["title_font"]
    banner.alignment = styles["left"]
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    meta_cell = ws.cell(
        row=row,
        column=1,
        value=f"Period: {_period_label(period)}    ·    Rows: {len(rows)}",
    )
    meta_cell.font = styles["subtitle_font"]
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).fill = styles["section_fill"]
    row += 1

    filters_label = meta.get("filters_label")
    if filters_label:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
        fcell = ws.cell(row=row, column=1, value=f"Applied filters: {filters_label}")
        fcell.font = styles["filter_font"]
        for col in range(1, col_count + 1):
            c = ws.cell(row=row, column=col)
            c.fill = styles["filter_fill"]
            c.border = styles["filter_border"]
        row += 1

    row += 1  # spacer before table

    if not headers:
        ws.cell(row=row, column=1, value="No tabular detail for this report.").font = (
            styles["meta_font"]
        )
        _set_col_widths(ws, {"A": 48})
        return ws

    header_row = row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = styles["center"]
    ws.row_dimensions[header_row].height = 22

    for r_idx, data_row in enumerate(rows):
        excel_row = header_row + 1 + r_idx
        alt = r_idx % 2 == 1
        for c_idx, raw in enumerate(data_row):
            header = headers[c_idx] if c_idx < len(headers) else ""
            value = _coerce_cell(raw)
            cell = ws.cell(row=excel_row, column=c_idx + 1, value=value)
            cell.font = styles["cell_font"]
            cell.border = styles["border"]
            cell.alignment = styles["left"]
            if alt:
                cell.fill = styles["alt_fill"]

            if isinstance(value, float) and _looks_percent_header(header):
                if value > 1:
                    cell.value = value / 100.0
                cell.number_format = "0.0%"
                cell.alignment = styles["right"]
            elif isinstance(value, (int, float)) and _looks_money_header(header):
                cell.number_format = money_format
                cell.alignment = styles["right"]
            elif isinstance(value, (int, float)):
                cell.number_format = "#,##0.##"
                cell.alignment = styles["right"]

    if rows:
        from openpyxl.styles import Font as XLFont

        total_row = header_row + 1 + len(rows)
        total_font = XLFont(name="Calibri", size=10, bold=True, color=COLOR_PRIMARY)
        ws.cell(row=total_row, column=1, value="Total / Count").font = total_font
        ws.cell(row=total_row, column=1).fill = styles["section_fill"]
        ws.cell(row=total_row, column=1).border = styles["border"]

        for c_idx, header in enumerate(headers):
            col = c_idx + 1
            col_letter = get_column_letter(col)
            data_start = header_row + 1
            data_end = header_row + len(rows)
            cell = ws.cell(row=total_row, column=col)
            cell.fill = styles["section_fill"]
            cell.border = styles["border"]
            cell.font = total_font
            if c_idx == 0:
                continue
            numeric_count = 0
            for data_row in rows:
                coerced = _coerce_cell(data_row[c_idx] if c_idx < len(data_row) else "")
                if isinstance(coerced, (int, float)):
                    numeric_count += 1
            if numeric_count >= max(1, len(rows) // 2):
                cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
                if _looks_money_header(header):
                    cell.number_format = money_format
                elif _looks_percent_header(header):
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = "#,##0.##"
                cell.alignment = styles["right"]

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(rows)}"
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    ws.page_setup.orientation = "landscape" if len(headers) > 5 else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"

    _autofit(ws, start_row=header_row)
    return ws


# Same brand hues as report/chart_render.py's PALETTE (plain hex — openpyxl
# graphicalProperties.solidFill wants strings, not reportlab Color objects).
_CHART_PALETTE_HEX = ["E54F38", "2563EB", "15803D", "7C3AED", "0D9488", "F59E0B"]


def _add_native_chart(
    ws, chart_payload, header_row, data_start, data_end, n_series, anchor_col
):
    """Attach a real openpyxl chart object (BarChart/LineChart/PieChart) —
    not just a data table — bound to the range already written on this
    sheet, so it stays live if the workbook is edited later."""
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    chart_type = (chart_payload.get("type") or "bar").lower()
    if chart_type in ("donut", "pie"):
        obj = PieChart()
    elif chart_type == "line":
        obj = LineChart()
    else:
        obj = BarChart()
        obj.type = "col"
        obj.grouping = "clustered"
        obj.gapWidth = 40

    obj.title = chart_payload.get("title") or chart_payload.get("id") or "Chart"
    obj.height = 8
    obj.width = 16

    data_ref = Reference(
        ws, min_col=2, max_col=1 + n_series, min_row=header_row, max_row=data_end
    )
    cats_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    obj.add_data(data_ref, titles_from_data=True)
    obj.set_categories(cats_ref)

    for i, series in enumerate(obj.series):
        color = _CHART_PALETTE_HEX[i % len(_CHART_PALETTE_HEX)]
        series.graphicalProperties.solidFill = color
        if chart_type == "line":
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 20000
            series.smooth = False

    obj.legend.position = "b"
    ws.add_chart(obj, f"{anchor_col}{header_row}")


def _write_chart_sheet(wb, payload: dict[str, Any], meta: Optional[dict] = None):
    from openpyxl.utils import get_column_letter

    styles = _styles()
    charts = payload.get("charts") or []
    if not charts:
        return None

    ws = wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COLOR_PRIMARY
    meta = meta or {}
    company = _company_from_meta(meta)

    row = _write_letterhead(ws, company, styles, meta, col_span=5)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    banner = ws.cell(row=row, column=1, value="Charts")
    banner.font = styles["title_font"]
    banner.alignment = styles["left"]
    ws.row_dimensions[row].height = 24
    row += 2

    for chart in charts:
        title = chart.get("title") or chart.get("id") or "Chart"
        row = _section_banner(ws, row, title, styles, cols=5)

        categories = chart.get("categories") or []
        series_list = [s for s in (chart.get("series") or []) if s.get("data")]
        if not categories or not series_list:
            continue

        headers = ["Category"] + [s.get("name") or "Series" for s in series_list]
        header_row = row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.border = styles["border"]
            cell.alignment = styles["center"]
        row += 1
        data_start = row

        for idx, cat in enumerate(categories):
            values = [cat]
            for series in series_list:
                data = series.get("data") or []
                values.append(_coerce_cell(data[idx] if idx < len(data) else ""))
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = styles["cell_font"]
                cell.border = styles["border"]
                if idx % 2 == 1:
                    cell.fill = styles["alt_fill"]
                if col > 1 and isinstance(val, (int, float)):
                    cell.number_format = "#,##0.##"
                    cell.alignment = styles["right"]
            row += 1
        data_end = row - 1

        anchor_col = get_column_letter(len(headers) + 2)
        _add_native_chart(
            ws, chart, header_row, data_start, data_end, len(series_list), anchor_col
        )
        row = max(row, header_row + 17) + 2  # clear the chart's own height

    _set_col_widths(ws, {"A": 28, "B": 16, "C": 16, "D": 16, "E": 16})
    return ws


def export_xlsx(
    payload: dict[str, Any],
    filename: str = "report.xlsx",
    meta: Optional[dict] = None,
) -> HttpResponse:
    """
    Export a structured enterprise workbook:
      Cover → Data → Chart Data (optional)
    """
    from openpyxl import Workbook

    wb = Workbook()
    _write_cover(wb, payload, meta=meta)
    _write_data_sheet(wb, payload, meta=meta)
    _write_chart_sheet(wb, payload, meta=meta)

    props = wb.properties
    props.title = str(payload.get("title") or "Horilla Report")
    company = _company_from_meta(meta)
    props.creator = company.get("name") or "Horilla HR"
    props.description = (
        f"Standard report export · {(payload.get('slug') or '')} · "
        f"{_period_label(payload.get('period') or {})}"
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_pdf(
    payload: dict[str, Any],
    filename: str = "report.pdf",
    meta: Optional[dict] = None,
) -> HttpResponse:
    """
    Server-side PDF via xhtml2pdf/pisa + letterhead template — an
    enterprise-ready deliverable: real chart images, page numbers/running
    footer, currency/percent-formatted figures, and a visible notice when
    the row cap truncates the detail table.
    """
    import os
    import tempfile
    from io import BytesIO

    from django.template.loader import render_to_string
    from django.utils.translation import gettext as _
    from xhtml2pdf import pisa

    from report.chart_render import render_chart_png
    from report.narrative import build_narrative

    meta = meta or {}
    company = _company_from_meta(meta)
    headers, _keys, data_rows = _table_from_payload(payload)
    max_rows = 500
    total_rows = len(data_rows)
    truncated = total_rows > max_rows
    data_rows = data_rows[:max_rows]

    currency_symbol, currency_position = _currency_format()
    formatted_rows = [
        [
            _format_table_cell(
                cell,
                headers[i] if i < len(headers) else "",
                currency_symbol,
                currency_position,
            )
            for i, cell in enumerate(row)
        ]
        for row in data_rows
    ]

    # Per-column alignment: right-align columns whose values are mostly
    # numeric (money/percent/count), matching the Excel side — a key part of
    # a clean tabular read in print.
    col_classes: list[str] = []
    for c_idx, header in enumerate(headers):
        numeric = 0
        non_empty = 0
        for row in data_rows:
            raw = row[c_idx] if c_idx < len(row) else ""
            coerced = _coerce_cell(raw)
            if coerced == "" or coerced is None:
                continue
            non_empty += 1
            if isinstance(coerced, (int, float)) and not isinstance(coerced, bool):
                numeric += 1
        is_num = non_empty > 0 and numeric >= max(1, non_empty // 2)
        col_classes.append("num" if is_num else "txt")
    # Zip class onto every cell so the template needs no positional lookups.
    classed_rows = [list(zip(row, col_classes)) for row in formatted_rows]
    classed_headers = list(zip(headers, col_classes))

    period = payload.get("period") or {}
    compare = payload.get("compare") or {}
    period_label = _period_label(period)
    compare_label = ""
    if compare.get("period"):
        compare_label = (
            f"{compare.get('label') or _('Compare')}: "
            f"{_period_label(compare['period'])}"
        )

    logo_path = company.get("logo_path")
    if logo_path and not os.path.exists(logo_path):
        logo_path = None

    address_lines = company.get("address_lines") or [
        company.get("name") or "All companies"
    ]
    kpis = payload.get("kpis") or []
    has_compare = any(
        k.get("prior_value") is not None or k.get("delta_label") for k in kpis
    )
    formatted_kpis = [
        {
            **kpi,
            "value": _format_kpi_value(
                kpi.get("value"),
                str(kpi.get("label", "")),
                currency_symbol,
                currency_position,
            ),
            "prior_value": _format_kpi_value(
                kpi.get("prior_value"),
                str(kpi.get("label", "")),
                currency_symbol,
                currency_position,
            ),
        }
        for kpi in kpis
    ]
    # Narrative built from the FORMATTED KPIs so the summary sentence reads
    # "Attendance rate: 93.4%", not the raw "0.934" stored in the payload.
    narrative = build_narrative({**payload, "kpis": formatted_kpis})

    generated_at = meta.get("generated_at")
    if generated_at and hasattr(generated_at, "strftime"):
        generated_str = timezone.localtime(generated_at).strftime("%Y-%m-%d %H:%M")
    else:
        generated_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    # Structured (label, value) filter rows preferred; chip list is the
    # legacy fallback for callers that never built summary_pairs.
    filter_pairs = meta.get("filters_pairs") or []
    filter_chips = []
    if not filter_pairs:
        filter_chips = payload.get("filters") or []
        if not filter_chips and meta.get("filters_label"):
            filter_chips = [meta["filters_label"]]

    # Render each chart to a PNG and reference it by file path — xhtml2pdf
    # resolves <img src> the same way it already does for the company logo,
    # so no new image-embedding mechanism is needed.
    chart_images = []
    tmp_paths = []
    for chart in payload.get("charts") or []:
        png_bytes = render_chart_png(chart)
        if not png_bytes:
            continue
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        try:
            tmp.write(png_bytes)
        finally:
            tmp.close()
        tmp_paths.append(tmp.name)
        chart_images.append(
            {
                "title": chart.get("title") or chart.get("id") or _("Chart"),
                "path": tmp.name,
            }
        )

    try:
        html = render_to_string(
            "report/standard_report_pdf.html",
            {
                "report_title": payload.get("title") or _("Standard Report"),
                "domain_label": (payload.get("domain") or meta.get("domain") or "")
                .replace("_", " ")
                .title(),
                "company_name": company.get("name") or _("All companies"),
                "address_lines": address_lines,
                "logo_path": logo_path,
                "period_label": period_label,
                "compare_label": compare_label,
                "narrative": narrative,
                "filter_pairs": filter_pairs,
                "filter_chips": filter_chips,
                "kpis": formatted_kpis,
                "has_compare": has_compare,
                "chart_images": chart_images,
                "headers": classed_headers,
                "rows": classed_rows,
                "col_count": len(headers),
                "truncated": truncated,
                "total_rows": total_rows,
                "shown_rows": len(data_rows),
                "landscape": len(headers) > 5,
                "dense": len(headers) > 8,
                "report_ref": payload.get("slug") or meta.get("slug") or "",
                "product_name": meta.get("product_name")
                or "Horilla HR · Standard Reports",
                "generated_at": generated_str,
                "generated_by": meta.get("user") or "",
            },
        )

        buf = BytesIO()
        result = pisa.CreatePDF(src=html, dest=buf)
        if result.err:
            raise RuntimeError(f"PDF generation failed ({result.err})")
    finally:
        for path in tmp_paths:
            try:
                os.remove(path)
            except OSError:
                pass

    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
